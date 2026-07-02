from __future__ import annotations

import io
import json
import re
from datetime import date, datetime, timedelta
from typing import Callable

from sqlalchemy.orm import Session

from app.models import models

from .schemas import DailyReportField, DailyReportIn, RevenueGridIn, RevenueRecordEntry


PAYMENT_FIELDS = (
    "total_medical_fee",
    "nhis_burden_total",
    "cash_amount",
    "card_amount",
    "receivable_income",
    "unpaid_amount",
    "health_living_fee",
    "certificate_amount",
    "disability_fund",
    "uninsured_amount",
    "meal_amount",
    "other_amount",
    "discount_amount",
    "free_amount",
    "cash_expense_amount",
    "transfer_amount",
)
PAYMENT_LABELS = {
    "total_medical_fee": "총진료비",
    "nhis_burden_total": "공단부담총액",
    "cash_amount": "현금수납액",
    "card_amount": "카드수납액",
    "receivable_income": "미수입금",
    "unpaid_amount": "미수발생",
    "health_living_fee": "건강생활유지비",
    "certificate_amount": "입,통원확인서",
    "disability_fund": "장애인기금",
    "uninsured_amount": "비급여",
    "meal_amount": "식대",
    "other_amount": "기타",
    "discount_amount": "할인",
    "free_amount": "FREE",
    "cash_expense_amount": "현금지출",
    "transfer_amount": "계좌입금",
}
REVENUE_TOTAL_LABELS = {
    "collected_amount": "수납액",
    "total_expense": "총지출",
    "cash_total": "현금",
}
REVENUE_TOTAL_FORMULAS = {
    "collected_amount": (
        ("field", "total_medical_fee", 1),
        ("field", "nhis_burden_total", -1),
        ("field", "unpaid_amount", -1),
        ("field", "health_living_fee", -1),
        ("field", "disability_fund", -1),
    ),
    "total_expense": (
        ("total", "collected_amount", 1),
        ("field", "card_amount", -1),
        ("field", "discount_amount", -1),
        ("field", "free_amount", -1),
        ("field", "cash_expense_amount", -1),
        ("field", "transfer_amount", -1),
    ),
    "cash_total": (
        ("total", "collected_amount", 1),
        ("total", "total_expense", -1),
    ),
}
REVENUE_FIELD_MEMO_KEYS = (*PAYMENT_FIELDS, *REVENUE_TOTAL_LABELS.keys())
# m032 이전 형식 감지용: 총진료비·공단부담총액 없이 아래 항목만 있으면
# 수납액 공식(총진료비 기반) 결과가 부정확하므로 화면에 재입력 안내를 띄운다.
LEGACY_VALUE_FIELDS = (
    "cash_amount",
    "card_amount",
    "transfer_amount",
    "unpaid_amount",
    "health_living_fee",
    "disability_fund",
    "other_amount",
)
REVENUE_RECORD_HEADER_ALIASES = {
    "record_date": ("날짜", "일자", "진료일자", "진료일", "내원일", "수납일", "record date", "date"),
    "total_medical_fee": ("총진료비", "총 진료비", "진료비총액", "진료비 총액"),
    "nhis_burden_total": ("공단부담총액", "공단 부담 총액", "공단부담금", "공단부담", "공담부담총액"),
    "cash_amount": ("현금수납액", "현금수납앱", "현금 수납액", "현금", "수납현금", "수납(현금)"),
    "card_amount": ("카드수납액", "카드 수납액", "카드", "수납카드", "수납(카드)"),
    "receivable_income": ("미수입금", "미수 입금", "미수수납", "미수 회수", "미수회수"),
    "unpaid_amount": ("미수발생", "미수 발생", "미수", "미수금", "미수납"),
    "health_living_fee": ("건강생활유지비", "건강 생활 유지비", "건생비"),
    "certificate_amount": ("입,통원확인서", "입통원확인서", "입 통원 확인서", "확인서", "통원확인서"),
    "disability_fund": ("장애인기금", "장애인 기금"),
    "uninsured_amount": ("비급여", "비급여액", "비급여총액"),
    "meal_amount": ("식대", "식대총액"),
    "other_amount": ("기타", "기타금액"),
    "discount_amount": ("할인", "할인액", "할인금액"),
    "free_amount": ("FREE", "free", "무료", "프리"),
    "cash_expense_amount": ("현금지출", "현금 지출", "지출현금"),
    "transfer_amount": ("계좌입금", "계좌 입금", "계좌", "입금계좌", "계좌이체"),
    "memo": ("메모", "비고", "참고"),
}
CASH_DENOMINATIONS = (50000, 10000, 5000, 1000, 500, 100, 50, 10)

MEDICAL_SUMMARY_FIELDS = (
    "total_medical_fee",
    "nhis_burden_total",
    "patient_burden_total",
    "covered_total",
    "uncovered_total",
)
MEDICAL_SUMMARY_LABELS = {
    "total_medical_fee": "총진료비",
    "nhis_burden_total": "공단부담총액",
    "patient_burden_total": "본인부담총액",
    "covered_total": "급여총액",
    "uncovered_total": "비급여총액",
}
MEDICAL_SUMMARY_HEADER_ALIASES = {
    "summary_date": ("날짜", "일자", "진료일자", "진료일", "내원일", "수납일"),
    "total_medical_fee": ("총진료비", "총 진료비", "진료비총액", "진료비 총액"),
    "nhis_burden_total": ("공단부담총액", "공단 부담 총액", "공단부담금", "공단부담"),
    "patient_burden_total": ("본인부담총액", "본인 부담 총액", "본인부담금", "본인부담"),
    "covered_total": ("급여총액", "급여 총액", "급여"),
    "uncovered_total": ("비급여총액", "비급여 총액", "비급여"),
}

REPORT_FIELD_TYPES = {"short_text", "long_text", "number", "checkbox"}
REPORT_FIELD_TYPE_LABELS = {
    "short_text": "짧은글",
    "long_text": "긴글",
    "number": "숫자",
    "checkbox": "체크",
}


def _parse_date(value: str) -> date:
    try:
        return date.fromisoformat((value or "").strip())
    except Exception as exc:
        raise ValueError("날짜 형식은 YYYY-MM-DD 이어야 합니다.") from exc


def resolve_range(date_from: str, date_to: str) -> tuple[date, date, list[str]]:
    start = _parse_date(date_from)
    end = _parse_date(date_to)
    if start > end:
        raise ValueError("시작일이 종료일보다 늦을 수 없습니다.")
    days = (end - start).days + 1
    if days > 370:
        raise ValueError("조회 기간은 최대 370일입니다.")
    keys = [(start + timedelta(days=i)).isoformat() for i in range(days)]
    return start, end, keys


def _category_id(value: str | None) -> str:
    return (value or "").strip()


def _amount(value) -> int:
    try:
        return max(0, int(value or 0))
    except Exception:
        return 0


def _money_amount(value) -> int:
    try:
        return int(value or 0)
    except Exception:
        return 0


def _normalize_excel_header(value) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").strip().lower())


def _medical_header_lookup() -> dict[str, str]:
    out = {}
    for field, aliases in MEDICAL_SUMMARY_HEADER_ALIASES.items():
        for alias in aliases:
            key = _normalize_excel_header(alias)
            if key:
                out[key] = field
    return out


def _medical_header_field(value) -> str | None:
    norm = _normalize_excel_header(value)
    if not norm:
        return None
    lookup = _medical_header_lookup()
    if norm in lookup:
        return lookup[norm]
    for key, field in sorted(lookup.items(), key=lambda x: len(x[0]), reverse=True):
        if key and key in norm:
            return field
    return None


def _revenue_header_lookup(labels: dict[str, str] | None = None) -> dict[str, str]:
    out = {}
    for field, aliases in REVENUE_RECORD_HEADER_ALIASES.items():
        for alias in aliases:
            key = _normalize_excel_header(alias)
            if key:
                out[key] = field
    for field, label in (labels or {}).items():
        if field not in (*PAYMENT_FIELDS, "record_date", "memo"):
            continue
        key = _normalize_excel_header(label)
        if key:
            out[key] = field
    return out


def _revenue_header_field(value, labels: dict[str, str] | None = None) -> str | None:
    norm = _normalize_excel_header(value)
    if not norm:
        return None
    lookup = _revenue_header_lookup(labels)
    if norm in lookup:
        return lookup[norm]
    for key, field in sorted(lookup.items(), key=lambda x: len(x[0]), reverse=True):
        if key and key in norm:
            return field
    return None


def _parse_excel_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        number = int(value)
        text_number = str(number)
        if len(text_number) == 8:
            try:
                return date(int(text_number[:4]), int(text_number[4:6]), int(text_number[6:8])).isoformat()
            except Exception:
                pass
        try:
            from openpyxl.utils.datetime import from_excel

            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date().isoformat()
            if isinstance(parsed, date):
                return parsed.isoformat()
        except Exception:
            return None
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"(\d{4})\D{0,3}(\d{1,2})\D{0,3}(\d{1,2})", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
        except Exception:
            return None
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 8:
        try:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8])).isoformat()
        except Exception:
            return None
    match = re.fullmatch(r"(\d{1,2})\s*[월./\-]\s*(\d{1,2})\s*일?\s*\.?", text)
    if match:
        month, day_num = int(match.group(1)), int(match.group(2))
        if 1 <= month <= 12 and 1 <= day_num <= 31:
            today = date.today()
            try:
                parsed = date(today.year, month, day_num)
                # 매출은 과거 데이터이므로 미래 날짜가 되면 작년으로 본다
                # (예: 1월에 작년 12월 일계표를 가져오는 경우).
                if parsed > today:
                    parsed = date(today.year - 1, month, day_num)
            except Exception:
                return None
            return parsed.isoformat()
    return None


def _parse_excel_amount(value) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value or "").strip()
    if not text:
        return 0
    negative = text.startswith("(") and text.endswith(")")
    cleaned = (
        text.replace(",", "")
        .replace("원", "")
        .replace("₩", "")
        .replace(" ", "")
    )
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return 0
    amount = int(round(float(match.group(0))))
    return -abs(amount) if negative else amount


def _cash_counts(value) -> dict[str, int]:
    raw = value or {}
    if isinstance(raw, str):
        try:
            raw = json.loads(raw or "{}")
        except Exception:
            raw = {}
    if not isinstance(raw, dict):
        raw = {}
    return {
        str(denom): _amount(raw.get(str(denom), raw.get(denom, 0)))
        for denom in CASH_DENOMINATIONS
    }


def _cash_counts_total(counts: dict[str, int]) -> int:
    return sum(denom * _amount(counts.get(str(denom))) for denom in CASH_DENOMINATIONS)


def _field_memos(value) -> dict[str, str]:
    raw = value or {}
    if isinstance(raw, str):
        try:
            raw = json.loads(raw or "{}")
        except Exception:
            raw = {}
    if not isinstance(raw, dict):
        raw = {}
    out = {}
    for key in REVENUE_FIELD_MEMO_KEYS:
        text = str(raw.get(key) or "").strip()
        if text:
            out[key] = text[:300]
    return out


def _payment_applied_amount(field: str, value) -> int:
    return _money_amount(value)


def _field_value(data, field: str) -> int:
    return _money_amount(getattr(data, field, 0) if not isinstance(data, dict) else data.get(field))


def _calculated_revenue_totals(data) -> dict[str, int]:
    totals: dict[str, int] = {}
    for total_key, terms in REVENUE_TOTAL_FORMULAS.items():
        amount = 0
        for source, field, sign in terms:
            base = totals.get(field, 0) if source == "total" else _field_value(data, field)
            amount += base * sign
        totals[total_key] = amount
    return totals


def _total_from_parts(data) -> int:
    return _calculated_revenue_totals(data)["collected_amount"]


def serialize_category(c: models.EmployeeCategory) -> dict:
    return {
        "id": c.id,
        "name": c.name,
        "color": c.color or "#9CA3AF",
        "active": bool(c.active),
        "sort_order": c.sort_order or 0,
    }


def active_categories(db: Session) -> list[dict]:
    return [
        serialize_category(c)
        for c in db.query(models.EmployeeCategory)
        .filter(models.EmployeeCategory.active == True)  # noqa: E712
        .order_by(models.EmployeeCategory.sort_order, models.EmployeeCategory.name)
        .all()
    ]


def category_name_map(db: Session) -> dict[str, str]:
    return {
        c.id: c.name
        for c in db.query(models.EmployeeCategory).all()
    }


def serialize_record(rec: models.RevenueRecord | None, record_date: str, category_id: str, categories: dict[str, str] | None = None) -> dict:
    categories = categories or {}
    data = {
        "id": rec.id if rec else "",
        "record_date": record_date,
        "category_id": category_id,
        "category_name": categories.get(category_id, "전체") if category_id else "전체",
        "cash_counts": _cash_counts(rec.cash_counts_json if rec else None),
        "field_memos": _field_memos(getattr(rec, "field_memos_json", "{}") if rec else None),
        "memo": (rec.memo or "") if rec else "",
        "created_at": rec.created_at.isoformat() if rec and rec.created_at else None,
        "updated_at": rec.updated_at.isoformat() if rec and rec.updated_at else None,
    }
    for field in PAYMENT_FIELDS:
        data[field] = int(getattr(rec, field, 0) or 0) if rec else 0
    data["legacy_format"] = bool(
        rec
        and not data["total_medical_fee"]
        and not data["nhis_burden_total"]
        and any(data[field] for field in LEGACY_VALUE_FIELDS)
    )
    data["unpaid_applied_amount"] = _payment_applied_amount("unpaid_amount", data["unpaid_amount"])
    data.update(_calculated_revenue_totals(data))
    data["total_amount"] = data["collected_amount"]
    return data


def list_records(db: Session, date_from: str, date_to: str, category_id: str = "") -> dict:
    start, end, date_keys = resolve_range(date_from, date_to)
    category_id = _category_id(category_id)
    categories = category_name_map(db)
    rows = (
        db.query(models.RevenueRecord)
        .filter(
            models.RevenueRecord.record_date >= start.isoformat(),
            models.RevenueRecord.record_date <= end.isoformat(),
            models.RevenueRecord.category_id == category_id,
        )
        .all()
    )
    by_date = {r.record_date: r for r in rows}
    return {
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "category_id": category_id,
        "categories": active_categories(db),
        "records": [
            serialize_record(by_date.get(day), day, category_id, categories)
            for day in date_keys
        ],
    }


def _entry_has_value(entry: RevenueRecordEntry) -> bool:
    if entry.cash_counts is not None and _cash_counts_total(_cash_counts(entry.cash_counts)) > 0:
        return True
    return (
        any(_money_amount(getattr(entry, f, 0)) != 0 for f in PAYMENT_FIELDS)
        or bool((entry.memo or "").strip())
        or bool(_field_memos(entry.field_memos))
    )


def upsert_grid(
    db: Session,
    payload: RevenueGridIn,
    *,
    log_callback: Callable | None = None,
    audit_callback: Callable | None = None,
) -> dict:
    start, end, date_keys = resolve_range(payload.date_from, payload.date_to)
    allowed_dates = set(date_keys)
    category_id = _category_id(payload.category_id)
    changed = {"upserted": 0, "deleted": 0}

    for entry in payload.entries:
        record_date = (entry.record_date or "").strip()
        if record_date not in allowed_dates:
            continue
        row_category = category_id
        rec = (
            db.query(models.RevenueRecord)
            .filter(
                models.RevenueRecord.record_date == record_date,
                models.RevenueRecord.category_id == row_category,
            )
            .first()
        )
        if not _entry_has_value(entry):
            if rec:
                rec_id = rec.id
                if log_callback:
                    log_callback(db, "revenue_record", rec_id, "delete", rec)
                db.delete(rec)
                db.flush()
                changed["deleted"] += 1
            continue
        if not rec:
            rec = models.RevenueRecord(record_date=record_date, category_id=row_category)
            db.add(rec)
        counts = _cash_counts(entry.cash_counts)
        if entry.cash_counts is not None:
            rec.cash_amount = _cash_counts_total(counts)
            rec.cash_counts_json = json.dumps(counts, ensure_ascii=False)
        else:
            rec.cash_amount = _money_amount(entry.cash_amount)
            rec.cash_counts_json = "{}"
        for field in PAYMENT_FIELDS:
            if field == "cash_amount":
                continue
            setattr(rec, field, _money_amount(getattr(entry, field, 0)))
        rec.field_memos_json = json.dumps(_field_memos(entry.field_memos), ensure_ascii=False)
        rec.memo = (entry.memo or "").strip()[:500]
        rec.updated_at = datetime.utcnow()
        db.flush()
        if log_callback:
            log_callback(db, "revenue_record", rec.id, "upsert", rec)
        changed["upserted"] += 1

    if audit_callback:
        audit_callback(
            db,
            "revenue.records.grid.save",
            "",
            f"{start.isoformat()}~{end.isoformat()} category_id={category_id}",
        )
    return changed


def _records_for_period(db: Session, start: date, end: date, category_id: str) -> list[models.RevenueRecord]:
    return (
        db.query(models.RevenueRecord)
        .filter(
            models.RevenueRecord.record_date >= start.isoformat(),
            models.RevenueRecord.record_date <= end.isoformat(),
            models.RevenueRecord.category_id == category_id,
        )
        .order_by(models.RevenueRecord.record_date)
        .all()
    )


def _revenue_summary(records: list[models.RevenueRecord]) -> dict:
    by_payment = {
        key: {"key": key, "label": PAYMENT_LABELS[key], "amount": 0}
        for key in PAYMENT_FIELDS
    }
    daily = []
    total = 0
    total_expense = 0
    cash_total = 0
    legacy_count = 0
    for rec in records:
        row = serialize_record(rec, rec.record_date, rec.category_id)
        total += row["total_amount"]
        total_expense += row["total_expense"]
        cash_total += row["cash_total"]
        if row["legacy_format"]:
            legacy_count += 1
        for key in PAYMENT_FIELDS:
            by_payment[key]["amount"] += _payment_applied_amount(key, row[key])
        daily_row = {
            "date": rec.record_date,
            "total_amount": row["total_amount"],
            "collected_amount": row["collected_amount"],
            "total_expense": row["total_expense"],
            "cash_total": row["cash_total"],
            "unpaid_raw_amount": row["unpaid_amount"],
        }
        for key in PAYMENT_FIELDS:
            daily_row[key] = row["unpaid_applied_amount"] if key == "unpaid_amount" else row[key]
        daily.append(daily_row)
    summary = {
        "record_count": len(records),
        "revenue_total": total,
        "collected_amount": total,
        "total_expense": total_expense,
        "cash_total": cash_total,
        "legacy_count": legacy_count,
        "by_payment": list(by_payment.values()),
        "daily": daily,
    }
    for key in PAYMENT_FIELDS:
        summary[key] = by_payment[key]["amount"]
    return summary


def _settlement_summary(db: Session, start: date, end: date, category_id: str) -> dict:
    q = db.query(models.SettlementRecord).filter(
        models.SettlementRecord.performed_on >= start.isoformat(),
        models.SettlementRecord.performed_on <= end.isoformat(),
    )
    if category_id:
        q = q.filter(models.SettlementRecord.employee_category_id_snapshot == category_id)
    records = q.all()
    quantity_total = 0
    price_total = 0
    incentive_total = 0
    employee_ids = set()
    treatment_ids = set()
    by_treatment: dict[str, dict] = {}
    for rec in records:
        qty = _amount(rec.quantity)
        rec_price_total = _amount(rec.price_snapshot) * qty
        rec_incentive_total = _amount(rec.incentive_amount)
        quantity_total += qty
        price_total += rec_price_total
        incentive_total += rec_incentive_total
        if rec.employee_id:
            employee_ids.add(rec.employee_id)
        if rec.treatment_id:
            treatment_ids.add(rec.treatment_id)
        treatment_key = rec.treatment_id or rec.treatment_code_snapshot or rec.treatment_name_snapshot or "-"
        treatment_row = by_treatment.setdefault(treatment_key, {
            "key": treatment_key,
            "treatment_id": rec.treatment_id or "",
            "treatment_name": rec.treatment_name_snapshot or "-",
            "treatment_short": rec.treatment_short_snapshot or rec.treatment_name_snapshot or "-",
            "quantity_total": 0,
            "price_total": 0,
            "incentive_total": 0,
            "record_count": 0,
        })
        treatment_row["quantity_total"] += qty
        treatment_row["price_total"] += rec_price_total
        treatment_row["incentive_total"] += rec_incentive_total
        treatment_row["record_count"] += 1
    return {
        "record_count": len(records),
        "quantity_total": quantity_total,
        "price_total": price_total,
        "incentive_total": incentive_total,
        "employee_count": len(employee_ids),
        "treatment_count": len(treatment_ids),
        "by_treatment": sorted(
            by_treatment.values(),
            key=lambda row: (-row["price_total"], str(row["treatment_short"] or row["treatment_name"])),
        ),
    }


def _previous_period(start: date, end: date) -> tuple[date, date]:
    days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days - 1)
    return prev_start, prev_end


def _delta(current: int, previous: int) -> dict:
    amount = current - previous
    pct = None if previous == 0 else round((amount / previous) * 100, 1)
    return {"amount": amount, "pct": pct}


def stats(db: Session, date_from: str, date_to: str, category_id: str = "") -> dict:
    start, end, _ = resolve_range(date_from, date_to)
    category_id = _category_id(category_id)
    prev_start, prev_end = _previous_period(start, end)

    current_records = _records_for_period(db, start, end, category_id)
    previous_records = _records_for_period(db, prev_start, prev_end, category_id)
    current = _revenue_summary(current_records)
    previous = _revenue_summary(previous_records)
    settlement = _settlement_summary(db, start, end, category_id)

    revenue_total = current["revenue_total"]
    settlement["revenue_minus_settlement_price"] = revenue_total - settlement["price_total"]
    settlement["revenue_after_incentive"] = revenue_total - settlement["incentive_total"]
    settlement["incentive_rate"] = round((settlement["incentive_total"] / revenue_total) * 100, 1) if revenue_total else 0
    settlement["settlement_price_rate"] = round((settlement["price_total"] / revenue_total) * 100, 1) if revenue_total else 0

    return {
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "category_id": category_id,
        "categories": active_categories(db),
        "compare_from": prev_start.isoformat(),
        "compare_to": prev_end.isoformat(),
        "current": current,
        "previous": previous,
        "delta": {
            "revenue_total": _delta(current["revenue_total"], previous["revenue_total"]),
            "record_count": _delta(current["record_count"], previous["record_count"]),
        },
        "settlement": settlement,
    }


def _json_list(value: str) -> list:
    try:
        data = json.loads(value or "[]")
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _treatment_rows(db: Session) -> list[models.Treatment]:
    return (
        db.query(models.Treatment)
        .order_by(models.Treatment.sort_order, models.Treatment.name)
        .all()
    )


def _treatment_maps(db: Session) -> tuple[list[models.Treatment], dict[str, models.Treatment], dict[str, models.Treatment]]:
    rows = _treatment_rows(db)
    return rows, {t.code: t for t in rows}, {t.id: t for t in rows}


def _serialize_report_treatment(t: models.Treatment) -> dict:
    return {
        "id": t.id,
        "code": t.code,
        "name": t.name,
        "short": t.short or t.name or t.code,
        "category_id": t.category_id or "",
        "category_name": t.category.name if getattr(t, "category", None) else "",
        "active": bool(t.active),
        "sort_order": int(t.sort_order or 0),
    }


def _normalize_selected_codes(db: Session, codes: list[str]) -> list[str]:
    _, by_code, _ = _treatment_maps(db)
    out = []
    seen = set()
    for raw in codes or []:
        code = str(raw or "").strip()
        if not code or code in seen:
            continue
        if code not in by_code:
            raise ValueError(f"존재하지 않는 치료항목입니다: {code}")
        seen.add(code)
        out.append(code)
    return out


def _number_value(value) -> int | float:
    if value in (None, ""):
        return 0
    try:
        cleaned = str(value).replace(",", "").strip()
        n = float(cleaned)
    except Exception as exc:
        raise ValueError("숫자 칸은 0 이상의 숫자여야 합니다.") from exc
    if n < 0:
        raise ValueError("숫자 칸은 0 이상의 숫자여야 합니다.")
    return int(n) if n.is_integer() else n


def _checkbox_value(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on", "checked"}


def _normalize_report_fields(fields: list[DailyReportField]) -> list[dict]:
    if len(fields or []) > 30:
        raise ValueError("보고 칸은 최대 30개까지 저장할 수 있습니다.")
    out = []
    for idx, field in enumerate(fields or []):
        label = (field.label or "").strip()
        if not label:
            raise ValueError("보고 칸 제목을 입력하세요.")
        if len(label) > 50:
            raise ValueError("보고 칸 제목은 최대 50자입니다.")
        field_type = (field.type or "").strip()
        if field_type not in REPORT_FIELD_TYPES:
            raise ValueError(f"지원하지 않는 보고 칸 타입입니다: {field_type}")
        if field_type == "short_text":
            value = str(field.value or "").strip()
            if len(value) > 200:
                raise ValueError("짧은글 칸은 최대 200자입니다.")
        elif field_type == "long_text":
            value = str(field.value or "").strip()
            if len(value) > 2000:
                raise ValueError("긴글 칸은 최대 2000자입니다.")
        elif field_type == "number":
            value = _number_value(field.value)
        else:
            value = _checkbox_value(field.value)
        field_id = (field.id or "").strip()
        if not field_id or len(field_id) > 60:
            field_id = f"field_{models.uid()[:10]}"
        out.append({
            "id": field_id,
            "label": label,
            "type": field_type,
            "type_label": REPORT_FIELD_TYPE_LABELS[field_type],
            "value": value,
            "sort_order": idx,
        })
    return out


def _saved_report_fields(report: models.DailyWorkReport | None) -> list[dict]:
    fields = []
    for idx, raw in enumerate(_json_list(report.custom_fields_json if report else "")):
        if not isinstance(raw, dict):
            continue
        field_type = raw.get("type")
        if field_type not in REPORT_FIELD_TYPES:
            field_type = "long_text"
        fields.append({
            "id": str(raw.get("id") or f"field_{idx}"),
            "label": str(raw.get("label") or ""),
            "type": field_type,
            "type_label": REPORT_FIELD_TYPE_LABELS[field_type],
            "value": raw.get("value"),
            "sort_order": int(raw.get("sort_order") or idx),
        })
    return sorted(fields, key=lambda x: x["sort_order"])


def _daily_revenue_record(db: Session, report_date: str) -> dict:
    rec = (
        db.query(models.RevenueRecord)
        .filter(
            models.RevenueRecord.record_date == report_date,
            models.RevenueRecord.category_id == "",
        )
        .first()
    )
    data = serialize_record(rec, report_date, "")
    data["exists"] = bool(rec)
    return data


def serialize_medical_summary(rec: models.DailyMedicalSummary | None, report_date: str) -> dict:
    data = {
        "id": rec.id if rec else "",
        "report_date": report_date,
        "summary_date": rec.summary_date if rec else report_date,
        "total_medical_fee": int(rec.total_medical_fee or 0) if rec else 0,
        "nhis_burden_total": int(rec.nhis_burden_total or 0) if rec else 0,
        "patient_burden_total": int(rec.patient_burden_total or 0) if rec else 0,
        "covered_total": int(rec.covered_total or 0) if rec else 0,
        "uncovered_total": int(rec.uncovered_total or 0) if rec else 0,
        "source_filename": rec.source_filename if rec else "",
        "created_at": rec.created_at.isoformat() if rec and rec.created_at else None,
        "updated_at": rec.updated_at.isoformat() if rec and rec.updated_at else None,
        "exists": bool(rec),
    }
    data["lines"] = [
        {"key": field, "label": MEDICAL_SUMMARY_LABELS[field], "amount": data[field], "source": "기간별 데이터"}
        for field in MEDICAL_SUMMARY_FIELDS
    ]
    return data


def _daily_medical_summary(db: Session, report_date: str) -> dict:
    rec = (
        db.query(models.DailyMedicalSummary)
        .filter(models.DailyMedicalSummary.summary_date == report_date)
        .first()
    )
    return serialize_medical_summary(rec, report_date)


def _detect_medical_summary_header(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    required = ("summary_date", *MEDICAL_SUMMARY_FIELDS)
    for row_idx, row in enumerate(rows[:30]):
        mapping: dict[str, int] = {}
        for col_idx, value in enumerate(row):
            field = _medical_header_field(value)
            if field and field not in mapping:
                mapping[field] = col_idx
        if all(field in mapping for field in required):
            return row_idx, mapping
    labels = ["날짜", *[MEDICAL_SUMMARY_LABELS[field] for field in MEDICAL_SUMMARY_FIELDS]]
    raise ValueError("엑셀 헤더에 " + ", ".join(labels) + " 항목이 필요합니다.")


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


def _detect_revenue_record_header(rows: list[tuple], labels: dict[str, str] | None = None) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(rows[:30]):
        mapping: dict[str, int] = {}
        for col_idx, value in enumerate(row):
            field = _revenue_header_field(value, labels)
            if field and field not in mapping:
                mapping[field] = col_idx
        if "record_date" in mapping and any(field in mapping for field in PAYMENT_FIELDS):
            return row_idx, mapping
    raise ValueError("엑셀 헤더에 날짜와 매출 입력 항목이 필요합니다.")


def _serialize_record_data(
    record_date: str,
    category_id: str = "",
    values: dict[str, int] | None = None,
    memo: str = "",
    categories: dict[str, str] | None = None,
) -> dict:
    data = {
        "id": "",
        "record_date": record_date,
        "category_id": category_id,
        "category_name": (categories or {}).get(category_id, "전체") if category_id else "전체",
        "cash_counts": {},
        "field_memos": {},
        "memo": memo,
        "created_at": None,
        "updated_at": None,
    }
    values = values or {}
    for field in PAYMENT_FIELDS:
        data[field] = _money_amount(values.get(field))
    data["unpaid_applied_amount"] = _payment_applied_amount("unpaid_amount", data["unpaid_amount"])
    data.update(_calculated_revenue_totals(data))
    data["total_amount"] = data["collected_amount"]
    return data


def preview_records_from_excel(
    db: Session,
    file_bytes: bytes,
    filename: str = "",
    labels: dict[str, str] | None = None,
) -> dict:
    if not file_bytes:
        raise ValueError("엑셀 파일이 비어 있습니다.")
    try:
        import openpyxl
    except Exception as exc:
        raise ValueError("엑셀 파싱 라이브러리(openpyxl)를 불러올 수 없습니다.") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("엑셀 파일에 데이터가 없습니다.")
    header_idx, mapping = _detect_revenue_record_header(rows, labels)

    grouped: dict[str, dict[str, int]] = {}
    memo_by_day: dict[str, list[str]] = {}
    skipped = 0
    for row in rows[header_idx + 1:]:
        if not any(str(value or "").strip() for value in row):
            continue
        raw_day = _cell(row, mapping["record_date"])
        day = _parse_excel_date(raw_day)
        if not day:
            total_label = _normalize_excel_header(raw_day)
            if not str(raw_day or "").strip() or total_label in {"합계", "총계", "소계", "total", "sum"}:
                continue
            skipped += 1
            continue
        item = grouped.setdefault(day, {field: 0 for field in PAYMENT_FIELDS})
        for field in PAYMENT_FIELDS:
            if field in mapping:
                item[field] += _parse_excel_amount(_cell(row, mapping[field]))
        if "memo" in mapping:
            memo = str(_cell(row, mapping["memo"]) or "").strip()
            if memo:
                memo_by_day.setdefault(day, []).append(memo)

    if not grouped:
        raise ValueError("가져올 매출 기록 데이터가 없습니다.")

    dates = sorted(grouped.keys())
    # 이미 저장된 기록이 있는 날짜 — 프론트에서 덮어쓰기 확인창을 띄우는 데 사용
    existing_dates = sorted({
        row[0]
        for row in db.query(models.RevenueRecord.record_date)
        .filter(
            models.RevenueRecord.record_date >= dates[0],
            models.RevenueRecord.record_date <= dates[-1],
            models.RevenueRecord.category_id == "",
        )
        .all()
        if row[0] in grouped
    })
    categories = category_name_map(db)
    records = [
        _serialize_record_data(
            day,
            "",
            grouped[day],
            " / ".join(memo_by_day.get(day, []))[:500],
            categories,
        )
        for day in dates
    ]
    return {
        "date_from": dates[0],
        "date_to": dates[-1],
        "category_id": "",
        "categories": active_categories(db),
        "records": records,
        "existing_dates": existing_dates,
        "imported": len(records),
        "skipped": skipped,
        "source_filename": (filename or "").strip()[:255],
        "fields": [
            {"key": field, "label": PAYMENT_LABELS[field]}
            for field in PAYMENT_FIELDS
        ],
    }


def import_medical_summaries_from_excel(
    db: Session,
    file_bytes: bytes,
    filename: str = "",
    *,
    log_callback: Callable | None = None,
    audit_callback: Callable | None = None,
) -> dict:
    if not file_bytes:
        raise ValueError("엑셀 파일이 비어 있습니다.")
    try:
        import openpyxl
    except Exception as exc:
        raise ValueError("엑셀 파싱 라이브러리(openpyxl)를 불러올 수 없습니다.") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("엑셀 파일에 데이터가 없습니다.")
    header_idx, mapping = _detect_medical_summary_header(rows)

    grouped: dict[str, dict[str, int]] = {}
    skipped = 0
    source_filename = (filename or "").strip()[:255]
    for row in rows[header_idx + 1:]:
        if not any(str(value or "").strip() for value in row):
            continue
        raw_day = _cell(row, mapping["summary_date"])
        day = _parse_excel_date(raw_day)
        if not day:
            total_label = _normalize_excel_header(raw_day)
            # Exported period reports often end with a summary row without a date.
            if not str(raw_day or "").strip() or total_label in {"합계", "총계", "소계", "total", "sum"}:
                continue
            skipped += 1
            continue
        item = grouped.setdefault(day, {field: 0 for field in MEDICAL_SUMMARY_FIELDS})
        for field in MEDICAL_SUMMARY_FIELDS:
            item[field] += _parse_excel_amount(_cell(row, mapping[field]))

    if not grouped:
        raise ValueError("가져올 날짜별 데이터가 없습니다.")

    now = datetime.utcnow()
    changed = {"upserted": 0}
    for day, amounts in sorted(grouped.items()):
        rec = (
            db.query(models.DailyMedicalSummary)
            .filter(models.DailyMedicalSummary.summary_date == day)
            .first()
        )
        if not rec:
            rec = models.DailyMedicalSummary(summary_date=day)
            db.add(rec)
        for field, amount in amounts.items():
            setattr(rec, field, int(amount or 0))
        rec.source_filename = source_filename
        rec.updated_at = now
        db.flush()
        if log_callback:
            log_callback(db, "daily_medical_summary", rec.id, "upsert", rec)
        changed["upserted"] += 1

    dates = sorted(grouped.keys())
    if audit_callback:
        audit_callback(
            db,
            "revenue.daily_medical_summary.import",
            "",
            f"{source_filename or 'uploaded'} {dates[0]}~{dates[-1]} rows={len(grouped)} skipped={skipped}",
        )
    return {
        "changed": changed,
        "imported": len(grouped),
        "skipped": skipped,
        "date_from": dates[0],
        "date_to": dates[-1],
        "fields": [
            {"key": field, "label": MEDICAL_SUMMARY_LABELS[field]}
            for field in MEDICAL_SUMMARY_FIELDS
        ],
    }


def _settlement_item_base(
    treatment: models.Treatment | None,
    code: str,
    rec: models.SettlementRecord | None = None,
) -> dict:
    name = treatment.name if treatment else (rec.treatment_name_snapshot if rec else code)
    short = treatment.short if treatment else (rec.treatment_short_snapshot if rec else name)
    category_id = treatment.category_id if treatment else (rec.employee_category_id_snapshot if rec else "")
    category_name = (
        treatment.category.name if treatment and getattr(treatment, "category", None)
        else (rec.employee_category_name_snapshot if rec else "")
    )
    return {
        "code": code,
        "treatment_id": treatment.id if treatment else (rec.treatment_id if rec else ""),
        "treatment_name": name or code,
        "treatment_short": short or name or code,
        "category_id": category_id or "",
        "category_name": category_name or "",
        "quantity_total": 0,
        "price_total": 0,
        "incentive_total": 0,
        "net_total": 0,
        "record_count": 0,
    }


def _daily_settlement_by_code(db: Session, report_date: str) -> dict[str, dict]:
    _, by_code, by_id = _treatment_maps(db)
    rows = (
        db.query(models.SettlementRecord)
        .filter(models.SettlementRecord.performed_on == report_date)
        .all()
    )
    by_code_data: dict[str, dict] = {}
    for rec in rows:
        treatment = by_id.get(rec.treatment_id) if rec.treatment_id else None
        code = (treatment.code if treatment else "") or rec.treatment_code_snapshot or ""
        if not code:
            continue
        treatment = treatment or by_code.get(code)
        row = by_code_data.setdefault(code, _settlement_item_base(treatment, code, rec))
        qty = _amount(rec.quantity)
        price_total = _amount(rec.price_snapshot) * qty
        incentive_total = _amount(rec.incentive_amount)
        row["quantity_total"] += qty
        row["price_total"] += price_total
        row["incentive_total"] += incentive_total
        row["net_total"] += price_total - incentive_total
        row["record_count"] += 1
    return by_code_data


def _auto_summary(db: Session, report_date: str, selected_codes: list[str]) -> dict:
    _, by_code, _ = _treatment_maps(db)
    settlement = _daily_settlement_by_code(db, report_date)
    items = []
    for code in selected_codes:
        if code in settlement:
            row = dict(settlement[code])
        else:
            row = _settlement_item_base(by_code.get(code), code)
        items.append(row)
    totals = {
        "quantity_total": sum(int(x.get("quantity_total") or 0) for x in items),
        "price_total": sum(int(x.get("price_total") or 0) for x in items),
        "incentive_total": sum(int(x.get("incentive_total") or 0) for x in items),
        "net_total": sum(int(x.get("net_total") or 0) for x in items),
        "record_count": sum(int(x.get("record_count") or 0) for x in items),
    }
    return {"items": items, "totals": totals}


def _daily_journal_summary(revenue_record: dict, settlement_items: list[dict], medical_summary: dict) -> dict:
    revenue_lines = [
        {"key": "collected_amount", "label": REVENUE_TOTAL_LABELS["collected_amount"], "amount": _money_amount(revenue_record.get("collected_amount")), "source": "매출 기록"},
        {"key": "total_expense", "label": REVENUE_TOTAL_LABELS["total_expense"], "amount": _money_amount(revenue_record.get("total_expense")), "source": "매출 기록"},
        {"key": "cash_total", "label": REVENUE_TOTAL_LABELS["cash_total"], "amount": _money_amount(revenue_record.get("cash_total")), "source": "매출 기록"},
    ]
    revenue_lines.extend(
        {
            "key": field,
            "label": PAYMENT_LABELS[field],
            "amount": _money_amount(
                revenue_record.get("unpaid_applied_amount")
                if field == "unpaid_amount"
                else revenue_record.get(field)
            ),
            "source": "매출 기록",
        }
        for field in PAYMENT_FIELDS
    )
    treatment_lines = sorted(
        [
            {
                "code": str(row.get("code") or ""),
                "label": str(row.get("treatment_short") or row.get("treatment_name") or row.get("code") or ""),
                "quantity_total": _amount(row.get("quantity_total")),
                "price_total": _amount(row.get("price_total")),
                "incentive_total": _amount(row.get("incentive_total")),
                "net_total": _amount(row.get("net_total")),
                "source": "정산",
            }
            for row in settlement_items
        ],
        key=lambda row: (-row["price_total"], row["label"]),
    )
    totals = {
        "quantity_total": sum(row["quantity_total"] for row in treatment_lines),
        "price_total": sum(row["price_total"] for row in treatment_lines),
        "incentive_total": sum(row["incentive_total"] for row in treatment_lines),
        "net_total": sum(row["net_total"] for row in treatment_lines),
    }
    return {
        "revenue_lines": revenue_lines,
        "medical_lines": medical_summary.get("lines", []) if medical_summary.get("exists") else [],
        "treatment_lines": treatment_lines,
        "totals": totals,
    }


def _default_selected_codes(db: Session, report_date: str) -> list[str]:
    settlement = _daily_settlement_by_code(db, report_date)
    rows, by_code, _ = _treatment_maps(db)
    sort_index = {t.code: idx for idx, t in enumerate(rows)}
    return sorted(
        [
            code for code, row in settlement.items()
            if int(row.get("quantity_total") or 0) > 0 and code in by_code
        ],
        key=lambda code: sort_index.get(code, 9999),
    )


def get_daily_report(db: Session, report_date: str) -> dict:
    day = _parse_date(report_date).isoformat()
    report = (
        db.query(models.DailyWorkReport)
        .filter(models.DailyWorkReport.report_date == day)
        .first()
    )
    rows, by_code, _ = _treatment_maps(db)
    if report:
        selected = []
        seen = set()
        for raw in _json_list(report.selected_treatment_codes_json):
            code = str(raw or "").strip()
            if code and code in by_code and code not in seen:
                selected.append(code)
                seen.add(code)
    else:
        selected = _default_selected_codes(db, day)
    settlement_by_code = _daily_settlement_by_code(db, day)
    revenue_record = _daily_revenue_record(db, day)
    medical_summary = _daily_medical_summary(db, day)
    settlement_items = list(settlement_by_code.values())
    return {
        "report_date": day,
        "exists": bool(report),
        "selected_treatment_codes": selected,
        "custom_fields": _saved_report_fields(report),
        "revenue_record": revenue_record,
        "medical_summary": medical_summary,
        "treatments": [_serialize_report_treatment(t) for t in rows],
        "settlement_codes": list(settlement_by_code.keys()),
        "settlement_items": settlement_items,
        "auto": _auto_summary(db, day, selected),
        "journal": _daily_journal_summary(revenue_record, settlement_items, medical_summary),
        "updated_at": report.updated_at.isoformat() if report and report.updated_at else None,
    }


def save_daily_report(
    db: Session,
    payload: DailyReportIn,
    *,
    log_callback: Callable | None = None,
    audit_callback: Callable | None = None,
) -> dict:
    day = _parse_date(payload.report_date).isoformat()
    selected = _normalize_selected_codes(db, payload.selected_treatment_codes)
    fields = _normalize_report_fields(payload.custom_fields)
    report = (
        db.query(models.DailyWorkReport)
        .filter(models.DailyWorkReport.report_date == day)
        .first()
    )
    changed = {"upserted": 0, "deleted": 0}
    if not selected and not fields:
        if report:
            report_id = report.id
            if log_callback:
                log_callback(db, "daily_work_report", report_id, "delete", report)
            db.delete(report)
            db.flush()
            changed["deleted"] = 1
        if audit_callback:
            audit_callback(db, "revenue.daily_report.delete", "", day)
        return changed

    if not report:
        report = models.DailyWorkReport(report_date=day)
        db.add(report)
    report.selected_treatment_codes_json = json.dumps(selected, ensure_ascii=False)
    report.custom_fields_json = json.dumps(fields, ensure_ascii=False)
    report.updated_at = datetime.utcnow()
    db.flush()
    if log_callback:
        log_callback(db, "daily_work_report", report.id, "upsert", report)
    if audit_callback:
        audit_callback(db, "revenue.daily_report.save", report.id, day)
    changed["upserted"] = 1
    return changed
