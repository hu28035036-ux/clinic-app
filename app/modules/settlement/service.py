from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Callable
import re

from sqlalchemy.orm import Session

from ...models import constants as C
from ...models import models
from .rules import calculate_incentive_amount, incentive_snapshot_for_treatment
from .schemas import SettlementGridIn


def parse_date(value: str, field: str = "date") -> date:
    try:
        return datetime.strptime((value or "").strip(), "%Y-%m-%d").date()
    except Exception as exc:
        raise ValueError(f"{field} must be YYYY-MM-DD") from exc


def resolve_range(date_from: str, date_to: str) -> tuple[date, date, list[str], str]:
    start = parse_date(date_from, "date_from")
    end = parse_date(date_to, "date_to")
    if end < start:
        raise ValueError("date_to must be on or after date_from")
    days = []
    cur = start
    while cur <= end:
        days.append(cur.isoformat())
        cur += timedelta(days=1)
    return start, end, days, f"{start.isoformat()}~{end.isoformat()}"


def _serialize_category(c: models.EmployeeCategory) -> dict:
    return {
        "id": c.id,
        "name": c.name,
        "color": c.color or "#9CA3AF",
        "active": bool(c.active),
        "sort_order": c.sort_order or 0,
        "default_can_doctor_treatment": bool(c.default_can_doctor_treatment),
        "default_can_manual": bool(c.default_can_manual),
        "default_can_eswt": bool(c.default_can_eswt),
    }


def _unit_incentive_for_treatment(t: models.Treatment) -> int:
    rule, value = incentive_snapshot_for_treatment(t)
    return calculate_incentive_amount(getattr(t, "price", 0) or 0, rule, value, 1)


def _serialize_treatment(t: models.Treatment) -> dict:
    category = getattr(t, "category", None)
    return {
        "id": t.id,
        "code": t.code,
        "name": t.name,
        "category_id": t.category_id,
        "category_name": category.name if category else "",
        "short": t.short,
        "default_minutes": t.default_minutes,
        "role": t.role,
        "count_increment": t.count_increment,
        "show_in_patient": bool(t.show_in_patient),
        "active": bool(t.active),
        "sort_order": t.sort_order or 0,
        "price": int(getattr(t, "price", 0) or 0),
        "incentive_pct": getattr(t, "incentive_pct", None),
        "incentive_amount": getattr(t, "incentive_amount", None),
        "unit_incentive": _unit_incentive_for_treatment(t),
    }


def employee_treatment_ids(db: Session, e: models.Employee) -> list[str]:
    if getattr(e, "treatment_override_enabled", False):
        q = (
            db.query(models.Treatment)
            .join(
                models.EmployeeTreatment,
                models.EmployeeTreatment.treatment_id == models.Treatment.id,
            )
            .filter(
                models.EmployeeTreatment.employee_id == e.id,
                models.Treatment.active == True,  # noqa: E712
            )
        )
        if e.category_id:
            q = q.filter(models.Treatment.category_id == e.category_id)
        return [
            t.id
            for t in q.order_by(models.Treatment.sort_order, models.Treatment.name).all()
        ]
    if not e.category_id:
        return []
    return [
        t.id
        for t in db.query(models.Treatment)
        .filter(
            models.Treatment.category_id == e.category_id,
            models.Treatment.active == True,  # noqa: E712
        )
        .order_by(models.Treatment.sort_order, models.Treatment.name)
        .all()
    ]


def _employee_can_doctor_treatment(e: models.Employee) -> bool:
    override = getattr(e, "can_doctor_treatment_override", None)
    if override is not None:
        return bool(override)
    category = getattr(e, "category", None)
    if category is not None:
        return bool(category.default_can_doctor_treatment)
    return getattr(e, "role", "") == C.ROLE_DOCTOR


def _employee_can_manual(e: models.Employee) -> bool:
    override = getattr(e, "can_manual_override", None)
    if override is not None:
        return bool(override)
    category = getattr(e, "category", None)
    if category is not None:
        return bool(category.default_can_manual)
    return bool(getattr(e, "can_manual", True))


def _employee_can_eswt(e: models.Employee) -> bool:
    override = getattr(e, "can_eswt_override", None)
    if override is not None:
        return bool(override)
    category = getattr(e, "category", None)
    if category is not None:
        return bool(category.default_can_eswt)
    return bool(getattr(e, "can_eswt", True))


def _serialize_employee(db: Session, e: models.Employee) -> dict:
    category = getattr(e, "category", None)
    return {
        "id": e.id,
        "name": e.name,
        "category_id": e.category_id,
        "category_name": category.name if category else "",
        "color": e.color or "#9CA3AF",
        "active": bool(e.active),
        "treatment_override_enabled": bool(getattr(e, "treatment_override_enabled", False)),
        "treatment_ids": employee_treatment_ids(db, e),
        "sort_order": e.sort_order or 0,
    }


def _active_categories(db: Session) -> list[models.EmployeeCategory]:
    return (
        db.query(models.EmployeeCategory)
        .filter(models.EmployeeCategory.active == True)  # noqa: E712
        .order_by(models.EmployeeCategory.sort_order, models.EmployeeCategory.name)
        .all()
    )


def _choose_category_id(db: Session, category_id: str, categories: list[models.EmployeeCategory]) -> str:
    if category_id and any(c.id == category_id for c in categories):
        return category_id
    if not categories:
        return ""
    treatment_category_ids = {
        row[0]
        for row in db.query(models.Treatment.category_id)
        .filter(
            models.Treatment.active == True,  # noqa: E712
            models.Treatment.category_id.isnot(None),
        )
        .distinct()
        .all()
    }
    employee_category_ids = {
        row[0]
        for row in db.query(models.Employee.category_id)
        .filter(
            models.Employee.active == True,  # noqa: E712
            models.Employee.category_id.isnot(None),
        )
        .distinct()
        .all()
    }

    def usable(c: models.EmployeeCategory) -> bool:
        return c.id in treatment_category_ids and c.id in employee_category_ids

    selected = (
        next((c for c in categories if c.default_can_manual and usable(c)), None)
        or next((c for c in categories if usable(c)), None)
        or next((c for c in categories if c.id in treatment_category_ids), None)
        or categories[0]
    )
    return selected.id


def _treatments_for_category(
    db: Session,
    category_id: str,
    selected_category: models.EmployeeCategory | None,
) -> list[models.Treatment]:
    q = db.query(models.Treatment).filter(models.Treatment.active == True)  # noqa: E712
    if category_id:
        q = q.filter(models.Treatment.category_id == category_id)
    treatments = q.order_by(models.Treatment.sort_order, models.Treatment.name).all()
    return treatments


def _employees_for_category(
    db: Session,
    category_id: str,
    treatment_ids: set[str],
    selected_category: models.EmployeeCategory | None,
) -> list[models.Employee]:
    if not treatment_ids:
        return []
    employees_q = db.query(models.Employee).filter(models.Employee.active == True)  # noqa: E712
    if category_id:
        employees_q = employees_q.filter(models.Employee.category_id == category_id)
    employees = employees_q.order_by(models.Employee.sort_order, models.Employee.name).all()
    selected = []
    for e in employees:
        ids = set(employee_treatment_ids(db, e))
        if getattr(e, "treatment_override_enabled", False):
            if not ids or not (ids & treatment_ids):
                continue
        elif ids and not (ids & treatment_ids):
            continue
        elif not ids:
            continue
        selected.append(e)
    return selected


def employee_can_perform(db: Session, e: models.Employee, t: models.Treatment) -> bool:
    if not e or not t or not getattr(t, "active", False):
        return False
    # 기록필요 항목은 과(category) 기준으로만 판정 — 기록 입력/집계와 동일 규칙.
    # (개별 EmployeeTreatment 권한과 무관하게 같은 과 직원이면 정산 반영 가능)
    if getattr(t, "requires_record", False):
        if t.category_id:
            return e.category_id == t.category_id
        return True
    if getattr(e, "treatment_override_enabled", False):
        if e.category_id and t.category_id != e.category_id:
            return False
        return t.id in set(employee_treatment_ids(db, e))
    if e.category_id:
        return t.category_id == e.category_id
    return True


def _record_dict(rec: models.SettlementRecord) -> dict:
    qty = max(0, int(rec.quantity or 0))
    price_total = int(rec.price_snapshot or 0) * qty
    return {
        "id": rec.id,
        "performed_on": rec.performed_on,
        "employee_id": rec.employee_id,
        "treatment_id": rec.treatment_id,
        "treatment_code": rec.treatment_code or rec.treatment_code_snapshot,
        "quantity": qty,
        "memo": rec.memo or "",
        "employee_name_snapshot": rec.employee_name_snapshot or "",
        "employee_category_id_snapshot": rec.employee_category_id_snapshot or "",
        "employee_category_name_snapshot": rec.employee_category_name_snapshot or "",
        "treatment_name_snapshot": rec.treatment_name_snapshot or "",
        "treatment_short_snapshot": rec.treatment_short_snapshot or "",
        "treatment_code_snapshot": rec.treatment_code_snapshot or rec.treatment_code or "",
        "price_snapshot": int(rec.price_snapshot or 0),
        "price_total": price_total,
        "incentive_type_snapshot": rec.incentive_type_snapshot or "none",
        "incentive_value_snapshot": rec.incentive_value_snapshot or 0,
        "incentive_amount": int(rec.incentive_amount or 0),
        "unit_incentive": int((rec.incentive_amount or 0) / qty) if qty else 0,
        "created_at": rec.created_at.isoformat() if rec.created_at else None,
        "updated_at": rec.updated_at.isoformat() if rec.updated_at else None,
    }


def _summary(records: list[models.SettlementRecord]) -> dict:
    by_employee = {}
    by_treatment = {}
    by_category = {}
    quantity_total = 0
    price_total = 0
    incentive_total = 0

    for rec in records:
        qty = max(0, int(rec.quantity or 0))
        line_price = int(rec.price_snapshot or 0) * qty
        inc = int(rec.incentive_amount or 0)
        quantity_total += qty
        price_total += line_price
        incentive_total += inc

        emp = by_employee.setdefault(
            rec.employee_id,
            {
                "employee_id": rec.employee_id,
                "employee_name": rec.employee_name_snapshot or "",
                "category_id": rec.employee_category_id_snapshot or "",
                "category_name": rec.employee_category_name_snapshot or "",
                "quantity_total": 0,
                "price_total": 0,
                "incentive_total": 0,
                "adjustment_total": 0,
                "payment_total": 0,
            },
        )
        emp["quantity_total"] += qty
        emp["price_total"] += line_price
        emp["incentive_total"] += inc

        tx_key = rec.treatment_id
        tx = by_treatment.setdefault(
            tx_key,
            {
                "treatment_id": rec.treatment_id,
                "treatment_code": rec.treatment_code_snapshot or rec.treatment_code or "",
                "treatment_name": rec.treatment_name_snapshot or "",
                "treatment_short": rec.treatment_short_snapshot or "",
                "quantity_total": 0,
                "price_total": 0,
                "incentive_total": 0,
            },
        )
        tx["quantity_total"] += qty
        tx["price_total"] += line_price
        tx["incentive_total"] += inc

        cat_key = rec.employee_category_id_snapshot or ""
        cat = by_category.setdefault(
            cat_key,
            {
                "category_id": cat_key,
                "category_name": rec.employee_category_name_snapshot or "",
                "quantity_total": 0,
                "price_total": 0,
                "incentive_total": 0,
            },
        )
        cat["quantity_total"] += qty
        cat["price_total"] += line_price
        cat["incentive_total"] += inc

    for emp in by_employee.values():
        emp["payment_total"] = int(emp["incentive_total"] or 0) + int(emp["adjustment_total"] or 0)

    return {
        "record_count": len(records),
        "quantity_total": quantity_total,
        "price_total": price_total,
        "incentive_total": incentive_total,
        "adjustment_total": 0,
        "payment_total": incentive_total,
        "by_employee": sorted(by_employee.values(), key=lambda x: x["employee_name"]),
        "by_treatment": sorted(by_treatment.values(), key=lambda x: x["treatment_name"]),
        "by_category": sorted(by_category.values(), key=lambda x: x["category_name"]),
    }


def get_grid(db: Session, date_from: str, date_to: str, category_id: str = "") -> dict:
    start, end, date_keys, range_label = resolve_range(date_from, date_to)
    categories = _active_categories(db)
    category_id = _choose_category_id(db, category_id, categories)
    selected_category = next((c for c in categories if c.id == category_id), None)
    treatments = _treatments_for_category(db, category_id, selected_category)
    treatment_ids = {t.id for t in treatments}
    employees = _employees_for_category(db, category_id, treatment_ids, selected_category)
    employee_ids = {e.id for e in employees}

    records = (
        db.query(models.SettlementRecord)
        .filter(
            models.SettlementRecord.performed_on >= start.isoformat(),
            models.SettlementRecord.performed_on <= end.isoformat(),
        )
        .all()
    )
    records = [
        rec for rec in records
        if rec.treatment_id in treatment_ids and rec.employee_id in employee_ids
    ]
    record_map = {
        (rec.performed_on, rec.employee_id, rec.treatment_id): rec
        for rec in records
    }

    items = []
    for day in date_keys:
        employee_data = {}
        for e in employees:
            cell = {"counts": {}, "incentives": {}, "record_ids": {}, "memos": {}}
            for t in treatments:
                rec = record_map.get((day, e.id, t.id))
                cell["counts"][t.id] = int(rec.quantity or 0) if rec else 0
                cell["incentives"][t.id] = int(rec.incentive_amount or 0) if rec else 0
                cell["record_ids"][t.id] = rec.id if rec else None
                cell["memos"][t.id] = rec.memo if rec else ""
            employee_data[e.id] = cell
        items.append({"date": day, "employee_data": employee_data})

    return {
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "range_label": range_label,
        "categories": [_serialize_category(c) for c in categories],
        "category_id": category_id,
        "treatments": [_serialize_treatment(t) for t in treatments],
        "employees": [_serialize_employee(db, e) for e in employees],
        "items": items,
        "records": [_record_dict(rec) for rec in sorted(records, key=lambda r: (r.performed_on, r.employee_name_snapshot, r.treatment_name_snapshot))],
        "summary": _summary(records),
    }


def _apply_new_snapshot(rec: models.SettlementRecord, e: models.Employee, t: models.Treatment):
    category = getattr(e, "category", None)
    rule, value = incentive_snapshot_for_treatment(t)
    rec.employee_name_snapshot = e.name or ""
    rec.employee_category_id_snapshot = e.category_id
    rec.employee_category_name_snapshot = category.name if category else ""
    rec.treatment_name_snapshot = t.name or ""
    rec.treatment_short_snapshot = t.short or ""
    rec.treatment_code_snapshot = t.code or ""
    rec.treatment_code = t.code or ""
    rec.price_snapshot = int(getattr(t, "price", 0) or 0)
    rec.incentive_type_snapshot = rule
    rec.incentive_value_snapshot = value


def _recalculate_record(rec: models.SettlementRecord):
    rec.incentive_amount = calculate_incentive_amount(
        rec.price_snapshot or 0,
        rec.incentive_type_snapshot or "none",
        rec.incentive_value_snapshot or 0,
        rec.quantity,
    )


# v1.3.37+: 집계가 실시간 계산(치료완료/기록 자동 + manual_counts 델타)으로 바뀌어
# 정산→manual_counts 역미러는 제거됨. 남기면 manual_counts 가 '수동 보정 델타' 의미를
# 잃고 자동 집계값과 이중으로 더해진다. 정산은 집계의 단방향 다운스트림 스냅샷이다.


def upsert_grid(
    db: Session,
    payload: SettlementGridIn,
    log_callback: Callable | None = None,
    audit_callback: Callable | None = None,
) -> dict:
    start, end, _, _ = resolve_range(payload.date_from, payload.date_to)
    category_id = (payload.category_id or "").strip()
    changed = {"upserted": 0, "deleted": 0}
    seen = set()

    for entry in payload.entries:
        performed_on = parse_date(entry.performed_on, "performed_on")
        if performed_on < start or performed_on > end:
            raise ValueError("performed_on is outside the requested range")
        try:
            quantity = int(entry.quantity or 0)
        except Exception as exc:
            raise ValueError("quantity must be zero or a positive integer") from exc
        if quantity < 0:
            raise ValueError("quantity must be zero or a positive integer")

        key = (performed_on.isoformat(), entry.employee_id, entry.treatment_id)
        if key in seen:
            continue
        seen.add(key)

        employee = db.get(models.Employee, entry.employee_id)
        if not employee or not getattr(employee, "active", False):
            raise ValueError("employee not found or inactive")
        treatment = db.get(models.Treatment, entry.treatment_id)
        if not treatment or not getattr(treatment, "active", False):
            raise ValueError("treatment not found or inactive")
        if category_id and treatment.category_id != category_id:
            raise ValueError("treatment does not belong to the selected category")
        if not employee_can_perform(db, employee, treatment):
            raise ValueError("employee is not assigned to the treatment")

        existing = (
            db.query(models.SettlementRecord)
            .filter(
                models.SettlementRecord.performed_on == performed_on.isoformat(),
                models.SettlementRecord.employee_id == employee.id,
                models.SettlementRecord.treatment_id == treatment.id,
            )
            .first()
        )

        if quantity == 0:
            if existing:
                rec_id = existing.id
                if log_callback:
                    log_callback(db, "settlement_record", rec_id, "delete", None)
                if audit_callback:
                    audit_callback(db, "settlement_record.delete", rec_id,
                                   f"date={performed_on.isoformat()} employee={employee.name} treatment={treatment.name}")
                db.delete(existing)
                changed["deleted"] += 1
            continue

        if existing:
            rec = existing
            rec.quantity = quantity
            rec.memo = entry.memo or ""
            _apply_new_snapshot(rec, employee, treatment)
        else:
            rec = models.SettlementRecord(
                performed_on=performed_on.isoformat(),
                employee_id=employee.id,
                treatment_id=treatment.id,
                quantity=quantity,
                memo=entry.memo or "",
            )
            _apply_new_snapshot(rec, employee, treatment)
            db.add(rec)

        _recalculate_record(rec)
        rec.updated_at = datetime.utcnow()
        db.flush()
        if log_callback:
            log_callback(db, "settlement_record", rec.id, "upsert", rec)
        if audit_callback:
            audit_callback(db, "settlement_record.upsert", rec.id,
                           f"date={rec.performed_on} employee={rec.employee_name_snapshot} treatment={rec.treatment_name_snapshot} quantity={rec.quantity}")
        changed["upserted"] += 1

    return changed


def report_incentives(db: Session, date_from: str, date_to: str, category_id: str = "") -> dict:
    start, end, _, range_label = resolve_range(date_from, date_to)
    categories = _active_categories(db)
    category_id = _choose_category_id(db, category_id, categories) if categories else ""
    q = db.query(models.SettlementRecord).filter(
        models.SettlementRecord.performed_on >= start.isoformat(),
        models.SettlementRecord.performed_on <= end.isoformat(),
    )
    if category_id:
        q = q.filter(models.SettlementRecord.employee_category_id_snapshot == category_id)
    records = q.order_by(
        models.SettlementRecord.performed_on,
        models.SettlementRecord.employee_name_snapshot,
        models.SettlementRecord.treatment_name_snapshot,
    ).all()
    return {
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "range_label": range_label,
        "category_id": category_id or "",
        "categories": [_serialize_category(c) for c in categories],
        "records": [_record_dict(rec) for rec in records],
        "summary": _summary(records),
    }


def _incentive_rule_label(rule: str, value) -> str:
    rule = (rule or "none").strip()
    if rule == "fixed":
        return f"고정 {int(value or 0):,}원"
    if rule == "percent":
        return f"{float(value or 0):g}%"
    return "없음"


def _safe_filename_part(value: str) -> str:
    text = (value or "전체").strip() or "전체"
    return re.sub(r'[\\/:*?"<>|]+', "_", text)


def build_incentive_workbook(
    db: Session,
    date_from: str,
    date_to: str,
    category_id: str = "",
) -> tuple[BytesIO, str]:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    report = report_incentives(db, date_from, date_to, category_id)
    summary = report["summary"]
    records = report["records"]
    categories = report.get("categories") or []
    selected_category_id = report.get("category_id") or ""
    selected_category = next((c for c in categories if c.get("id") == selected_category_id), None)
    category_name = (selected_category or {}).get("name") or "전체"

    THIN = Side(style="thin", color="CBD5E1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
    TOTAL_FILL = PatternFill("solid", fgColor="E2E8F0")

    F_TITLE = Font(name="맑은 고딕", size=15, bold=True, color="1E293B")
    F_SUB = Font(name="맑은 고딕", size=10, color="64748B")
    F_HEADER = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    F_BODY = Font(name="맑은 고딕", size=10, color="1E293B")
    F_TOTAL = Font(name="맑은 고딕", size=10, bold=True, color="0F172A")

    A_LEFT = Alignment(horizontal="left", vertical="center")
    A_CENTER = Alignment(horizontal="center", vertical="center")
    A_RIGHT = Alignment(horizontal="right", vertical="center")

    def setup_sheet(ws):
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.45, bottom=0.45)

    def write_title(ws, title: str):
        ws.cell(row=1, column=1, value=title).font = F_TITLE
        ws.cell(row=2, column=1, value=f"{report['date_from']} ~ {report['date_to']} / {category_name}").font = F_SUB
        ws.cell(row=2, column=1).alignment = A_LEFT

    def header_row(ws, row: int, headers: list[str]):
        for idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=idx, value=label)
            cell.font = F_HEADER
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = A_CENTER

    def body_cell(cell, *, alt=False, total=False, align=None):
        cell.font = F_TOTAL if total else F_BODY
        cell.border = BORDER
        if total:
            cell.fill = TOTAL_FILL
        elif alt:
            cell.fill = ALT_FILL
        if align:
            cell.alignment = align

    def money_cell(cell, value, *, alt=False, total=False):
        cell.value = int(value or 0)
        cell.number_format = '"₩"#,##0'
        body_cell(cell, alt=alt, total=total, align=A_RIGHT)

    def count_cell(cell, value, *, alt=False, total=False):
        cell.value = int(value or 0)
        cell.number_format = '#,##0'
        body_cell(cell, alt=alt, total=total, align=A_RIGHT)

    def autosize(ws, widths: dict[int, int] | None = None):
        widths = widths or {}
        for col_idx in range(1, ws.max_column + 1):
            width = widths.get(col_idx)
            if width is None:
                width = min(28, max(10, max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, ws.max_row + 1)
                ) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb = openpyxl.Workbook()
    ws_emp = wb.active
    ws_emp.title = "직원별 정산 요약"
    ws_detail = wb.create_sheet("상세 내역")
    ws_tx = wb.create_sheet("치료항목별 합계")

    # 직원별 정산 요약
    setup_sheet(ws_emp)
    write_title(ws_emp, "직원별 정산 요약")
    headers = ["직원", "과", "총 건수", "총 수가", "세전 인센티브", "조정금액", "최종 지급액"]
    header_row(ws_emp, 4, headers)
    row = 5
    for idx, emp in enumerate(summary.get("by_employee") or []):
        alt = idx % 2 == 1
        values = [emp["employee_name"], emp["category_name"]]
        for col, value in enumerate(values, start=1):
            cell = ws_emp.cell(row=row, column=col, value=value)
            body_cell(cell, alt=alt, align=A_LEFT)
        count_cell(ws_emp.cell(row=row, column=3), emp.get("quantity_total"), alt=alt)
        money_cell(ws_emp.cell(row=row, column=4), emp.get("price_total"), alt=alt)
        money_cell(ws_emp.cell(row=row, column=5), emp.get("incentive_total"), alt=alt)
        money_cell(ws_emp.cell(row=row, column=6), emp.get("adjustment_total"), alt=alt)
        money_cell(ws_emp.cell(row=row, column=7), emp.get("payment_total"), alt=alt)
        row += 1
    ws_emp.cell(row=row, column=1, value="합계")
    body_cell(ws_emp.cell(row=row, column=1), total=True, align=A_CENTER)
    body_cell(ws_emp.cell(row=row, column=2), total=True, align=A_CENTER)
    count_cell(ws_emp.cell(row=row, column=3), summary.get("quantity_total"), total=True)
    money_cell(ws_emp.cell(row=row, column=4), summary.get("price_total"), total=True)
    money_cell(ws_emp.cell(row=row, column=5), summary.get("incentive_total"), total=True)
    money_cell(ws_emp.cell(row=row, column=6), summary.get("adjustment_total"), total=True)
    money_cell(ws_emp.cell(row=row, column=7), summary.get("payment_total"), total=True)
    autosize(ws_emp, {1: 16, 2: 16, 3: 10, 4: 14, 5: 16, 6: 12, 7: 14})

    # 상세 내역: 사용자 요청에 따라 코드 컬럼은 제외한다.
    setup_sheet(ws_detail)
    write_title(ws_detail, "상세 내역")
    detail_headers = ["날짜", "과", "직원", "치료항목", "수량", "수가", "금액", "인센티브 규칙", "세전 인센티브", "메모"]
    header_row(ws_detail, 4, detail_headers)
    row = 5
    for idx, rec in enumerate(records):
        alt = idx % 2 == 1
        text_values = [
            rec["performed_on"],
            rec["employee_category_name_snapshot"],
            rec["employee_name_snapshot"],
            rec["treatment_name_snapshot"],
        ]
        for col, value in enumerate(text_values, start=1):
            cell = ws_detail.cell(row=row, column=col, value=value)
            body_cell(cell, alt=alt, align=A_LEFT if col != 1 else A_CENTER)
        count_cell(ws_detail.cell(row=row, column=5), rec.get("quantity"), alt=alt)
        money_cell(ws_detail.cell(row=row, column=6), rec.get("price_snapshot"), alt=alt)
        money_cell(ws_detail.cell(row=row, column=7), rec.get("price_total"), alt=alt)
        cell = ws_detail.cell(
            row=row,
            column=8,
            value=_incentive_rule_label(rec.get("incentive_type_snapshot"), rec.get("incentive_value_snapshot")),
        )
        body_cell(cell, alt=alt, align=A_LEFT)
        money_cell(ws_detail.cell(row=row, column=9), rec.get("incentive_amount"), alt=alt)
        cell = ws_detail.cell(row=row, column=10, value=rec.get("memo") or "")
        body_cell(cell, alt=alt, align=A_LEFT)
        row += 1
    autosize(ws_detail, {1: 12, 2: 14, 3: 14, 4: 22, 5: 8, 6: 12, 7: 12, 8: 14, 9: 16, 10: 20})

    # 치료항목별 합계
    setup_sheet(ws_tx)
    write_title(ws_tx, "치료항목별 합계")
    tx_headers = ["치료항목", "총 건수", "총 수가", "세전 인센티브"]
    header_row(ws_tx, 4, tx_headers)
    row = 5
    for idx, tx in enumerate(summary.get("by_treatment") or []):
        alt = idx % 2 == 1
        cell = ws_tx.cell(row=row, column=1, value=tx.get("treatment_name") or tx.get("treatment_short") or "")
        body_cell(cell, alt=alt, align=A_LEFT)
        count_cell(ws_tx.cell(row=row, column=2), tx.get("quantity_total"), alt=alt)
        money_cell(ws_tx.cell(row=row, column=3), tx.get("price_total"), alt=alt)
        money_cell(ws_tx.cell(row=row, column=4), tx.get("incentive_total"), alt=alt)
        row += 1
    ws_tx.cell(row=row, column=1, value="합계")
    body_cell(ws_tx.cell(row=row, column=1), total=True, align=A_CENTER)
    count_cell(ws_tx.cell(row=row, column=2), summary.get("quantity_total"), total=True)
    money_cell(ws_tx.cell(row=row, column=3), summary.get("price_total"), total=True)
    money_cell(ws_tx.cell(row=row, column=4), summary.get("incentive_total"), total=True)
    autosize(ws_tx, {1: 24, 2: 10, 3: 14, 4: 16})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = (
        f"정산_{report['date_from']}_{report['date_to']}_"
        f"{_safe_filename_part(category_name)}.xlsx"
    )
    return buf, filename
