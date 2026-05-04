"""REST API - 2단계 재작성

새 모델 기준:
- Employee (의사/치료사 통합, role 필드)
- EmployeeLeave (rename from therapist_leaves)
- Patient (항목별 처방/완료 횟수)
- Appointment (treatment_codes JSON 배열, 원무과 승인 1단계만)
- TreatmentAssignment (체외충격파/주사/연골주사 담당 추적)
- SystemSetting (관리자 시스템탭 설정)

프론트 호환을 위해 /api/therapists, /api/therapist-leaves 등 기존 엔드포인트는
alias로 유지 (내부적으로 Employee 참조).
"""
import shutil, json
from datetime import datetime, timedelta, date
from typing import List, Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Request, UploadFile, File, Query, Header
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..database import get_db
from ..config import load_config, save_config, get_db_path, get_backup_dir, APP_VERSION, APP_BUILD_DATE
from ..models import models, schemas, constants as C
from ..services.sync import record_op
from ..services import auth

router = APIRouter(prefix="/api")


# ──────────────── 공통 유틸 ────────────────

def require_admin(x_admin_token: str = Header(default="")):
    if not auth.is_valid(x_admin_token):
        raise HTTPException(401, "관리자 인증이 필요합니다.")
    return True


def require_admin_or_sync_token(
    x_admin_token: str = Header(default=""),
    x_sync_token: str = Header(default=""),
):
    """관리자 세션 토큰 또는 노드 간 동기화 토큰 둘 중 하나로 인증.

    sync pull/push 는 외부 무인 노드(다른 PC) 가 자동으로 호출하는 엔드포인트라
    관리자 세션 토큰만으로는 인증할 수 없음. 그러나 인증을 풀면 같은 네트워크에서
    누구든 SyncOp 큐를 노출/주입할 수 있어 위험.

    해결: config 의 sync_secret 과 X-Sync-Token 헤더를 비교 (constant-time).
    페어링된 노드끼리는 같은 sync_secret 을 공유해야 함.
    """
    if auth.is_valid(x_admin_token):
        return True
    cfg = load_config()
    secret = (cfg.get("sync_secret") or "").strip()
    if secret and x_sync_token:
        import hmac
        if hmac.compare_digest(x_sync_token, secret):
            return True
    raise HTTPException(401, "관리자 인증 또는 X-Sync-Token 헤더가 필요합니다.")


def _lunch_window():
    """현재 config 의 점심시간을 (start_min, end_min, start_str, end_str) 로 반환.
    lunch_enabled=False 이거나 형식이 이상하거나 end<=start 면 None (=차단 비활성).
    """
    cfg = load_config()
    if not cfg.get("lunch_enabled"):
        return None
    try:
        ls_str = (cfg.get("lunch_start") or "").strip()
        le_str = (cfg.get("lunch_end") or "").strip()
        sh, sm = ls_str.split(":")
        eh, em = le_str.split(":")
        s = int(sh) * 60 + int(sm)
        e = int(eh) * 60 + int(em)
    except Exception:
        return None
    if not (0 <= s < 24 * 60 and 0 <= e <= 24 * 60):
        return None
    if e <= s:
        return None
    return (s, e, ls_str, le_str)


def _check_lunch_block(start_at: datetime, duration_min):
    """예약 시간창 [start_at, start_at+duration) 가 점심창과 겹치면 400 차단.
    호출 사이트: 신규 예약, 시간/길이 수정, split-code (start_at 변경 또는 새 예약 생성).
    """
    win = _lunch_window()
    if not win:
        return
    try:
        dur = int(duration_min or 0)
    except Exception:
        return
    if dur <= 0:
        return
    s_min, e_min, ls_str, le_str = win
    sm = start_at.hour * 60 + start_at.minute
    em = sm + dur
    if em > s_min and sm < e_min:
        raise HTTPException(
            400,
            f"점심시간({ls_str}~{le_str})에는 예약을 잡을 수 없습니다.",
        )


def audit(db, action: str, entity_id: str = "", detail: str = "", actor: str = "system"):
    cfg = load_config()
    db.add(models.AuditLog(
        node_id=cfg.get("node_id"), actor=actor,
        action=action, entity_id=entity_id, detail=detail[:500],
    ))


def _log(db, entity, eid, op, obj):
    payload = {}
    if obj is not None:
        for c in obj.__table__.columns:
            v = getattr(obj, c.name)
            if isinstance(v, datetime):
                v = v.isoformat()
            payload[c.name] = v
    record_op(db, entity, eid, op, payload)


def _parse_codes(codes_json: str) -> list:
    try:
        v = json.loads(codes_json or "[]")
        return v if isinstance(v, list) else []
    except Exception:
        return []


# ──────────────── 치료항목 동적 헬퍼 (DB 기반) ────────────────

def _all_treatments(db) -> list:
    """모든 치료항목 (활성+비활성) 캐시 없이 매 호출마다 조회 — 단순함 우선."""
    return db.query(models.Treatment).all()


def _treatments_by_code(db) -> dict:
    return {t.code: t for t in _all_treatments(db)}


def _existing_codes_set(db) -> set:
    """현재 DB 에 존재하는 모든 치료항목 코드 (활성+비활성)."""
    return {t.code for t in _all_treatments(db)}


def _doctor_codes_set(db) -> set:
    """역할=doctor 인 치료항목 코드 집합."""
    return {t.code for t in _all_treatments(db) if t.role == "doctor"}


def _therapist_codes_set(db) -> set:
    """역할=therapist 인 치료항목 코드 집합."""
    return {t.code for t in _all_treatments(db) if t.role == "therapist"}


def _therapist_only_codes_set(db) -> set:
    """치료사 역할이지만 체외충격파 제외 — 도수치료/기타 (담당 치료사 필드로 관리)."""
    return {t.code for t in _all_treatments(db)
            if t.role == "therapist" and t.code != C.ESWT_CODE}


def _serialize_employee(e: models.Employee) -> dict:
    return {
        "id": e.id, "name": e.name, "role": e.role, "color": e.color,
        "active": bool(e.active), "birth_date": e.birth_date, "phone": e.phone,
        "hire_date": e.hire_date,
        "can_eswt": bool(e.can_eswt), "can_manual": bool(e.can_manual),
        "sort_order": e.sort_order or 0,
    }


def _serialize_assignment(a: models.TreatmentAssignment) -> dict:
    return {
        "treatment_code": a.treatment_code,
        "handler_id": a.handler_id,
    }


def _serialize_appointment(a: models.Appointment) -> dict:
    codes = _parse_codes(a.treatment_codes)
    color = a.therapist.color if a.therapist else "#9CA3AF"
    opacity = {"reserved": 1.0, "approved": 1.0, "canceled": 0.3}.get(a.status, 1.0)
    # 대량 데이터 대응: 프론트가 예약표 렌더에 필요한 환자 기본정보를 embed.
    #   이전엔 PATIENTS 전체(80K+) 를 preload 해서 id→name 을 조회했으나,
    #   이제 appointment 응답 자체에 담아 전역 preload 를 제거.
    pt = a.patient
    return {
        "id": a.id,
        "start": a.start_at.isoformat(),
        "end": a.end_at.isoformat(),
        "color": color, "textColor": "#fff",
       "extendedProps": {
            "patient_id": a.patient_id,
            # 환자 기본정보 embed (예약표 + 각종 모달에서 사용)
            "patient_name": (pt.name if pt else ""),
            "patient_chart_no": (pt.chart_no if pt else ""),
            "patient_phone": (pt.phone if pt else ""),
            "patient_birth_date": (pt.birth_date if pt else ""),
            "patient_memo": ((pt.memo or "") if pt else ""),
            "therapist_id": a.therapist_id,
            "treatment_codes": codes,
            "status": a.status,
            "memo": a.memo or "",
            "approved_at": a.approved_at.isoformat() if a.approved_at else None,
            "approved_by": a.approved_by,
            "opacity": opacity,
            "duration_min": a.duration_min,
            "assignments": [_serialize_assignment(x) for x in a.assignments],
            "is_new_patient": bool(a.is_new_patient),
            "version": int(a.version or 0),
            # 20-3-1 (post-19-P / F-10): 노쇼 별도 필드 — status="canceled" 와 동시 가능
            "no_show": bool(a.no_show),
        },
    }


# ──────────────── 관리자 인증 ────────────────

@router.get("/admin/status")
def admin_status(x_admin_token: str = Header(default="")):
    return {
        "authenticated": auth.is_valid(x_admin_token),
        "is_default_password": auth.is_default_password(),
    }


@router.post("/admin/login")
def admin_login(payload: dict):
    pw = (payload or {}).get("password", "")
    rem = auth.get_lock_remaining()
    if rem > 0:
        raise HTTPException(429, f"로그인이 잠겼습니다. {rem}초 후 다시 시도하세요.")
    token = auth.login(pw)
    if not token:
        rem2 = auth.get_lock_remaining()
        if rem2 > 0:
            raise HTTPException(429, f"5회 연속 실패. {rem2}초 동안 잠금됩니다.")
        raise HTTPException(401, "비밀번호가 올바르지 않습니다.")
    return {"token": token, "is_default_password": auth.is_default_password()}


@router.post("/admin/logout")
def admin_logout(x_admin_token: str = Header(default="")):
    auth.logout(x_admin_token)
    return {"ok": True}


@router.post("/admin/change-password")
def admin_change_password(payload: dict, db: Session = Depends(get_db),
                          _: bool = Depends(require_admin)):
    cur = (payload or {}).get("current_password", "")
    new = (payload or {}).get("new_password", "")
    if not auth.verify_password(cur, auth.get_admin_hash()):
        raise HTTPException(401, "현재 비밀번호가 올바르지 않습니다.")
    try:
        auth.set_admin_password(new)
    except ValueError as e:
        raise HTTPException(400, str(e))
    audit(db, "admin.password_change", "", "관리자 비밀번호 변경")
    db.commit()
    return {"ok": True, "msg": "비밀번호가 변경되었습니다. 다시 로그인하세요."}


# ──────────────── 앱 정보 / 업데이트 ────────────────

@router.get("/about")
def about():
    """앱 버전 / 데이터 경로 / 업데이트 매니페스트 URL 노출.
    업데이트 확인 UI가 참조하는 기본 정보.
    """
    cfg = load_config()
    return {
        "app_name": "도수치료예약",
        "version": APP_VERSION,
        "build_date": APP_BUILD_DATE,
        "data_dir": str(get_db_path().parent),
        "db_path": str(get_db_path()),
        "backup_dir": str(get_backup_dir()),
        "update_manifest_url": cfg.get("update_manifest_url", ""),
        # 프론트가 dev/uvicorn 환경 여부를 알 수 있게 노출.
        # False 면 자동 업데이트(다운로드/설치) 가 _is_frozen() 가드로 차단됨 →
        # UI 가 사전에 안내 배너 + 버튼 비활성화로 헛클릭을 막음.
        "is_frozen": _is_frozen(),
    }


@router.post("/about/check-update")
def check_update(_: bool = Depends(require_admin)):
    """원격 manifest URL 이 설정되어 있으면 최신 버전 정보를 가져와 비교.
    manifest 형식:
        {"version":"1.2.3","download_url":"...","notes":"...","mandatory":false}

    반환:
      - configured:False → URL 미설정 (UI에 수동 업데이트 안내)
      - up_to_date:True  → 최신
      - available:True   → 새 버전 있음 + 다운로드/릴리즈노트
    """
    import urllib.request, urllib.error, json as _json
    cfg = load_config()
    url = (cfg.get("update_manifest_url") or "").strip()
    base = {
        "current_version": APP_VERSION,
        "checked_at": datetime.utcnow().isoformat() + "Z",
    }
    if not url:
        return {**base, "configured": False,
                "message": "업데이트 매니페스트 URL이 설정되지 않았습니다. 시스템 설정에서 URL을 입력하거나, 배포처 안내에 따라 수동으로 교체하세요."}
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "dosu-clinic-updater/1.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            raw = r.read().decode("utf-8", errors="ignore")
        data = _json.loads(raw)
    except (urllib.error.URLError, urllib.error.HTTPError, ValueError, TimeoutError) as e:
        return {**base, "configured": True, "error": f"매니페스트 조회 실패: {e}"}
    except Exception as e:
        return {**base, "configured": True, "error": f"매니페스트 파싱 실패: {e}"}

    remote_ver = str(data.get("version") or "").strip()
    if not remote_ver:
        return {**base, "configured": True, "error": "매니페스트에 version 필드가 없습니다."}

    def _vt(s):
        try: return tuple(int(x) for x in s.split("."))
        except Exception: return (0,)

    up_to_date = _vt(remote_ver) <= _vt(APP_VERSION)
    return {
        **base,
        "configured": True,
        "latest_version": remote_ver,
        "up_to_date": up_to_date,
        "available": (not up_to_date),
        "download_url": data.get("download_url", ""),
        "sha256": data.get("sha256", ""),          # 검증용 (optional)
        "notes": data.get("notes", ""),
        "mandatory": bool(data.get("mandatory", False)),
    }


# ──────────────── 업데이트 다운로드 / 설치 ────────────────
#
# 흐름:
#   1) check-update → 최신 버전 / download_url / sha256 확인
#   2) download-update → ZIP 다운로드 + SHA256 검증 → <app>/.update/new.zip
#   3) apply-update → updater.bat 실행(detached) + 서버 자기 종료
#
# ⚠ 이 기능은 PyInstaller 로 빌드된 onedir exe 환경에서만 동작.
#    개발 모드(python run.py)에서는 "frozen=False" 이므로 거부.


def _get_app_folder():
    """PyInstaller 빌드 시 프로그램 폴더(exe 위치) 반환.
    개발 모드에서는 프로젝트 루트 (업데이트 경로엔 쓰지 않음, 안내용)."""
    import sys as _sys
    from pathlib import Path as _Path
    if getattr(_sys, 'frozen', False):
        return _Path(_sys.executable).parent
    return _Path(__file__).resolve().parent.parent.parent


def _is_frozen() -> bool:
    import sys as _sys
    return bool(getattr(_sys, 'frozen', False))


@router.post("/about/download-update")
def download_update(_: bool = Depends(require_admin)):
    """매니페스트에서 받은 download_url 로 ZIP 다운로드 + SHA256 검증.

    반환:
      - ok: True / False
      - path: 저장 경로 (검증 완료 시)
      - size_mb: 다운로드 크기
      - sha256_matched: 서버가 sha256 제공했을 경우 검증 결과
      - error: 실패 사유
    """
    import urllib.request, urllib.error, hashlib, json as _json
    from pathlib import Path

    if not _is_frozen():
        raise HTTPException(400,
            "개발 모드에서는 자동 업데이트가 지원되지 않습니다. "
            "배포본(exe)으로 실행 중일 때만 동작합니다.")

    # manifest 재조회 (check-update 와 동일 로직)
    cfg = load_config()
    url = (cfg.get("update_manifest_url") or "").strip()
    if not url:
        raise HTTPException(400, "업데이트 매니페스트 URL이 설정되지 않았습니다.")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "dosu-clinic-updater/1.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            manifest = _json.loads(r.read().decode("utf-8", errors="ignore"))
    except Exception as e:
        raise HTTPException(502, f"매니페스트 조회 실패: {e}")

    dl_url = (manifest.get("download_url") or "").strip()
    expected_sha = (manifest.get("sha256") or "").strip().lower()
    # GitHub UI 의 복사 버튼으로 복사하면 "sha256:<해시>" 형태가 돼서
    # 자주 붙는 접두사 자동 제거. 공백/탭/개행도 정리.
    for prefix in ("sha256:", "sha-256:", "sha256 :"):
        if expected_sha.startswith(prefix):
            expected_sha = expected_sha[len(prefix):].strip()
            break
    expected_sha = "".join(expected_sha.split())  # 내부 공백 제거 (혹시 모를 붙여넣기 사고)
    if not dl_url:
        raise HTTPException(400, "매니페스트에 download_url 이 없습니다.")

    app_folder = _get_app_folder()
    update_dir = app_folder / ".update"
    update_dir.mkdir(parents=True, exist_ok=True)
    zip_path = update_dir / "new.zip"
    # 기존 임시 파일 제거
    for old in update_dir.iterdir():
        try:
            if old.is_file(): old.unlink()
            elif old.is_dir(): shutil.rmtree(old, ignore_errors=True)
        except Exception: pass

    # 다운로드 (스트리밍 + sha256 누적 계산)
    try:
        req = urllib.request.Request(dl_url, headers={"User-Agent": "dosu-clinic-updater/1.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            h = hashlib.sha256()
            total = 0
            with open(zip_path, "wb") as f:
                while True:
                    chunk = resp.read(64 * 1024)
                    if not chunk: break
                    f.write(chunk)
                    h.update(chunk)
                    total += len(chunk)
            got_sha = h.hexdigest()
    except Exception as e:
        if zip_path.exists():
            try: zip_path.unlink()
            except Exception: pass
        raise HTTPException(502, f"다운로드 실패: {e}")

    # SHA256 검증 (매니페스트에 제공된 경우에만)
    sha_matched = None
    if expected_sha:
        sha_matched = (got_sha == expected_sha)
        if not sha_matched:
            try: zip_path.unlink()
            except Exception: pass
            raise HTTPException(400,
                f"SHA256 검증 실패: 예상={expected_sha[:16]}... / 실제={got_sha[:16]}..."
                " — 다운로드 파일이 손상되었거나 변조되었을 수 있습니다.")

    return {
        "ok": True,
        "path": str(zip_path),
        "size_mb": round(total / 1024 / 1024, 2),
        "sha256": got_sha,
        "sha256_expected": expected_sha,
        "sha256_matched": sha_matched,  # None = 검증 생략, True = 일치
    }


def _backup_db_before_update() -> dict:
    """업데이트 '지금 설치' 직전에 DB 스냅샷 생성.

    SQLite 공식 backup API 사용 — 본체가 DB 를 쓰고 있어도 트랜잭션 안전하게 복사.
    WAL/잠금 문제 없음.

    파일명 규칙:
        clinic_before_update_v<현재버전>_<YYYYMMDD_HHMMSS>.db
        → 백업임이 명확하고, 원래 auto-backup 파일과 구분됨.

    정책:
      - 실패해도 예외를 던지지 않음 → 업데이트를 막지 않는다 (안전망일 뿐).
      - 결과(ok/path/size/error)를 dict 로 반환해서 프론트가 표시할 수 있게 함.
    """
    import sqlite3
    from pathlib import Path
    try:
        src = Path(get_db_path())
        if not src.exists():
            return {"ok": False, "error": "DB 파일이 아직 생성되지 않았습니다."}

        backup_dir = Path(get_backup_dir())
        backup_dir.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = backup_dir / f"clinic_before_update_v{APP_VERSION}_{ts}.db"

        # SQLite 의 공식 online-backup API — 락 걱정 없이 안전.
        # with 블록은 connection 을 자동 close 하지 않으므로 명시적으로 close.
        s = sqlite3.connect(str(src))
        d = sqlite3.connect(str(dst))
        try:
            s.backup(d)
        finally:
            d.close()
            s.close()

        size_mb = round(dst.stat().st_size / 1024 / 1024, 2)
        return {"ok": True, "path": str(dst), "filename": dst.name, "size_mb": size_mb}
    except Exception as e:
        return {"ok": False, "error": f"백업 실패: {e}"}


@router.post("/about/apply-update")
def apply_update(_: bool = Depends(require_admin)):
    """updater.bat 실행(detached) + 서버 자기 종료.
    1~2초 뒤 uvicorn 을 죽여 updater 가 파일 교체할 수 있게 함.

    ⚠ 실행 직전에 DB 자동 백업 수행 (%APPDATA%/도수치료예약/backups/).
    백업은 실패해도 업데이트를 막지 않음.
    """
    import sys as _sys, os as _os, subprocess, threading
    if not _is_frozen():
        raise HTTPException(400, "개발 모드에서는 자동 업데이트가 지원되지 않습니다.")

    app_folder = _get_app_folder()
    zip_path = app_folder / ".update" / "new.zip"
    if not zip_path.exists():
        raise HTTPException(400, "업데이트 파일(.update/new.zip)이 없습니다. 먼저 다운로드하세요.")

    # updater.bat 우선순위: 루트 → _internal/ (PyInstaller 기본 datas 위치)
    updater = app_folder / "updater.bat"
    if not updater.exists():
        fallback = app_folder / "_internal" / "updater.bat"
        if fallback.exists():
            # 다음 실행부턴 루트에서 바로 찾도록 루트로 복사
            try:
                shutil.copy2(str(fallback), str(updater))
            except Exception:
                updater = fallback  # 복사 실패해도 fallback 경로로 실행
    if not updater.exists():
        raise HTTPException(500,
            "updater.bat 파일이 없습니다. 배포 패키지가 손상됐을 수 있습니다.")

    # ── 업데이트 직전 DB 자동 백업 (안전망) ──
    backup_result = _backup_db_before_update()
    # 백업 실패 시 업데이트 중단 — 백업 없는 채로 강제 진행하면 롤백 수단이 없어짐.
    # 이전 코드는 backup_result.ok 를 검사 안 해 디스크 부족/권한 오류 상황에서도 진행했음.
    if not backup_result.get("ok"):
        raise HTTPException(500,
            f"업데이트 직전 DB 백업이 실패했습니다 — 안전을 위해 업데이트를 중단합니다. "
            f"사유: {backup_result.get('error') or '알 수 없음'}. "
            f"디스크 여유 공간과 권한을 확인 후 다시 시도하세요.")

    # 부모(GUI exe, console=False)와 완전히 분리해서 실행.
    # CREATE_NEW_CONSOLE 단독은 PyInstaller console=False 환경에서
    # 부모가 os._exit(0) 으로 죽을 때 자식 cmd 도 같이 잡힐 수 있어,
    # DETACHED_PROCESS + CREATE_NEW_PROCESS_GROUP + CREATE_BREAKAWAY_FROM_JOB 으로 변경.
    # 자식 콘솔 가시성은 updater.bat 내부에서 자체 콘솔로 재실행해 확보.
    #
    # cmd 의 두 번째 단계는 /c (not /k):
    #   /k 였을 때는 updater.bat 가 정상 종료(exit /b 0) 한 뒤에도 콘솔이 그대로 남아
    #   사용자가 직접 닫아야 했음. /c 로 바꾸면 성공 경로는 자동 종료되고,
    #   실패(rollback) 경로는 updater.bat 안의 `pause` 가 사용자 입력을 기다림.
    DETACHED_PROCESS = 0x00000008
    CREATE_NEW_PROCESS_GROUP = 0x00000200
    CREATE_BREAKAWAY_FROM_JOB = 0x01000000
    flags = DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP | CREATE_BREAKAWAY_FROM_JOB
    try:
        subprocess.Popen(
            ["cmd", "/c", "start", "", "cmd", "/c", str(updater)],
            cwd=str(app_folder),
            creationflags=flags,
            close_fds=True,
        )
    except Exception as e:
        raise HTTPException(500, f"updater 실행 실패: {e}")

    # ⚠ 이전엔 proc.poll() 로 "즉시 종료 검증" 을 시도했으나 의미가 없었음:
    #   여기서 proc 는 `start` 를 호출하는 중간 cmd /c 셸이고, start 가 실제 updater
    #   윈도우를 띄운 직후 그 셸은 자기 임무를 마치고 곧바로 returncode=0 으로 종료됨.
    #   따라서 poll() 결과는 updater 의 생존 여부와 무관하다.
    #   updater 자체가 안 떴는지 확인하려면 _get_updater_log_path() 의 로그 파일 mtime 을
    #   별도로 검사해야 하는데, 이는 프론트의 30초/60초 진단 UX 가 이미 다루고 있음.

    # 본체 종료 예약 — 3초 후 종료.
    # PyInstaller _internal/ 의 .pyd / .dll 파일이 OS 에서 unlock 되기까지 시간 필요.
    # 1.5초였을 때 updater.bat 의 rename 단계에서 잠금 충돌 발생 사례 있음.
    #
    # 또한 backup/sync worker daemon thread 를 graceful 하게 정지시켜
    # 진행 중인 shutil.copy2 / sync_with_peer 가 mid-operation 에서 끊기지 않게 함.
    # (정지 못 시키면 부분 백업 파일 / 부분 적용된 sync op 가 잔존할 수 있음)
    try:
        from ..services.backup import stop_auto_backup
        stop_auto_backup()
    except Exception:
        pass
    try:
        from ..services.sync import stop_sync_worker
        stop_sync_worker(timeout=2.0)
    except Exception:
        pass

    def _delayed_exit():
        import time as _t
        _t.sleep(3.0)
        _os._exit(0)  # uvicorn / threads 전부 즉시 종료
    threading.Thread(target=_delayed_exit, daemon=True).start()

    return {
        "ok": True,
        "message": "업데이터 실행 완료. 프로그램이 곧 재시작됩니다.",
        "backup": backup_result,   # 프론트에서 "백업 파일명/크기" 안내에 사용
        "updater_log_path": _get_updater_log_path(),  # 사용자가 멈춤 시 직접 확인 가능
    }


def _get_updater_log_path() -> str:
    """updater.bat 이 기록하는 로그 파일 경로 (%TEMP%\\도수치료예약_updater.log).

    Python 의 tempfile.gettempdir() 는 Windows 의 %TEMP% 와 동일하게 결정됨.
    """
    import tempfile
    from pathlib import Path
    return str(Path(tempfile.gettempdir()) / "도수치료예약_updater.log")


@router.get("/about/update-log")
def get_update_log(_: bool = Depends(require_admin), tail: int = Query(200, ge=1, le=2000)):
    """업데이터 실행 로그 조회 — 자동 업데이트가 멈췄을 때 진단용.

    프론트의 "업데이트 진행 중" 화면에서 30~60초 이상 응답이 없을 때 호출,
    어느 단계에서 실패/지연되고 있는지 사용자가 직접 확인할 수 있게 한다.

    반환:
      - exists: 로그 파일 존재 여부
      - path: 절대경로 (사용자에게 "직접 열어보세요" 안내용)
      - lines: 마지막 N 줄 (기본 200, 최대 2000)
      - mtime: 마지막 수정 시각 (ISO)
      - size_bytes: 파일 크기
    """
    from pathlib import Path

    log_path = Path(_get_updater_log_path())
    base = {"path": str(log_path)}
    if not log_path.exists():
        return {**base, "exists": False, "lines": [], "mtime": None, "size_bytes": 0}

    try:
        size = log_path.stat().st_size
        mtime = datetime.fromtimestamp(log_path.stat().st_mtime).isoformat()
        # cp949 / utf-8 둘 다 시도 (bat 출력 인코딩 환경 의존)
        raw = log_path.read_bytes()
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("cp949", errors="replace")
        all_lines = text.splitlines()
        lines = all_lines[-tail:] if len(all_lines) > tail else all_lines
        return {
            **base,
            "exists": True,
            "lines": lines,
            "total_lines": len(all_lines),
            "mtime": mtime,
            "size_bytes": size,
        }
    except Exception as e:
        return {**base, "exists": True, "error": f"로그 읽기 실패: {e}", "lines": []}


# ──────────────── 설정 / 모드 ────────────────

@router.get("/config")
def get_config():
    """공개 가능한 config 만 반환.

    ⚠ 비밀 값은 반드시 제거: admin_password_hash, sync_secret.
       sync_secret 이 노출되면 누구나 X-Sync-Token 으로 sync pull/push 를 호출할 수 있어
       sync 인증이 무력화됨. 별도 관리자 전용 엔드포인트(/config/sync-secret)에서만 조회 가능.
    """
    cfg = dict(load_config())
    cfg.setdefault("leave_am_until", "14:00")
    cfg.setdefault("leave_pm_from", "13:00")
    cfg.setdefault("lunch_enabled", False)
    cfg.setdefault("lunch_start", "12:30")
    cfg.setdefault("lunch_end", "13:30")
    cfg.pop("admin_password_hash", None)
    cfg.pop("sync_secret", None)
    return cfg


@router.get("/config/sync-secret")
def get_sync_secret(_: bool = Depends(require_admin)):
    """관리자 전용 — peer 노드 페어링에 쓸 sync_secret 원문 조회.

    이 값을 다른 노드의 config.json `sync_secret` 에 동일하게 입력하면
    페어링이 완료되어 양 방향 sync pull/push 가 통과함.
    """
    cfg = load_config()
    return {"sync_secret": cfg.get("sync_secret") or ""}


@router.post("/config/regenerate-sync-secret")
def regenerate_sync_secret(_: bool = Depends(require_admin)):
    """관리자 전용 — sync_secret 을 새 랜덤값으로 재생성.

    재생성 후엔 페어링된 다른 노드들도 같은 값으로 갱신해줘야 sync 가 다시 통함.
    """
    import secrets as _secrets
    cfg = load_config()
    cfg["sync_secret"] = _secrets.token_urlsafe(32)
    save_config(cfg)
    return {"ok": True, "sync_secret": cfg["sync_secret"]}


@router.post("/config")
def update_config(payload: dict, db: Session = Depends(get_db), _: bool = Depends(require_admin)):
    # ⚠ 일반 config 갱신 경로로 sync_secret 을 덮어쓰지 못하게 차단.
    #    실수로 빈 값을 보내면 인증이 망가지므로 — 갱신은 전용 regenerate 엔드포인트로만.
    payload = dict(payload or {})
    payload.pop("sync_secret", None)
    payload.pop("admin_password_hash", None)
    cfg = load_config(); cfg.update(payload)
    # 점심시간 검증 (lunch_enabled=true 일 때만 형식/범위 강제)
    if cfg.get("lunch_enabled"):
        try:
            ls_str = (cfg.get("lunch_start") or "").strip()
            le_str = (cfg.get("lunch_end") or "").strip()
            sh, sm = ls_str.split(":")
            eh, em = le_str.split(":")
            s = int(sh) * 60 + int(sm)
            e = int(eh) * 60 + int(em)
        except Exception:
            raise HTTPException(400, "점심시간 형식이 올바르지 않습니다 (HH:MM 필요).")
        if not (0 <= s < 24 * 60 and 0 <= e <= 24 * 60):
            raise HTTPException(400, "점심시간이 유효한 시각 범위를 벗어났습니다.")
        if e <= s:
            raise HTTPException(400, "점심 종료 시간은 시작 시간보다 뒤여야 합니다.")
    save_config(cfg)
    audit(db, "config.update", "", str(payload)); db.commit()
    # 응답에서 비밀 값 제거 — admin 이라도 일반 config 응답에 secret echo 안 시킴.
    out = dict(cfg)
    out.pop("sync_secret", None)
    out.pop("admin_password_hash", None)
    return out


@router.post("/mode")
def set_mode(p: schemas.ModeSelect, x_admin_token: str = Header(default="")):
    """노드 모드(main/sub) 변경.

    인증 정책: 항상 관리자 인증 필요.
    이전 코드는 cfg.get("mode") 가 None 일 때(첫 실행) 인증 없이 변경 허용했지만,
    이는 외부에서 누구나 main↔sub 전환을 트리거할 수 있어 위험.
    """
    if not auth.is_valid(x_admin_token):
        raise HTTPException(401, "모드 변경은 관리자 인증이 필요합니다.")
    cfg = load_config()
    if p.mode not in ("main", "sub"):
        raise HTTPException(400, "mode must be 'main' or 'sub'")
    cfg["mode"] = p.mode
    if p.mode == "sub":
        if not p.main_url:
            raise HTTPException(400, "main_url required for sub")
        cfg["main_url"] = p.main_url.rstrip("/")
    save_config(cfg); return cfg


# ──────────────── 치료항목 메타 (DB 기반) ────────────────

def _serialize_treatment(t: models.Treatment) -> dict:
    return {
        "id": t.id,
        "code": t.code,
        "name": t.name,
        "short": t.short,
        "default_minutes": t.default_minutes,
        "role": t.role,
        "count_increment": t.count_increment,
        "show_in_patient": t.show_in_patient,
        "active": t.active,
        "sort_order": t.sort_order,
        # 수가/인센티브 (v1.2.3+). 기존 DB 에는 NULL 이거나 0 이어서 안전.
        "price": int(getattr(t, "price", 0) or 0),
        "incentive_pct": getattr(t, "incentive_pct", None),
        "incentive_amount": getattr(t, "incentive_amount", None),
    }


def _normalize_incentive(pct, amount):
    """인센티브 XOR 정규화 + 검증.
    - 둘 다 값이 있으면 400
    - pct 는 0 ~ 100, amount 는 0 이상
    - 빈값/None/0 이하는 NULL 로 통일 (DB 에 "입력 안 함" 의미)
    반환: (pct_or_none, amount_or_none)
    """
    def _to_float(v):
        try:
            if v is None or v == "": return None
            f = float(v)
            return f if f > 0 else None
        except Exception:
            return None
    def _to_int(v):
        try:
            if v is None or v == "": return None
            i = int(v)
            return i if i > 0 else None
        except Exception:
            return None
    p = _to_float(pct)
    a = _to_int(amount)
    if p is not None and a is not None:
        raise HTTPException(400, "인센티브는 '퍼센티지' 또는 '고정 금액' 중 하나만 입력하세요.")
    if p is not None and (p < 0 or p > 100):
        raise HTTPException(400, "인센티브 퍼센티지는 0~100 사이여야 합니다.")
    return p, a


def _build_treatment_meta(db) -> dict:
    """프론트가 사용할 치료항목 메타 (DB 기반)."""
    treatments = db.query(models.Treatment).order_by(models.Treatment.sort_order).all()

    treatment_codes = [t.code for t in treatments if t.active]
    treatment_names = {t.code: t.name for t in treatments}
    treatment_short = {t.code: t.short for t in treatments}
    doctor_treatments = [t.code for t in treatments if t.active and t.role == "doctor"]
    therapist_treatments = [t.code for t in treatments if t.active and t.role == "therapist"]
    # 도수치료 = 체외충격파가 아닌 치료사 항목 (B-2 한도 계산용)
    manual_treatments = [t.code for t in treatments
                         if t.active and t.role == "therapist" and t.code != C.ESWT_CODE]
    treatment_minutes = {t.code: t.default_minutes for t in treatments}
    count_increment = {t.code: t.count_increment for t in treatments}
    treatment_role = {t.code: t.role for t in treatments}
    treatment_show = {t.code: t.show_in_patient for t in treatments}
    # 수가/인센티브 맵 — 프론트에서 매출·인센티브 계산 시 사용
    treatment_price = {t.code: int(getattr(t, "price", 0) or 0) for t in treatments}
    treatment_incentive_pct = {t.code: getattr(t, "incentive_pct", None) for t in treatments}
    treatment_incentive_amount = {t.code: getattr(t, "incentive_amount", None) for t in treatments}

    return {
        "treatment_codes": treatment_codes,
        "treatment_names": treatment_names,
        "treatment_short": treatment_short,
        "treatment_minutes": treatment_minutes,
        "treatment_role": treatment_role,
        "treatment_show": treatment_show,
        "doctor_treatments": doctor_treatments,
        "therapist_treatments": therapist_treatments,
        "manual_treatments": manual_treatments,
        "count_increment": count_increment,
        "eswt_code": C.ESWT_CODE,
        # 수가/인센티브 (v1.2.3+)
        "treatment_price": treatment_price,
        "treatment_incentive_pct": treatment_incentive_pct,
        "treatment_incentive_amount": treatment_incentive_amount,
        # 모든 치료항목 (활성 + 비활성) — 관리자 화면용
        "all_treatments": [_serialize_treatment(t) for t in treatments],
    }


@router.get("/treatment-meta")
def treatment_meta(db: Session = Depends(get_db)):
    return _build_treatment_meta(db)


# ──────────────── 치료항목 CRUD ────────────────

@router.get("/treatments")
def list_treatments(db: Session = Depends(get_db)):
    """모든 치료항목 (활성+비활성) — 관리자 화면용."""
    items = db.query(models.Treatment).order_by(models.Treatment.sort_order).all()
    return [_serialize_treatment(t) for t in items]


def _slug_code(name: str) -> str:
    """이름에서 코드 자동 생성 (영문/숫자 외 'tx_<random>')."""
    import re, secrets
    cleaned = re.sub(r"[^a-zA-Z0-9_]", "", name)
    if cleaned and cleaned[0].isalpha():
        return cleaned.lower()[:30]
    return f"tx_{secrets.token_hex(4)}"


@router.post("/treatments")
def create_treatment(p: schemas.TreatmentIn, db: Session = Depends(get_db),
                     _: bool = Depends(require_admin)):
    # 약자 중복 검사
    dup = db.query(models.Treatment).filter(models.Treatment.short == p.short).first()
    if dup:
        raise HTTPException(400, f"약자 '{p.short}' 가 이미 사용 중입니다 (항목: {dup.name})")
    # 코드 결정
    code = (p.code or "").strip() or _slug_code(p.name)
    code_dup = db.query(models.Treatment).filter(models.Treatment.code == code).first()
    if code_dup:
        # 코드 충돌 시 자동 변경
        import secrets
        code = f"{code}_{secrets.token_hex(2)}"
    # sort_order 자동: 기존 max+1
    max_sort = db.query(models.Treatment).count()
    if p.role not in ("doctor", "therapist"):
        raise HTTPException(400, "역할은 doctor 또는 therapist 여야 합니다.")

    inc_pct, inc_amount = _normalize_incentive(p.incentive_pct, p.incentive_amount)
    t = models.Treatment(
        code=code,
        name=p.name.strip(),
        short=p.short.strip(),
        default_minutes=max(5, p.default_minutes),
        role=p.role,
        count_increment=max(0, p.count_increment),
        show_in_patient=p.show_in_patient,
        active=p.active,
        sort_order=p.sort_order or (max_sort + 1),
        price=max(0, int(p.price or 0)),
        incentive_pct=inc_pct,
        incentive_amount=inc_amount,
    )
    db.add(t); db.flush()
    audit(db, "treatment.create", t.id, t.name)
    _log(db, "treatment", t.id, "upsert", t)
    db.commit()
    return _serialize_treatment(t)


@router.put("/treatments/{tid}")
def update_treatment(tid: str, p: schemas.TreatmentIn, db: Session = Depends(get_db),
                     _: bool = Depends(require_admin)):
    t = db.query(models.Treatment).filter_by(id=tid).first()
    if not t:
        raise HTTPException(404, "치료항목 없음")
    # 약자 중복 (자기 제외)
    dup = db.query(models.Treatment).filter(
        models.Treatment.short == p.short,
        models.Treatment.id != tid,
    ).first()
    if dup:
        raise HTTPException(400, f"약자 '{p.short}' 가 이미 사용 중입니다 (항목: {dup.name})")
    if p.role not in ("doctor", "therapist"):
        raise HTTPException(400, "역할은 doctor 또는 therapist 여야 합니다.")

    inc_pct, inc_amount = _normalize_incentive(p.incentive_pct, p.incentive_amount)
    t.name = p.name.strip()
    t.short = p.short.strip()
    t.default_minutes = max(5, p.default_minutes)
    t.role = p.role
    t.count_increment = max(0, p.count_increment)
    t.show_in_patient = p.show_in_patient
    t.active = p.active
    if p.sort_order:
        t.sort_order = p.sort_order
    t.price = max(0, int(p.price or 0))
    t.incentive_pct = inc_pct
    t.incentive_amount = inc_amount
    db.flush()
    audit(db, "treatment.update", t.id, t.name)
    _log(db, "treatment", t.id, "upsert", t)
    db.commit()
    return _serialize_treatment(t)


@router.get("/treatments/{tid}/references")
def treatment_references(tid: str, db: Session = Depends(get_db),
                         _: bool = Depends(require_admin)):
    """치료항목을 참조하는 예약 목록 (삭제 전 확인용)."""
    t = db.query(models.Treatment).filter_by(id=tid).first()
    if not t:
        raise HTTPException(404, "치료항목 없음")
    # treatment_codes JSON에 t.code 가 들어있는 모든 예약
    appts = db.query(models.Appointment).all()
    refs = []
    for a in appts:
        codes = _parse_codes(a.treatment_codes or "[]")
        if t.code in codes:
            refs.append({
                "appt_id": a.id,
                "start_at": a.start_at.isoformat(),
                "patient_name": a.patient.name if a.patient else "?",
                "chart_no": a.patient.chart_no if a.patient else "-",
                "status": a.status,
            })
    refs.sort(key=lambda x: x["start_at"], reverse=True)
    return {"treatment": _serialize_treatment(t), "references": refs, "count": len(refs)}


@router.delete("/treatments/{tid}")
def delete_treatment(tid: str, db: Session = Depends(get_db),
                     _: bool = Depends(require_admin)):
    t = db.query(models.Treatment).filter_by(id=tid).first()
    if not t:
        raise HTTPException(404, "치료항목 없음")
    # 참조 예약 검증
    appts = db.query(models.Appointment).all()
    refs_count = 0
    for a in appts:
        codes = _parse_codes(a.treatment_codes or "[]")
        if t.code in codes:
            refs_count += 1
    if refs_count > 0:
        raise HTTPException(400,
            f"이 치료항목을 사용하는 예약이 {refs_count}건 있어 삭제할 수 없습니다. "
            f"비활성화로 전환하거나, 참조 예약을 정리한 후 다시 시도하세요.")
    # PatientTreatmentCount 도 cascade 로 삭제됨
    audit(db, "treatment.delete", tid, t.name)
    _log(db, "treatment", tid, "delete", t)
    db.delete(t)
    db.commit()
    return {"ok": True}


# ──────────────── 직원 (Employee) ────────────────

@router.get("/employees")
def list_employees(role: str = "", active: Optional[bool] = None,
                   db: Session = Depends(get_db)):
    q = db.query(models.Employee)
    if role:
        q = q.filter(models.Employee.role == role)
    if active is not None:
        q = q.filter(models.Employee.active == active)
    rows = q.order_by(models.Employee.sort_order, models.Employee.name).all()
    return [_serialize_employee(e) for e in rows]


@router.post("/employees/reorder")
async def reorder_employees(request: Request, db: Session = Depends(get_db)):
    """직원 순서 저장. body = [{"id": "...", "sort_order": N}, ...]"""
    payload = await request.json()
    for item in payload:
        e = db.get(models.Employee, item["id"])
        if e:
            e.sort_order = item["sort_order"]
    db.commit()
    return {"ok": True}


@router.post("/employees")
def create_employee(p: schemas.EmployeeIn, db: Session = Depends(get_db)):
    if p.role not in C.ROLES:
        raise HTTPException(400, f"role은 {C.ROLES} 중 하나여야 합니다.")
    # 신규 직원 sort_order = 같은 역할 내 최댓값 + 1
    max_order = db.query(models.Employee).filter(
        models.Employee.role == p.role
    ).count()
    e = models.Employee(**p.model_dump())
    e.sort_order = max_order + 1
    db.add(e); db.flush()
    _log(db, "employee", e.id, "upsert", e)
    audit(db, "employee.create", e.id, f"name={e.name} role={e.role}")
    db.commit(); db.refresh(e)
    return _serialize_employee(e)


@router.put("/employees/{eid}")
def update_employee(eid: str, p: schemas.EmployeeIn, db: Session = Depends(get_db)):
    e = db.get(models.Employee, eid)
    if not e:
        raise HTTPException(404)
    if p.role not in C.ROLES:
        raise HTTPException(400, f"role은 {C.ROLES} 중 하나여야 합니다.")
    for k, v in p.model_dump().items():
        setattr(e, k, v)
    db.flush()
    _log(db, "employee", e.id, "upsert", e)
    audit(db, "employee.update", e.id, f"name={e.name}")
    db.commit(); db.refresh(e)
    return _serialize_employee(e)


@router.delete("/employees/{eid}")
def delete_employee(eid: str, db: Session = Depends(get_db),
                    _: bool = Depends(require_admin)):
    """삭제는 관리자 비밀번호 필요."""
    e = db.get(models.Employee, eid)
    if not e:
        raise HTTPException(404)
    db.delete(e)
    _log(db, "employee", eid, "delete", None)
    audit(db, "employee.delete", eid, f"name={e.name}")
    db.commit()
    return {"ok": True}


# ──────────────── 직원 휴무 (EmployeeLeave) ────────────────

@router.get("/employee-leaves")
def list_employee_leaves(date: str = "", db: Session = Depends(get_db)):
    q = db.query(models.EmployeeLeave)
    if date:
        q = q.filter(models.EmployeeLeave.leave_date == date)
    rows = q.order_by(models.EmployeeLeave.leave_date.asc()).all()
    return [{
        "id": r.id,
        "employee_id": r.employee_id,
        "leave_date": r.leave_date,
        "leave_type": r.leave_type or "full",
        "leave_kind": r.leave_kind or "annual",
        "memo": r.memo or "",
    } for r in rows]


def _upsert_employee_leave_core(db: Session, p: schemas.EmployeeLeaveIn) -> models.EmployeeLeave:
    """동일 (employee_id, leave_date) 키면 update, 아니면 insert. commit 안 함.

    세션 13: AI 자연어 휴무 등록 흐름이 같은 헬퍼를 트랜잭션 안에서 호출.
    sync 로깅 (_log) 도 여기서 처리 — 호출자는 commit 만 책임.
    """
    exists = db.query(models.EmployeeLeave).filter(
        models.EmployeeLeave.employee_id == p.employee_id,
        models.EmployeeLeave.leave_date == p.leave_date,
    ).first()
    if exists:
        exists.leave_type = p.leave_type
        exists.leave_kind = p.leave_kind
        exists.memo = p.memo
        db.flush()
        _log(db, "employee_leave", exists.id, "upsert", exists)
        return exists
    obj = models.EmployeeLeave(**p.model_dump())
    db.add(obj); db.flush()
    _log(db, "employee_leave", obj.id, "upsert", obj)
    return obj


@router.post("/employee-leaves")
def create_employee_leave(p: schemas.EmployeeLeaveIn, db: Session = Depends(get_db)):
    obj = _upsert_employee_leave_core(db, p)
    db.commit(); db.refresh(obj)
    return {
        "id": obj.id, "employee_id": obj.employee_id,
        "leave_date": obj.leave_date, "leave_type": obj.leave_type,
        "leave_kind": obj.leave_kind or "annual",
        "memo": obj.memo or "",
    }


@router.delete("/employee-leaves/{lid}")
def delete_employee_leave(lid: str, db: Session = Depends(get_db)):
    obj = db.get(models.EmployeeLeave, lid)
    if not obj:
        raise HTTPException(404, "휴무 데이터가 없습니다.")
    db.delete(obj)
    _log(db, "employee_leave", lid, "delete", None)
    db.commit()
    return {"ok": True}


@router.post("/employee-leaves/bulk-set")
def bulk_set_employee_leaves(payload: dict, db: Session = Depends(get_db)):
    leave_date = (payload or {}).get("leave_date", "")
    items = (payload or {}).get("items", [])
    memo = (payload or {}).get("memo", "")
    if not leave_date:
        raise HTTPException(400, "leave_date가 필요합니다.")

    db.query(models.EmployeeLeave).filter(
        models.EmployeeLeave.leave_date == leave_date
    ).delete()

    count = 0
    for item in items:
        emp_id = item.get("employee_id") or item.get("therapist_id")  # 호환
        leave_type = item.get("leave_type", "full")
        leave_kind = item.get("leave_kind", "annual")
        if not emp_id:
            continue
        db.add(models.EmployeeLeave(
            employee_id=emp_id, leave_date=leave_date,
            leave_type=leave_type, leave_kind=leave_kind, memo=memo,
        ))
        count += 1
    db.commit()
    return {"ok": True, "count": count}


# ──────────────── 호환 alias: /api/therapists, /api/therapist-leaves ────────────────
# 프론트가 단계별로 migrate되는 동안 안 깨지게.

@router.get("/therapists")
def list_therapists_alias(db: Session = Depends(get_db)):
    """치료사 역할만 반환 (프론트 호환)."""
    rows = (db.query(models.Employee)
            .filter(models.Employee.role == C.ROLE_THERAPIST)
            .order_by(models.Employee.name).all())
    return [_serialize_employee(e) for e in rows]


@router.get("/therapist-leaves")
def list_therapist_leaves_alias(date: str = "", db: Session = Depends(get_db)):
    q = db.query(models.EmployeeLeave)
    if date:
        q = q.filter(models.EmployeeLeave.leave_date == date)
    rows = q.order_by(models.EmployeeLeave.leave_date.asc()).all()
    # 프론트 호환 위해 therapist_id 키로도 반환
    return [{
        "id": r.id,
        "therapist_id": r.employee_id,
        "employee_id": r.employee_id,
        "leave_date": r.leave_date,
        "leave_type": r.leave_type or "full",
        "leave_kind": r.leave_kind or "annual",
        "memo": r.memo or "",
    } for r in rows]


@router.post("/therapist-leaves/bulk-set")
def bulk_set_therapist_leaves_alias(payload: dict, db: Session = Depends(get_db)):
    items = (payload or {}).get("items", [])
    for it in items:
        if "therapist_id" in it and "employee_id" not in it:
            it["employee_id"] = it["therapist_id"]
    return bulk_set_employee_leaves(payload, db)


# ──────────────── 환자 ────────────────

def _patient_counts_dict(p: models.Patient, db) -> dict:
    """환자별 치료항목 처방/완료 카운트 dict (treatment_id → {rx, done, code, name, short, show, role})."""
    items = db.query(models.PatientTreatmentCount).filter_by(patient_id=p.id).all()
    by_id = {x.treatment_id: x for x in items}
    treatments = db.query(models.Treatment).order_by(models.Treatment.sort_order).all()
    counts = {}
    for t in treatments:
        c = by_id.get(t.id)
        counts[t.id] = {
            "treatment_id": t.id,
            "code": t.code,
            "name": t.name,
            "short": t.short,
            "role": t.role,
            "show": t.show_in_patient,
            "active": t.active,
            "rx_count": c.rx_count if c else 0,
            "done_count": c.done_count if c else 0,
        }
    return counts


def _patient_to_dict(p: models.Patient, db) -> dict:
    counts = _patient_counts_dict(p, db)
    # 표 표시용 축약: show=True 항목만, code 순
    show_items = [c for c in counts.values() if c["show"] and c["active"]]
    show_items.sort(key=lambda x: x["code"])
    return {
        "id": p.id, "name": p.name, "birth_date": p.birth_date,
        "phone": p.phone, "chart_no": p.chart_no,
        "gender": getattr(p, "gender", "") or "",
        "memo": p.memo or "",
        "counts": counts,
        "counts_show": show_items,  # 환자 관리 표 축약 표시용
    }


def _apply_patient_counts(db, patient_id: str, count_inputs):
    """환자 카운트 일괄 upsert (Lazy: 0/0 인 항목은 row 안 만듦)."""
    for ci in count_inputs:
        rx = max(0, ci.rx_count or 0)
        done = max(0, ci.done_count or 0)
        existing = db.query(models.PatientTreatmentCount).filter_by(
            patient_id=patient_id, treatment_id=ci.treatment_id,
        ).first()
        if rx == 0 and done == 0:
            # 0/0 이고 row 없으면 그대로 두고, row 있으면 0/0 으로 업데이트 (수동 리셋 가능)
            if existing:
                existing.rx_count = 0
                existing.done_count = 0
            continue
        if existing:
            existing.rx_count = rx
            existing.done_count = done
        else:
            # 치료항목 존재 검증
            t = db.query(models.Treatment).filter_by(id=ci.treatment_id).first()
            if not t:
                continue
            db.add(models.PatientTreatmentCount(
                patient_id=patient_id,
                treatment_id=ci.treatment_id,
                rx_count=rx,
                done_count=done,
            ))


@router.get("/patients")
def list_patients(light: int = 0, db: Session = Depends(get_db)):
    """환자 전체 목록.

    - light=1: counts 제외한 기본 필드만. 응답 크기 1/6, 파싱 속도 4~5배.
               loadMasters() 등 '환자 이름·차트만 필요한 캐시' 용도.
    - 기본(light=0): counts 포함 full 직렬화 (_serialize_patients_bulk 사용).
      대량(수만 건) 환경에서는 가능한 light 만 사용할 것.
    """
    rows = db.query(models.Patient).order_by(models.Patient.name).all()
    if light:
        return [
            {"id": p.id, "name": p.name, "chart_no": p.chart_no,
             "phone": p.phone, "birth_date": p.birth_date,
             "gender": getattr(p, "gender", "") or "",
             "memo": p.memo or ""}
            for p in rows
        ]
    return _serialize_patients_bulk(rows, db)


@router.get("/patients/search")
def search_patients(q: str = "", field: str = "all", limit: int = 20, offset: int = 0,
                    db: Session = Depends(get_db)):
    """서버 사이드 환자 검색 — 대량 데이터(수만 건) 대응.

    - q 비어있으면 최근 환자 limit 명 반환 (이름순)
    - field: all / name / chart_no / phone / birth_date
    - limit: 페이지 크기 (기본 20, 최대 500)
    - offset: 페이지 오프셋 (기본 0 — "더보기" 누적용)
    """
    from sqlalchemy import or_
    q = (q or "").strip()
    limit = max(1, min(500, int(limit or 20)))
    offset = max(0, int(offset or 0))
    qry = db.query(models.Patient)
    if q:
        like = f"%{q}%"
        P = models.Patient
        if field == "name":
            qry = qry.filter(P.name.like(like))
        elif field == "chart_no":
            qry = qry.filter(P.chart_no.like(like))
        elif field == "phone":
            qry = qry.filter(P.phone.like(like))
        elif field == "birth_date":
            qry = qry.filter(P.birth_date.like(like))
        else:
            qry = qry.filter(or_(
                P.name.like(like), P.chart_no.like(like),
                P.phone.like(like), P.birth_date.like(like),
            ))
    total_matched = qry.count() if q else db.query(models.Patient).count()
    rows = qry.order_by(models.Patient.name).offset(offset).limit(limit).all()
    items = [
        {"id": p.id, "name": p.name, "chart_no": p.chart_no,
         "phone": p.phone, "birth_date": p.birth_date,
         "gender": getattr(p, "gender", "") or "",
         "memo": p.memo or ""}
        for p in rows
    ]
    return {
        "items": items, "total": total_matched,
        "limit": limit, "offset": offset, "q": q,
        "has_more": (offset + len(items)) < total_matched,
    }


@router.get("/patients/{pid}")
def get_patient(pid: str, db: Session = Depends(get_db)):
    """단건 환자 상세 (counts 포함). 편집 모달용."""
    p = db.get(models.Patient, pid)
    if not p:
        raise HTTPException(404)
    return _serialize_patients_bulk([p], db)[0]


def _serialize_patients_bulk(patients, db):
    """여러 환자를 한 번에 직렬화 — counts 를 일괄 로드해서 N+1 회피."""
    from collections import defaultdict
    if not patients:
        return []
    treatments = db.query(models.Treatment).order_by(models.Treatment.sort_order).all()
    pids = [p.id for p in patients]
    # 환자 수가 많으면 IN 절 대신 전체 로드 (SQLite 는 IN 파라미터 수 한계 ~999)
    if len(pids) > 500:
        cnt_rows = db.query(models.PatientTreatmentCount).all()
    else:
        cnt_rows = (db.query(models.PatientTreatmentCount)
                    .filter(models.PatientTreatmentCount.patient_id.in_(pids))
                    .all())
    by_pid = defaultdict(dict)
    for c in cnt_rows:
        by_pid[c.patient_id][c.treatment_id] = c

    # 치료항목 스키마를 매번 새로 만들지 않도록 템플릿 캐시
    tx_templates = []
    for t in treatments:
        tx_templates.append({
            "treatment_id": t.id, "code": t.code, "name": t.name,
            "short": t.short, "role": t.role,
            "show": t.show_in_patient, "active": t.active,
        })

    out = []
    for p in patients:
        counts = {}
        counts_show = []
        row_counts = by_pid.get(p.id, {})
        for tpl in tx_templates:
            c = row_counts.get(tpl["treatment_id"])
            item = {**tpl,
                    "rx_count": c.rx_count if c else 0,
                    "done_count": c.done_count if c else 0}
            counts[tpl["treatment_id"]] = item
            if tpl["show"] and tpl["active"]:
                counts_show.append(item)
        counts_show.sort(key=lambda x: x["code"])
        out.append({
            "id": p.id, "name": p.name, "birth_date": p.birth_date,
            "phone": p.phone, "chart_no": p.chart_no,
            "gender": getattr(p, "gender", "") or "",
            "memo": p.memo or "",
            "counts": counts, "counts_show": counts_show,
        })
    return out


def _check_patient_duplicate(db: Session, name: str, birth_date, chart_no):
    """신규 환자 등록 시 중복 차단.
    1) chart_no 비어있지 않고 같은 chart_no 가 이미 있으면 409
    2) name + birth_date 둘 다 비어있지 않고 같은 조합이 이미 있으면 409
    이름만 같음 / 전화번호만 같음 / 이름+전화번호만 같음 → 차단하지 않음.
    data-convert 의 자체 중복 판정은 별개 — 이 함수는 POST /api/patients 전용.
    """
    nm = (name or "").strip()
    bd = (birth_date or "").strip() if isinstance(birth_date, str) else (birth_date or "")
    cn = (chart_no or "").strip()
    if cn:
        exists = db.query(models.Patient.id).filter(models.Patient.chart_no == cn).first()
        if exists:
            raise HTTPException(409, "이미 등록된 차트번호입니다.")
    if nm and bd:
        exists = (db.query(models.Patient.id)
                  .filter(models.Patient.name == nm,
                          models.Patient.birth_date == bd)
                  .first())
        if exists:
            raise HTTPException(409, "같은 이름과 생년월일의 환자가 이미 등록되어 있습니다.")


@router.post("/patients")
def create_patient(p: schemas.PatientIn, db: Session = Depends(get_db)):
    _check_patient_duplicate(db, p.name, p.birth_date, p.chart_no)
    payload = p.model_dump(exclude={"counts"})
    obj = models.Patient(**payload)
    db.add(obj); db.flush()
    if p.counts:
        _apply_patient_counts(db, obj.id, p.counts)
    _log(db, "patient", obj.id, "upsert", obj)
    db.commit(); db.refresh(obj)
    return _patient_to_dict(obj, db)


@router.patch("/patients/{pid}/memo")
def update_patient_memo(pid: str, body: dict, db: Session = Depends(get_db)):
    """예약 화면에서 환자 누적 메모만 단독 저장."""
    obj = db.get(models.Patient, pid)
    if not obj:
        raise HTTPException(404)
    obj.memo = body.get("memo", obj.memo or "")
    _log(db, "patient", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True}


@router.put("/patients/{pid}")
def update_patient(pid: str, p: schemas.PatientIn, db: Session = Depends(get_db)):
    """환자 관리 모달에서 전체 필드 저장 (이름/생년월일/연락처/차트/메모/카운트)."""
    obj = db.get(models.Patient, pid)
    if not obj:
        raise HTTPException(404)
    payload = p.model_dump(exclude={"counts"})
    for k, v in payload.items():
        setattr(obj, k, v)
    db.flush()
    if p.counts:
        _apply_patient_counts(db, obj.id, p.counts)
    _log(db, "patient", obj.id, "upsert", obj)
    audit(db, "patient.update", obj.id, f"name={obj.name}")
    db.commit(); db.refresh(obj)
    return _patient_to_dict(obj, db)


@router.delete("/patients/{pid}")
def delete_patient(pid: str, db: Session = Depends(get_db),
                   _: bool = Depends(require_admin)):
    obj = db.get(models.Patient, pid)
    if not obj:
        raise HTTPException(404)
    db.delete(obj)
    _log(db, "patient", pid, "delete", None)
    audit(db, "patient.delete", pid, f"name={obj.name}")
    db.commit()
    return {"ok": True}


@router.get("/patients/last-appointments")
def last_appointments(db: Session = Depends(get_db)):
    rows = db.query(
        models.Appointment.patient_id,
        func.max(models.Appointment.start_at).label("last")
    ).filter(
        models.Appointment.status != "canceled"
    ).group_by(models.Appointment.patient_id).all()
    return {r[0]: r[1].isoformat() if r[1] else None for r in rows}


@router.get("/patients/{pid}/manual-history-summary")
def patient_manual_history_summary(pid: str, db: Session = Depends(get_db)):
    """도수치료 이력 요약 — 신환 체크박스 자동 토글용.
    canceled 제외 모든 상태(reserved/approved) 중 도수치료 코드 포함 여부 체크.
    """
    manual_codes = set(_get_manual_therapy_codes(db))
    rows = (db.query(models.Appointment)
            .filter(models.Appointment.patient_id == pid,
                    models.Appointment.status != "canceled")
            .all())
    manual_appt_ids = []
    has_new_patient_flag = False
    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        if any(c in manual_codes for c in codes):
            manual_appt_ids.append(a.id)
        if getattr(a, 'is_new_patient', False):
            has_new_patient_flag = True
    return {
        "patient_id": pid,
        "has_manual_history": len(manual_appt_ids) > 0,
        "manual_count": len(manual_appt_ids),
        "has_new_patient_flag": has_new_patient_flag,
        "manual_appointment_ids": manual_appt_ids,
    }


@router.get("/patients/{pid}/history")
def patient_history(pid: str, offset: int = 0, limit: int = 30,
                    db: Session = Depends(get_db)):
    """환자 치료이력 (approved만).

    같은 날 여러 예약이 있어도 **방문일 단위(날짜)로 1줄**로 묶어 반환.
    - offset/limit 은 '방문일 수' 기준
    - total = 총 방문일 수 (예약 건수 아님)
    - days[*].appointments = 그 날의 예약들 (시간 오름차순)
    - 하위 호환을 위해 기존 items (평면화된 예약 리스트) 도 함께 반환.
    """
    from collections import OrderedDict
    # approved 전량을 최신순으로 불러와 날짜별 묶기
    rows_all = (db.query(models.Appointment)
                .filter(models.Appointment.patient_id == pid,
                        models.Appointment.status == "approved")
                .order_by(models.Appointment.start_at.desc())
                .all())

    # 직원 이름 캐시
    emp_ids = set()
    for a in rows_all:
        if a.therapist_id: emp_ids.add(a.therapist_id)
        for asn in a.assignments:
            if asn.handler_id: emp_ids.add(asn.handler_id)
    emp_map = {}
    if emp_ids:
        for e in db.query(models.Employee).filter(models.Employee.id.in_(emp_ids)).all():
            emp_map[e.id] = e.name

    # 날짜별 그룹 (최신 방문일 순 — rows_all 이 이미 desc)
    grouped = OrderedDict()
    for a in rows_all:
        dkey = a.start_at.date().isoformat()
        grouped.setdefault(dkey, []).append(a)

    total_days = len(grouped)
    all_dates = list(grouped.keys())
    sliced = all_dates[offset: offset + limit]

    def _serialize_appt(a):
        codes = _parse_codes(a.treatment_codes)
        asn_map = {x.treatment_code: x.handler_id for x in a.assignments}
        return {
            "id": a.id,
            "start_at": a.start_at.isoformat(),
            "treatment_codes": codes,
            "therapist_id": a.therapist_id,
            "therapist_name": emp_map.get(a.therapist_id),
            "assignments": {
                code: {"handler_id": hid, "handler_name": emp_map.get(hid)}
                for code, hid in asn_map.items()
            },
        }

    days = []
    legacy_items = []  # fetchLastManualTherapist 등 호환용 (최신→과거)
    for dkey in sliced:
        day_appts = grouped[dkey]
        # 날짜 내부는 시간 오름차순으로 표시
        day_sorted_asc = sorted(day_appts, key=lambda a: a.start_at)
        # 평면 items 는 최신 순 유지 (fetchLastManualTherapist 가 [0] = 가장 최근 기대)
        day_sorted_desc = sorted(day_appts, key=lambda a: a.start_at, reverse=True)

        days.append({
            "date": dkey,
            "visit_start_at": day_sorted_asc[0].start_at.isoformat(),
            "appointments": [_serialize_appt(a) for a in day_sorted_asc],
        })
        for a in day_sorted_desc:
            legacy_items.append(_serialize_appt(a))

    return {
        "total": total_days,       # 총 방문일 수
        "offset": offset,
        "limit": limit,
        "days": days,
        "items": legacy_items,     # 하위 호환 (평면 예약 리스트, 최신순)
    }


# ──────────────── 예약 (단순 버전, 4단계에서 확장) ────────────────

@router.get("/appointments")
def list_appointments(start: str, end: str, category: str = "",  # category는 호환 위해 받고 무시
                      db: Session = Depends(get_db)):
    ts = datetime.fromisoformat(start.replace("Z", "+00:00").replace(" ", "+"))
    te = datetime.fromisoformat(end.replace("Z", "+00:00").replace(" ", "+"))
    if ts.tzinfo: ts = ts.replace(tzinfo=None)
    if te.tzinfo: te = te.replace(tzinfo=None)
    rows = (db.query(models.Appointment)
            .filter(models.Appointment.start_at >= ts,
                    models.Appointment.start_at < te).all())
    return [_serialize_appointment(a) for a in rows]


@router.post("/appointments")
def create_appointment(p: schemas.AppointmentIn, db: Session = Depends(get_db)):
    valid_codes = _existing_codes_set(db)
    therapist_only_codes = _therapist_only_codes_set(db)  # 치료사 + 체외충격파 제외 → assignment 만들지 않음
    codes = [c for c in (p.treatment_codes or []) if c in valid_codes]
    if not codes:
        raise HTTPException(400, "치료항목(treatment_codes)을 하나 이상 선택하세요.")
    _check_lunch_block(p.start_at, p.duration_min)
    obj = models.Appointment(
        patient_id=p.patient_id,
        therapist_id=p.therapist_id,
        start_at=p.start_at,
        end_at=p.start_at + timedelta(minutes=p.duration_min),
        duration_min=p.duration_min,
        treatment_codes=json.dumps(codes, ensure_ascii=False),
        memo=p.memo, status="reserved",
        is_new_patient=getattr(p, 'is_new_patient', False) or False,
    )
    db.add(obj); db.flush()
    # 초기 assignments — 치료사 항목(체외충격파 제외)은 만들지 않음 (담당은 appointment.therapist_id 로)
    for a in (p.assignments or []):
        if a.treatment_code in codes and a.treatment_code not in therapist_only_codes:
            db.add(models.TreatmentAssignment(
                appointment_id=obj.id,
                treatment_code=a.treatment_code,
                handler_id=a.handler_id,
            ))
    # 누락된 항목: NULL handler 로 채우기 (의사 공용 또는 체외충격파 미배정)
    supplied = {a.treatment_code for a in (p.assignments or [])}
    for code in codes:
        if code in therapist_only_codes:
            continue
        if code in supplied:
            continue
        db.add(models.TreatmentAssignment(
            appointment_id=obj.id, treatment_code=code, handler_id=None,
        ))
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit(); db.refresh(obj)
    return {"id": obj.id, "status": obj.status}


def _check_version(obj: models.Appointment, client_version):
    """낙관적 락: 클라이언트가 보낸 version 이 DB 와 다르면 409."""
    if client_version is None:
        return
    if (obj.version or 0) != client_version:
        raise HTTPException(status_code=409, detail={
            "error": "version_conflict",
            "message": "다른 PC에서 먼저 수정되었습니다. 최신 정보를 불러오세요.",
            "current_version": int(obj.version or 0),
        })


def _bump_version(obj: models.Appointment):
    obj.version = (obj.version or 0) + 1


@router.put("/appointments/{aid}")
def update_appointment(aid: str, p: schemas.AppointmentUpdate,
                       db: Session = Depends(get_db)):
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    if obj.status in ("approved", "canceled"):
        raise HTTPException(400, "확정/취소된 예약은 수정할 수 없습니다.")
    _check_version(obj, p.version)
    valid_codes = _existing_codes_set(db)
    therapist_only_codes = _therapist_only_codes_set(db)
    data = p.model_dump(exclude_unset=True)
    data.pop("version", None)  # 낙관적 락 필드 — 모델 속성 아님
    codes_changed = "treatment_codes" in data
    for k, v in data.items():
        if k == "treatment_codes":
            filtered = [c for c in (v or []) if c in valid_codes]
            if not filtered:
                raise HTTPException(400, "치료항목이 비어 있습니다.")
            obj.treatment_codes = json.dumps(filtered, ensure_ascii=False)
        elif k == "assignments":
            continue
        else:
            setattr(obj, k, v)
    if "start_at" in data or "duration_min" in data:
        obj.end_at = obj.start_at + timedelta(minutes=obj.duration_min)
        _check_lunch_block(obj.start_at, obj.duration_min)

    # assignments 갱신
    if "assignments" in data and data["assignments"] is not None:
        existing = {a.treatment_code: a for a in obj.assignments}
        for a in data["assignments"]:
            code = a.get("treatment_code") if isinstance(a, dict) else a.treatment_code
            hid = a.get("handler_id") if isinstance(a, dict) else a.handler_id
            if code in therapist_only_codes:
                continue
            if code in existing:
                existing[code].handler_id = hid
            else:
                db.add(models.TreatmentAssignment(
                    appointment_id=obj.id, treatment_code=code, handler_id=hid,
                ))
        current_codes = _parse_codes(obj.treatment_codes)
        for code, asn in existing.items():
            if code not in current_codes:
                db.delete(asn)
    elif codes_changed:
        existing = {a.treatment_code: a for a in obj.assignments}
        current_codes = _parse_codes(obj.treatment_codes)
        for code, asn in existing.items():
            if code not in current_codes:
                db.delete(asn)
        for code in current_codes:
            if code in therapist_only_codes:
                continue
            if code not in existing:
                db.add(models.TreatmentAssignment(
                    appointment_id=obj.id, treatment_code=code, handler_id=None,
                ))

    _bump_version(obj)
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True, "version": int(obj.version or 0)}


@router.post("/appointments/{aid}/assign")
def change_assignment(aid: str, p: schemas.AssignmentChange,
                      db: Session = Depends(get_db)):
    """특정 치료항목의 담당자 변경 (체외충격파 셀 이동 / 주사 의사 배정 등)."""
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    _check_version(obj, p.version)
    therapist_only_codes = _therapist_only_codes_set(db)
    valid_codes = _existing_codes_set(db)
    doctor_codes = _doctor_codes_set(db)

    if p.treatment_code in therapist_only_codes:
        raise HTTPException(400, "치료사 항목(체외충격파 제외)은 예약의 담당 치료사 필드로 관리합니다.")
    if p.treatment_code not in valid_codes:
        raise HTTPException(400, "알 수 없는 치료항목입니다.")
    codes = _parse_codes(obj.treatment_codes)
    if p.treatment_code not in codes:
        raise HTTPException(400, "이 예약에 해당 치료항목이 없습니다.")

    # 체외충격파 이동 시 대상자가 치료사 + can_eswt=True 여야
    if p.treatment_code == C.ESWT_CODE and p.handler_id:
        h = db.get(models.Employee, p.handler_id)
        if not h or h.role != C.ROLE_THERAPIST or not h.can_eswt:
            raise HTTPException(400, "선택한 직원은 체외충격파 담당이 불가합니다.")
    # 의사 항목 배정 시 대상자가 role=doctor 여야
    if p.treatment_code in doctor_codes and p.handler_id:
        h = db.get(models.Employee, p.handler_id)
        if not h or h.role != C.ROLE_DOCTOR:
            raise HTTPException(400, "이 항목은 의사 역할만 배정할 수 있습니다.")

    existing = next((a for a in obj.assignments
                     if a.treatment_code == p.treatment_code), None)
    if existing:
        existing.handler_id = p.handler_id
    else:
        db.add(models.TreatmentAssignment(
            appointment_id=obj.id, treatment_code=p.treatment_code,
            handler_id=p.handler_id,
        ))
    _bump_version(obj)
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True, "version": int(obj.version or 0)}


@router.post("/appointments/{aid}/split-code")
def split_appointment_code(aid: str, payload: dict, db: Session = Depends(get_db)):
    """한 예약에서 특정 치료항목(split_code)을 분리해 **새 예약**으로 만듭니다.
    같은 환자의 충격파 + 도수치료가 한 예약으로 묶여 있을 때,
    한쪽만 다른 시간/치료사로 이동시키고 나머지는 원래 자리에 남기려는 용도.

    payload:
      - treatment_code (필수): 분리할 치료 코드
      - start_at (선택): 새 예약 시작 시각 (ISO8601). 없으면 원본과 동일.
      - therapist_id (선택): 도수치료 담당. 없으면 원본과 동일.
      - handler_id (선택): 체외충격파/의사 등 assignments 담당자.
      - version (선택): 낙관적 락.

    반환:
      - split=False: 원본에 단독 코드만 있어 분리 불필요 → 원본을 업데이트한 경우
      - split=True : 원본에서 코드 제거 + 새 예약 생성
    """
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    if obj.status in ("approved", "canceled"):
        raise HTTPException(400, "확정/취소된 예약은 수정할 수 없습니다.")
    _check_version(obj, payload.get("version"))

    valid_codes = _existing_codes_set(db)
    therapist_only = _therapist_only_codes_set(db)
    split_code = (payload.get("treatment_code") or "").strip()
    if not split_code:
        raise HTTPException(400, "treatment_code 가 필요합니다.")
    if split_code not in valid_codes:
        raise HTTPException(400, "알 수 없는 치료항목입니다.")

    codes = _parse_codes(obj.treatment_codes)
    if split_code not in codes:
        raise HTTPException(400, "이 예약에 해당 치료항목이 없습니다.")

    # 새 예약에 반영할 값
    new_start = obj.start_at
    raw_start = payload.get("start_at")
    if raw_start:
        try:
            new_start = datetime.fromisoformat(
                str(raw_start).replace("Z", "").replace(" ", "T")
            )
            if new_start.tzinfo:
                new_start = new_start.replace(tzinfo=None)
        except Exception:
            raise HTTPException(400, "start_at 형식 오류")

    # 새 예약의 담당 치료사
    new_tid = payload.get("therapist_id", obj.therapist_id)
    if new_tid == "":
        new_tid = None
    # 새 예약의 assignment handler (체외충격파 / 주사 등)
    new_handler = payload.get("handler_id", None)
    if new_handler == "":
        new_handler = None

    remaining = [c for c in codes if c != split_code]

    # 원본에 split_code 만 있던 경우 → 분리 불필요, 원본 자체 업데이트
    if not remaining:
        if raw_start:
            obj.start_at = new_start
            obj.end_at = new_start + timedelta(minutes=obj.duration_min or 30)
            _check_lunch_block(obj.start_at, obj.duration_min or 30)
        if "therapist_id" in payload:
            obj.therapist_id = new_tid
        if "handler_id" in payload and split_code not in therapist_only:
            existing_asn = next(
                (a for a in obj.assignments if a.treatment_code == split_code), None
            )
            if existing_asn:
                existing_asn.handler_id = new_handler
            else:
                db.add(models.TreatmentAssignment(
                    appointment_id=obj.id, treatment_code=split_code,
                    handler_id=new_handler,
                ))
        _bump_version(obj)
        db.flush()
        _log(db, "appointment", obj.id, "upsert", obj)
        db.commit()
        return {"ok": True, "split": False, "id": obj.id, "version": int(obj.version or 0)}

    # ─── 실제 분리 ───
    # 새 예약 duration = 해당 치료 default_minutes (없으면 30)
    t_row = db.query(models.Treatment).filter_by(code=split_code).first()
    new_dur = (t_row.default_minutes if t_row else None) or 30
    _check_lunch_block(new_start, new_dur)

    new_appt = models.Appointment(
        patient_id=obj.patient_id,
        therapist_id=new_tid,
        start_at=new_start,
        end_at=new_start + timedelta(minutes=new_dur),
        duration_min=new_dur,
        treatment_codes=json.dumps([split_code], ensure_ascii=False),
        memo=obj.memo or "",
        status="reserved",
        # 신환 플래그는 원본에만 유지 (한 환자당 한 건 원칙)
        is_new_patient=False,
    )
    db.add(new_appt)
    db.flush()

    # assignment 이관 (도수치료 등 therapist_only 은 assignment 없음)
    if split_code not in therapist_only:
        src_asn = next(
            (a for a in obj.assignments if a.treatment_code == split_code), None
        )
        # 새 예약에 assignment 추가
        db.add(models.TreatmentAssignment(
            appointment_id=new_appt.id,
            treatment_code=split_code,
            handler_id=new_handler if new_handler is not None else (
                src_asn.handler_id if src_asn else None
            ),
        ))
        # 원본에서 해당 assignment 제거
        if src_asn:
            db.delete(src_asn)

    # 원본에서 split_code 제거
    obj.treatment_codes = json.dumps(remaining, ensure_ascii=False)
    _bump_version(obj)

    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    _log(db, "appointment", new_appt.id, "upsert", new_appt)
    db.commit()
    return {
        "ok": True,
        "split": True,
        "original_id": obj.id,
        "new_id": new_appt.id,
        "version": int(obj.version or 0),
    }


def _bump_patient_count(db, patient_id: str, treatment_code: str, delta: int):
    """환자 완료 카운트 증감 (Lazy 생성, 0 미만 방지)."""
    if delta == 0:
        return
    t = db.query(models.Treatment).filter_by(code=treatment_code).first()
    if not t:
        return
    inc = delta
    row = db.query(models.PatientTreatmentCount).filter_by(
        patient_id=patient_id, treatment_id=t.id,
    ).first()
    if row:
        row.done_count = max(0, (row.done_count or 0) + inc)
    elif delta > 0:
        db.add(models.PatientTreatmentCount(
            patient_id=patient_id,
            treatment_id=t.id,
            rx_count=0,
            done_count=max(0, inc),
        ))


@router.post("/appointments/{aid}/approve")
def approve_appointment(aid: str, p: schemas.ApproveAction,
                        db: Session = Depends(get_db)):
    """원무과 승인 1단계만. approve 시 PatientTreatmentCount.done_count 자동 +N."""
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    if obj.status == "approved":
        raise HTTPException(400, "이미 승인된 예약입니다.")
    if obj.status == "canceled":
        raise HTTPException(400, "취소된 예약은 승인할 수 없습니다.")
    _check_version(obj, p.version)
    obj.status = "approved"
    obj.approved_at = datetime.utcnow()
    obj.approved_by = (p.approved_by or "원무과").strip() or "원무과"
    # 환자 완료 카운트 증가 (PatientTreatmentCount)
    if obj.patient_id:
        for code in _parse_codes(obj.treatment_codes):
            _bump_patient_count(db, obj.patient_id, code, +1)
    _bump_version(obj)
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True, "status": obj.status, "version": int(obj.version or 0)}


@router.post("/appointments/{aid}/revert-approve")
def revert_approve(aid: str, payload: dict = Body(default=None),
                   db: Session = Depends(get_db)):
    """원무과 승인 취소. PatientTreatmentCount.done_count 자동 -N."""
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    if obj.status != "approved":
        raise HTTPException(400, "approved 상태에서만 되돌릴 수 있습니다.")
    client_version = (payload or {}).get("version") if isinstance(payload, dict) else None
    _check_version(obj, client_version)
    obj.status = "reserved"
    obj.approved_at = None
    obj.approved_by = None
    if obj.patient_id:
        for code in _parse_codes(obj.treatment_codes):
            _bump_patient_count(db, obj.patient_id, code, -1)
    _bump_version(obj)
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True, "version": int(obj.version or 0)}


@router.post("/appointments/{aid}/cancel")
def cancel_appointment(aid: str, p: schemas.CancelAction,
                       db: Session = Depends(get_db)):
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    if obj.status == "approved":
        raise HTTPException(400, "승인된 예약은 취소할 수 없습니다. 먼저 승인을 되돌리세요.")
    _check_version(obj, p.version)
    obj.status = "canceled"
    obj.memo = (obj.memo or "") + (f"\n[취소] {p.memo}" if p.memo else "\n[취소]")
    # 20-3-1 (post-19-P / F-10): 노쇼 동시 적용 (사용자 §3-7 권장값 (i)).
    if p.no_show:
        obj.no_show = True
    _bump_version(obj)
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True, "version": int(obj.version or 0), "no_show": bool(obj.no_show)}


@router.post("/appointments/{aid}/mark-no-show")
def mark_no_show(aid: str, db: Session = Depends(get_db)):
    """20-3-1 (post-19-P / F-10): 노쇼 마킹.

    # NOTE: 사용자 §3-7 권장값 (i) — 노쇼 = cancel 동시. obj.no_show=True +
    # status="canceled" 둘 다 적용. 메모에 [노쇼] prefix 자동 추가.
    """
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404)
    if obj.status == "approved":
        raise HTTPException(400, "승인된 예약은 노쇼로 처리할 수 없습니다. 먼저 승인을 되돌리세요.")
    obj.no_show = True
    obj.status = "canceled"
    obj.memo = (obj.memo or "") + "\n[노쇼]"
    _bump_version(obj)
    db.flush()
    _log(db, "appointment", obj.id, "upsert", obj)
    db.commit()
    return {"ok": True, "no_show": True, "status": obj.status, "version": int(obj.version or 0)}


@router.delete("/appointments/{aid}")
def delete_appointment(aid: str, db: Session = Depends(get_db),
                       _: bool = Depends(require_admin)):
    obj = db.get(models.Appointment, aid)
    if not obj:
        raise HTTPException(404, "예약이 없습니다.")
    # approved 되돌림 효과 (완료 카운트 보정)
    if obj.status == "approved" and obj.patient_id:
        for code in _parse_codes(obj.treatment_codes):
            _bump_patient_count(db, obj.patient_id, code, -1)
    db.delete(obj)
    _log(db, "appointment", aid, "delete", None)
    audit(db, "appointment.delete", aid)
    db.commit()
    return {"ok": True}


# ──────────────── 시스템 설정 ────────────────

def _get_or_create_system_setting(db) -> models.SystemSetting:
    ss = db.query(models.SystemSetting).first()
    if not ss:
        ss = models.SystemSetting(
            id=1,
            manual_slot_limit=None,
            sms_template="",
            auto_backup_enabled=True,
            auto_backup_interval_min=60,
            auto_backup_keep_count=30,
        )
        db.add(ss); db.commit(); db.refresh(ss)
    return ss


@router.get("/system-settings")
def system_settings_get(db: Session = Depends(get_db)):
    ss = _get_or_create_system_setting(db)
    # 치료항목 기본 시간은 Treatment.default_minutes 가 단일 진실 원천
    treatments = db.query(models.Treatment).order_by(models.Treatment.sort_order).all()
    minutes = {t.code: t.default_minutes for t in treatments}
    return {
        "manual_slot_limit": ss.manual_slot_limit,
        "treatment_minutes": minutes,   # 호환 (읽기 전용, Treatment 카드에서 편집)
        "sms_template": ss.sms_template or "",
        "auto_backup_enabled": bool(ss.auto_backup_enabled),
        "auto_backup_interval_min": ss.auto_backup_interval_min or 60,
        "auto_backup_keep_count": ss.auto_backup_keep_count or 30,
    }


@router.post("/system-settings")
def system_settings_set(payload: dict, db: Session = Depends(get_db)):
    ss = _get_or_create_system_setting(db)
    if "manual_slot_limit" in payload:
        v = payload["manual_slot_limit"]
        ss.manual_slot_limit = int(v) if v not in (None, "", "null") else None
    if "sms_template" in payload:
        ss.sms_template = (payload["sms_template"] or "")
    if "auto_backup_enabled" in payload:
        ss.auto_backup_enabled = bool(payload["auto_backup_enabled"])
    if "auto_backup_interval_min" in payload:
        v = int(payload["auto_backup_interval_min"] or 60)
        ss.auto_backup_interval_min = max(5, v)   # 최소 5분
    if "auto_backup_keep_count" in payload:
        ss.auto_backup_keep_count = max(1, int(payload["auto_backup_keep_count"] or 30))
    db.flush()
    _log(db, "system_setting", str(ss.id), "upsert", ss)
    audit(db, "system_setting.update", str(ss.id))
    db.commit()
    return {"ok": True}


# ──────────────── 동기화 ────────────────

@router.get("/sync/pull")
def sync_pull(since: str = Query(...), exclude_node: str = Query(""),
              db: Session = Depends(get_db),
              _: bool = Depends(require_admin_or_sync_token)):
    ts = datetime.fromisoformat(since)
    q = db.query(models.SyncOp).filter(models.SyncOp.ts > ts)
    if exclude_node:
        q = q.filter(models.SyncOp.node_id != exclude_node)
    ops = q.order_by(models.SyncOp.ts.asc()).limit(1000).all()
    return {"ops": [{
        "id": o.id, "node_id": o.node_id, "entity": o.entity,
        "entity_id": o.entity_id, "op": o.op, "payload": o.payload,
        "ts": o.ts.isoformat(),
    } for o in ops]}


@router.post("/sync/push")
def sync_push(batch: schemas.SyncBatch, db: Session = Depends(get_db),
              _: bool = Depends(require_admin_or_sync_token)):
    """다른 노드에서 보낸 op 묶음 적용.

    트랜잭션 정책: 각 op 를 개별 try/except 로 보호 — 한 op 가 깨지더라도
    나머지 op 가 적용되도록 하되, 실패 op 는 별도 카운트로 응답.
    이전 코드는 apply_op 가 raise 시 루프 자체가 깨져 db.commit() 미호출 → 부분 상태 잔존 위험.
    """
    from ..services.sync import apply_op
    applied = 0
    failed = 0
    failures = []
    for op in batch.ops:
        if isinstance(op.get("payload"), str):
            try: op["payload"] = json.loads(op["payload"])
            except Exception: op["payload"] = {}
        try:
            if apply_op(db, op):
                applied += 1
        except Exception as e:
            failed += 1
            db.rollback()
            failures.append({"op_id": op.get("id"), "entity": op.get("entity"), "error": str(e)[:200]})
            # rollback 후에도 같은 세션을 계속 쓰려면 다음 op 부터 새로 시작
            continue
    db.commit()
    return {"applied": applied, "failed": failed, "failures": failures}


@router.post("/sync/now")
def sync_now(_: bool = Depends(require_admin)):
    from ..services.sync import sync_with_peer
    cfg = load_config(); results = {}
    targets = []
    if cfg.get("mode") == "sub" and cfg.get("main_url"):
        targets.append(cfg["main_url"])
    targets += cfg.get("peers") or []
    for url in targets:
        results[url] = sync_with_peer(url)
    return results


# ──────────────── 백업 / 복원 ────────────────

@router.get("/backup")
def backup_now(_: bool = Depends(require_admin)):
    src = get_db_path()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = get_backup_dir() / f"clinic_{ts}.db"
    shutil.copy2(src, dst)
    return FileResponse(dst, filename=dst.name, media_type="application/octet-stream")


@router.post("/restore")
def restore(file: UploadFile = File(...), _: bool = Depends(require_admin)):
    """업로드된 DB 파일로 복원.

    정책: 임시 경로에 받아 SQLite integrity check 통과한 경우에만 실제 DB 위치로 atomic rename.

    Windows 안전성:
      - 이전엔 db: Session = Depends(get_db) 로 세션을 잡고 들어와 SQLAlchemy 가 운영 DB
        파일을 lock 해 tmp.replace(dst) 가 PermissionError 로 실패할 가능성이 있었음.
      - 이제 db dependency 를 제거하고, 파일 교체 직전 engine.dispose() 로 connection pool
        을 모두 닫음. 교체 후 새 SessionLocal() 로 audit log 만 새 DB 에 기록.

    감사 로그 폴백:
      - 새 DB 에 audit_logs 테이블이 없거나 호환성 문제가 있어 audit 호출이 실패하면
        backup 폴더의 restore_audit.log 파일에 평문으로 한 줄 기록.
    """
    import os as _os, sqlite3, tempfile
    from pathlib import Path

    from ..database import SessionLocal, engine

    dst = Path(get_db_path())
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(dst, get_backup_dir() / f"clinic_before_restore_{ts}.db")

    filename = file.filename or "(이름없음)"

    # 1) 임시 파일에 업로드 받기 (실패해도 운영 DB 영향 없음).
    #    mkstemp 의 fd 는 즉시 닫는다 — 안 닫으면 Windows 에서 파일 잠금 잔존 가능.
    fd, tmp_path = tempfile.mkstemp(prefix="clinic_restore_", suffix=".db",
                                    dir=str(dst.parent))
    _os.close(fd)
    tmp = Path(tmp_path)
    try:
        with open(tmp, "wb") as f:
            shutil.copyfileobj(file.file, f)

        # 2) 무결성 검증 — SQLite 가 읽고 PRAGMA integrity_check 통과해야 함
        try:
            con = sqlite3.connect(str(tmp))
            try:
                row = con.execute("PRAGMA integrity_check").fetchone()
            finally:
                con.close()
        except Exception as e:
            raise HTTPException(400, f"업로드 파일이 SQLite DB 가 아니거나 손상되었습니다: {e}")
        if not row or row[0] != "ok":
            raise HTTPException(400,
                f"무결성 검사 실패 (결과={row[0] if row else 'None'}). 복원 중단.")

        # 3) 파일 교체 직전: SQLAlchemy connection pool 의 모든 connection 을 닫음.
        #    Windows 에서 SQLite 가 dst 파일을 lock 한 상태에서 tmp.replace 하면 실패함.
        try:
            engine.dispose()
        except Exception:
            pass

        # 4) atomic 교체 — Path.replace 는 Windows 에서도 atomic
        tmp.replace(dst)
        tmp = None  # rename 성공 후 finally 에서 unlink 안 하도록
    finally:
        if tmp is not None and tmp.exists():
            try: tmp.unlink()
            except Exception: pass

    # 5) 감사 로그 — 새 SessionLocal() 로 (복원된) 새 DB 에 기록.
    #    엔진은 dispose 후 자동으로 lazy-reconnect 하므로 같은 DB_URL 로 새 파일 열림.
    audit_recorded = False
    try:
        new_db = SessionLocal()
        try:
            audit(new_db, "db.restore", "", f"file={filename}")
            new_db.commit()
            audit_recorded = True
        finally:
            new_db.close()
    except Exception:
        # audit 테이블이 없거나 신규 DB 가 호환 안 될 수 있음 — 파일 로그 폴백
        pass
    if not audit_recorded:
        try:
            log_path = get_backup_dir() / "restore_audit.log"
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{ts}] db.restore file={filename} (audit 테이블 미존재 폴백)\n")
        except Exception:
            pass

    return {"ok": True, "msg": "복원 완료. 프로그램을 재시작하세요."}


# ──────────────── AI 데이터변환 (엑셀 → 환자 DB) ────────────────

_DC_HEADER_ALIASES = {
    "name":       ["이름", "성명", "환자명", "환자", "name", "patient", "patient_name"],
    "chart_no":   ["차트", "차트번호", "차번", "chart", "chart_no", "chart no", "차트 번호"],
    # 메인 전화 (여러 후보 중 첫 매칭). 이동전화1/휴대폰1 등도 메인으로 간주
    "phone":      ["연락처", "전화", "휴대폰", "전화번호", "phone", "hp", "연락처1",
                   "휴대폰번호", "핸드폰", "mobile", "cell",
                   "이동전화", "이동전화1", "이동전화 1",
                   "휴대폰1", "휴대폰 1", "핸드폰1", "핸드폰 1"],
    "birth_date": ["생년월일", "생일", "birth", "birth_date", "birth date", "dob", "생년"],
    # 성별 — 컬럼명이 있으면 우선 사용
    "gender":     ["성별", "성", "sex", "gender", "m_f", "m/f"],
    # 주민번호 — 성별 컬럼이 없을 때 뒷자리 첫 숫자로 보조 판정
    "ssn":        ["주민번호", "주민", "주민등록번호", "ssn", "rrn", "id_no", "jumin"],
}


def _dc_normalize_gender(raw_gender, raw_ssn=None):
    """성별 추정 — 명확한 경우만 'M'/'F' 반환, 불명확 시 빈 문자열.

    규칙:
      1) raw_gender 가 있으면 우선 사용 (한국어/영문 매핑)
      2) 없으면 raw_ssn 뒷자리 첫 숫자로 홀수=M / 짝수=F
      3) 둘 다 불명확 → '' (검토 필요로 분류됨)
    """
    if raw_gender not in (None, ""):
        s = str(raw_gender).strip().upper()
        if s in ("M", "MALE", "남", "남자", "남성"):
            return "M"
        if s in ("F", "FEMALE", "여", "여자", "여성"):
            return "F"
    # 주민번호 보조 판정 — 뒷자리 첫 숫자
    if raw_ssn not in (None, ""):
        import re as _re
        s = _re.sub(r"[^0-9]", "", str(raw_ssn))
        # 주민번호는 13자리 (YYMMDD + 7자리). 7번째 = 인덱스 6
        if len(s) >= 7 and s[6].isdigit():
            d = int(s[6])
            if d in (1, 3, 5, 7, 9):
                return "M"
            if d in (2, 4, 6, 8, 0):
                return "F"
        # 짧게 뒷자리만 입력한 경우: "-1" "1" 등
        elif len(s) == 1 and s.isdigit():
            d = int(s)
            if d in (1, 3, 5, 7, 9):
                return "M"
            if d in (2, 4, 6, 8, 0):
                return "F"
    return ""

# 전화 추가 컬럼 — 여러 변형 모두 수집 (보조번호 여러 개 가능).
# 이동전화2/이동전화3/집전화 등이 추가 전화로 감지됨.
_DC_EXTRA_PHONE_ALIASES = [
    "연락처2", "연락처 2", "연락처3",
    "전화2", "전화 2", "전화번호2",
    "휴대폰2", "휴대폰 2", "휴대폰3", "핸드폰2", "핸드폰 2", "hp2",
    "phone2", "phone 2", "mobile2",
    "이동전화2", "이동전화 2", "이동전화3", "이동전화 3",
    "집전화", "집 전화", "자택전화", "자택 전화", "유선전화", "유선", "tel",
    "추가번호", "추가 번호", "보조연락처", "보조 연락처", "보조번호",
    "sub_phone", "secondary", "secondary_phone", "extra", "extra_phone",
]


def _dc_find_header_row(rows):
    """헤더가 1행이 아닐 수 있음(타이틀/설명 행이 있는 경우) → 최대 10행까지 스캔해서
    'name' alias 가 포함된 첫 행을 헤더로 선택."""
    name_aliases_low = set(a.lower() for a in _DC_HEADER_ALIASES["name"])
    for i, row in enumerate(rows[:10]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(c in name_aliases_low for c in cells):
            return i
    return 0  # 못 찾으면 1행으로 fallback (기존 동작)


def _dc_normalize_phone(s):
    """전화번호 포맷 통일. 형식 불명이면 원본 반환."""
    if not s:
        return None
    import re as _re
    digits = _re.sub(r"[^0-9]", "", str(s))
    if not digits:
        return None
    if len(digits) == 11 and digits.startswith("010"):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10 and digits[:2] in ("02",):
        return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return str(s).strip()


def _dc_is_valid_mobile(s):
    """이동전화(휴대폰) 유효성 판정 — 한국 010/011/016~019."""
    if not s:
        return False
    import re as _re
    d = _re.sub(r"[^0-9]", "", str(s))
    if d.startswith("010") and len(d) == 11:
        return True
    if d.startswith(("011", "016", "017", "018", "019")) and len(d) in (10, 11):
        return True
    return False


def _dc_split_phones(raw_list):
    """전화 후보 셀들을 모아 (메인, 보조이동전화, 집전화, 형식오류) 로 분리.

    - 이동전화(010/011/016~019) 여러 개면 첫 번째 = main, 나머지 = extra_mobiles
    - 이동전화가 하나도 없으면 집전화/지역번호를 main 으로 대체
    - 이동전화가 있는데 집전화도 있으면 집전화는 landlines (보조)
    - 완전 형식 오류 문자열 → invalid

    반환: (main, extra_mobiles, landlines, invalid)
      - main        : 메인 연락처 (이동전화 우선)
      - extra_mobiles: 메인 제외 추가 이동전화 — 있으면 "검토 필요" 사유
      - landlines   : 집전화/지역번호 (검토 필요 아님, memo 에 보조로 기록)
      - invalid     : 형식 불명
    """
    mobiles, landlines, invalid = [], [], []
    seen = set()
    for raw in raw_list:
        if raw is None or str(raw).strip() == "":
            continue
        norm = _dc_normalize_phone(raw)
        if norm is None:
            continue
        if norm in seen:
            continue
        seen.add(norm)
        if _dc_is_valid_mobile(raw):
            mobiles.append(norm)
        else:
            import re as _re
            d = _re.sub(r"[^0-9]", "", str(raw))
            if len(d) in (9, 10, 11) and d.startswith("0"):
                landlines.append(norm)  # 지역번호 · 집전화
            else:
                invalid.append(str(raw).strip())
    if mobiles:
        main = mobiles[0]
        extra_mobiles = mobiles[1:]
    elif landlines:
        # 이동전화가 없으면 집전화를 메인으로 대체
        main = landlines[0]
        extra_mobiles = []
        landlines = landlines[1:]
    else:
        main = None
        extra_mobiles = []
    return main, extra_mobiles, landlines, invalid


def _dc_normalize_date(v):
    """여러 포맷(1994-01-31 / 19940131 / 94.01.31 / datetime 객체) → YYYY-MM-DD."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    import re as _re
    m = _re.match(r"^(\d{4})[.\-/ ]?(\d{1,2})[.\-/ ]?(\d{1,2})$", s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    m = _re.match(r"^(\d{2})[.\-/ ]?(\d{1,2})[.\-/ ]?(\d{1,2})$", s)
    if m:
        yy = int(m.group(1))
        y = 1900 + yy if yy >= 30 else 2000 + yy
        return f"{y}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    return s  # 원본 그대로 (알려진 포맷 아님 — UI 에서 표시는 됨)


def _dc_dupe_key_in_file(e):
    """파일 내부 중복 판정용 키 — 차트번호 우선."""
    if e.get("chart_no"):
        return ("chart", e["chart_no"])
    return (
        "key",
        (e.get("name") or "").strip(),
        (e.get("birth_date") or "").strip(),
        ((e.get("phone") or "").strip())[-4:],
    )


def _dc_is_duplicate_in_db(e, chart_set, key_set):
    """DB 기존 환자 대비 중복 판정."""
    if e.get("chart_no") and e["chart_no"] in chart_set:
        return True
    key = (
        (e.get("name") or "").strip(),
        (e.get("birth_date") or "").strip(),
        ((e.get("phone") or "").strip())[-4:],
    )
    if key[0] and key in key_set:
        return True
    return False


def _dc_log(msg: str):
    """데이터변환 진행 로그 (콘솔)."""
    import sys, datetime as _dt
    try:
        print(f"[DC {_dt.datetime.now().strftime('%H:%M:%S')}] {msg}",
              file=sys.stderr, flush=True)
    except Exception:
        pass


def _dc_parse_csv_fallback(file_bytes: bytes):
    """xlsx 파싱 불가 시 CSV 로 fallback. 파이썬 기본 csv 모듈만 사용(라이브러리 의존 없음).

    - UTF-8 / UTF-8 BOM / CP949 순으로 디코딩 시도
    - 구분자는 자동 감지 (, \t ;)
    """
    import csv as _csv
    import io as _io
    text = None
    last_err = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = file_bytes.decode(enc)
            break
        except Exception as e:
            last_err = e
    if text is None:
        raise HTTPException(400, f"CSV 텍스트 디코딩 실패: {last_err}")
    # 구분자 자동 감지
    sample = text[:2048]
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except Exception:
        dialect = _csv.excel
    reader = _csv.reader(_io.StringIO(text), dialect)
    rows = [tuple(row) for row in reader]
    return rows


def _dc_parse_excel(file_bytes: bytes):
    """엑셀 → [{row, name, chart_no, phone, birth_date, gender}, ...] + errors.

    openpyxl 미설치 / 배포본 누락 / 파일 손상 등 여러 실패 상황을 모두 감지해서
    가능하면 CSV fallback 으로 계속 진행하고, 불가 시 명확한 안내 메시지 반환.
    """
    import io as _io
    # ── 1단계: xlsx 파일 여부 확인 (ZIP 헤더 PK\x03\x04) ──
    looks_like_xlsx = len(file_bytes) >= 4 and file_bytes[:4] == b'PK\x03\x04'
    rows = None
    fallback_used = False
    fallback_reason = None

    if looks_like_xlsx:
        # 정상적인 xlsx → openpyxl 로 파싱
        try:
            import openpyxl  # 지연 import
        except ImportError as e:
            _dc_log(f"openpyxl 로드 실패 ({e}) — CSV fallback 시도")
            fallback_reason = "openpyxl 라이브러리가 배포본에 포함되지 않음 (PyInstaller 번들 누락)"
            # CSV 로 fallback 시도 (실패 가능성 높음, xlsx 는 CSV 가 아님)
            try:
                rows = _dc_parse_csv_fallback(file_bytes)
                fallback_used = True
            except Exception:
                raise HTTPException(
                    500,
                    "엑셀 파싱 라이브러리(openpyxl)가 이 배포본에 포함되지 않았습니다.\n"
                    "임시 해결: 엑셀을 CSV(UTF-8) 형식으로 저장해서 다시 업로드해주세요.\n"
                    "영구 해결: 최신 배포본(v1.2.1 이상)으로 교체 후 재시도."
                )
        else:
            try:
                wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception as e:
                _dc_log(f"openpyxl 파싱 실패 ({e}) — CSV fallback 시도")
                fallback_reason = f"엑셀 파일 파싱 실패: {e}"
                try:
                    rows = _dc_parse_csv_fallback(file_bytes)
                    fallback_used = True
                except Exception:
                    raise HTTPException(400, f"엑셀 파일을 열 수 없습니다: {e}")
    else:
        # 파일 자체가 xlsx 가 아님 → CSV 로 시도
        _dc_log("파일이 xlsx 형식이 아님 — CSV 로 파싱 시도")
        try:
            rows = _dc_parse_csv_fallback(file_bytes)
            fallback_used = True
            fallback_reason = "업로드 파일이 xlsx 가 아니라 CSV 로 인식됨"
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"엑셀/CSV 모두 파싱 실패: {e}")

    if fallback_used:
        _dc_log(f"[FALLBACK] CSV 모드로 파싱 진행 · 사유: {fallback_reason}")
    rows = rows or []
    if not rows:
        raise HTTPException(400, "엑셀 파일에 데이터가 없습니다.")

    # 헤더가 1행이 아닐 수 있음(타이틀/설명 섞인 엑셀) → 자동 탐지
    header_row_idx = _dc_find_header_row(rows)
    header = [str(c).strip().lower() if c is not None else "" for c in rows[header_row_idx]]
    col_idx = {}
    for key, aliases in _DC_HEADER_ALIASES.items():
        low_al = [a.lower() for a in aliases]
        for i, h in enumerate(header):
            if h in low_al:
                col_idx[key] = i
                break

    # 추가 전화 컬럼들 (보조연락처/전화2/추가번호 등) — 여러 개 수집
    extra_phone_cols = []
    low_extras = [a.lower() for a in _DC_EXTRA_PHONE_ALIASES]
    for i, h in enumerate(header):
        if h in low_extras and i != col_idx.get("phone"):
            extra_phone_cols.append(i)

    if "name" not in col_idx:
        raise HTTPException(
            400,
            "엑셀 헤더에 '이름'/'성명'/'환자명' 중 하나가 있어야 합니다. "
            f"현재 헤더: {header}"
        )

    entries = []
    errors = []
    # 데이터는 헤더 바로 다음 행부터
    data_start = header_row_idx + 1
    for r_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        def _get(key):
            i = col_idx.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]
        raw_name = _get("name")
        name = (str(raw_name).strip() if raw_name not in (None, "") else "")
        if not name:
            continue  # 빈 행 스킵
        chart_raw = _get("chart_no")
        chart = (str(chart_raw).strip() if chart_raw not in (None, "") else None)

        # 전화 후보: 메인 컬럼 + 추가 컬럼 전부 모으기
        phone_candidates = [_get("phone")]
        for i in extra_phone_cols:
            if i < len(row):
                phone_candidates.append(row[i])
        main_phone, extra_mobiles, landlines, invalid_phones = _dc_split_phones(phone_candidates)
        # 기존 UI 호환: extra_phones = 이동+집전화 합본 (표 표시용)
        extra_phones_all = list(extra_mobiles) + list(landlines)

        birth_raw = _get("birth_date")
        birth = _dc_normalize_date(birth_raw)
        birth_format_bad = False
        if birth_raw not in (None, "") and birth:
            import re as _re
            if not _re.match(r"^\d{4}-\d{2}-\d{2}$", birth):
                birth_format_bad = True

        # 성별 — 컬럼 우선, 없으면 주민번호 뒷자리로 보조 판정
        raw_gender = _get("gender")
        raw_ssn = _get("ssn")
        gender = _dc_normalize_gender(raw_gender, raw_ssn)
        # 파일에 성별·주민번호 컬럼이 있는지 (검토 사유 판정용)
        has_gender_source = ("gender" in col_idx) or ("ssn" in col_idx)

        entries.append({
            "row": r_idx, "name": name,
            "chart_no": chart,
            "phone": main_phone,
            "extra_phones": extra_phones_all,       # UI 표시 호환
            "extra_mobiles": extra_mobiles,         # 검토 필요 판정용 (이동전화만)
            "landlines": landlines,                 # 집전화 (검토 대상 아님)
            "invalid_phones": invalid_phones,
            "birth_date": birth,
            "birth_format_bad": birth_format_bad,
            "gender": gender,
            "_has_gender_source": has_gender_source,
        })
    parse_info = {
        "fallback_used": fallback_used,
        "fallback_reason": fallback_reason or "",
        "mode": "csv" if fallback_used else "xlsx",
    }
    return entries, errors, header, parse_info


def _dc_classify_review(entry, file_key_counter):
    """검토 필요 사유 목록 반환 (사람이 읽는 문자열 리스트).

    사용자 요구:
      - 이동전화 2개 (다중 연락처)
      - 필수값 누락
      - 중복 의심 (파일 내 같은 이름+생일인데 차트/전화가 어긋남)
      - 형식 오류
    """
    reasons = []
    # ● 검토 필요는 "이동전화가 2개 이상" 일 때만 (집전화는 보조로 저장, 검토 대상 아님)
    extra_mobiles = entry.get("extra_mobiles") or []
    if extra_mobiles:
        reasons.append(f"다중 이동전화 · 확인 필요 ({len(extra_mobiles)+1}개)")
    if not entry.get("phone"):
        if entry.get("invalid_phones"):
            reasons.append("연락처 형식 오류 (원본 유지)")
        else:
            reasons.append("연락처 없음")
    if not entry.get("birth_date"):
        reasons.append("생년월일 없음")
    elif entry.get("birth_format_bad"):
        reasons.append("생년월일 형식 확인")
    # 성별 미확정: 파일에 성별/주민번호 컬럼이 있을 때만 경고 (AI 추측 금지).
    # 엑셀 원본에 성별 정보가 아예 없으면 경고 안 함 → 운영자가 나중에 개별 편집.
    if entry.get("_has_gender_source") and not entry.get("gender"):
        reasons.append("성별 미확정")
    # 파일 내부 중복 의심
    name_dob_key = (entry.get("name") or "", entry.get("birth_date") or "")
    if file_key_counter.get(name_dob_key, 0) >= 2 and not entry.get("chart_no"):
        reasons.append("중복 의심 (같은 이름/생일)")
    return reasons


@router.post("/data-convert/preview")
async def data_convert_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: bool = Depends(require_admin),
):
    """엑셀 업로드 → 파싱 + 중복 판정. DB에는 아직 반영하지 않음.

    중복 판정 순서:
      1) 차트번호 있으면: DB의 chart_no 집합과 비교
      2) 없으면: (name, birth_date, phone 뒤 4자리) 튜플로 비교
    """
    content = await file.read()
    if not content:
        raise HTTPException(400, "빈 파일입니다.")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "파일이 너무 큽니다 (최대 10MB).")

    entries, errors, header, parse_info = _dc_parse_excel(content)

    # 기존 환자 색인
    pats = db.query(models.Patient).all()
    chart_set = {p.chart_no for p in pats if p.chart_no}
    key_set = {
        ((p.name or "").strip(), (p.birth_date or ""), ((p.phone or "")[-4:]))
        for p in pats if p.name
    }

    # 파일 내 (이름, 생일) 빈도 — 중복 의심 판정용
    from collections import Counter as _Counter
    name_dob_counter = _Counter(
        ((e.get("name") or ""), (e.get("birth_date") or ""))
        for e in entries
    )

    new_list = []
    existing_count = 0
    dup_in_file = set()
    for e in entries:
        k = _dc_dupe_key_in_file(e)
        if k in dup_in_file:
            existing_count += 1  # 파일 내부 중복
            continue
        dup_in_file.add(k)
        if _dc_is_duplicate_in_db(e, chart_set, key_set):
            existing_count += 1
        else:
            # 검토 사유 태깅 (AI 보조 — 규칙 기반 분류)
            reasons = _dc_classify_review(e, name_dob_counter)
            e["review_reasons"] = reasons
            e["review_reason"] = " · ".join(reasons) if reasons else ""
            new_list.append(e)

    review_list = [e for e in new_list if e.get("review_reasons")]

    return {
        "total": len(entries),
        "new_count": len(new_list),
        "existing_count": existing_count,
        "error_count": len(errors),
        "header": header,
        "new_patients": new_list,
        "review_list": review_list,
        "review_count": len(review_list),
        "errors": errors,
        "file_name": file.filename,
        "parse_info": parse_info,
    }


@router.post("/data-convert/apply")
def data_convert_apply(
    payload: dict,
    db: Session = Depends(get_db),
    _: bool = Depends(require_admin),
):
    """preview 결과 중 new_patients 를 DB 에 insert.
    payload: { "items": [{name, chart_no, phone, birth_date}, ...] }
    적용 직전 DB 중복 재검증으로 동시성 안전.
    """
    items = payload.get("items", []) or []
    if not isinstance(items, list):
        raise HTTPException(400, "items 는 리스트여야 합니다.")

    # 적용 직전 DB 색인 다시 구축
    pats = db.query(models.Patient).all()
    chart_set = {p.chart_no for p in pats if p.chart_no}
    key_set = {
        ((p.name or "").strip(), (p.birth_date or ""), ((p.phone or "")[-4:]))
        for p in pats if p.name
    }

    # ── 1차 패스: 행 별 정제 + 중복 재검증 + memo 구성 ──
    import uuid as _uuid
    from datetime import datetime as _dt
    now = _dt.utcnow()

    inserted = []
    skipped = []
    review_inserted = 0
    bulk_rows = []  # bulk_insert_mappings 용

    for it in items:
        name = (it.get("name") or "").strip()
        if not name:
            skipped.append({"reason": "이름 없음", "item": it})
            continue
        chart = (it.get("chart_no") or "").strip() or None
        phone = (it.get("phone") or "").strip() or None
        bd    = (it.get("birth_date") or "").strip() or None
        # 성별: M/F 이외는 빈값으로 정규화
        gender = (it.get("gender") or "").strip().upper()
        if gender not in ("M", "F"):
            gender = ""
        extras = it.get("extra_phones") or []
        if not isinstance(extras, list):
            extras = []
        extras = [str(x).strip() for x in extras if str(x).strip()]
        review_reasons = it.get("review_reasons") or []
        if not isinstance(review_reasons, list):
            review_reasons = []

        e = {"name": name, "chart_no": chart, "phone": phone, "birth_date": bd}
        if _dc_is_duplicate_in_db(e, chart_set, key_set):
            skipped.append({"reason": "중복 (동시 등록)", "item": it})
            continue

        memo_lines = []
        if extras:
            memo_lines.append("[보조 연락처] " + ", ".join(extras))
        if review_reasons:
            memo_lines.append("[검토 필요] " + " · ".join(review_reasons))
        memo = "\n".join(memo_lines) if memo_lines else ""

        new_id = _uuid.uuid4().hex
        bulk_rows.append({
            "id": new_id, "name": name, "chart_no": chart, "phone": phone,
            "birth_date": bd, "gender": gender, "memo": memo,
            "created_at": now, "updated_at": now,
        })
        if review_reasons:
            review_inserted += 1
        inserted.append({
            "id": new_id, "name": name, "chart_no": chart,
            "phone": phone, "birth_date": bd,
            "gender": gender,
            "extra_phones": extras,
            "review_reasons": review_reasons,
        })
        if chart:
            chart_set.add(chart)
        key_set.add((name, bd or "", (phone or "")[-4:]))

    # ── 2차 패스: bulk insert (청크 단위) — 80K 기준 ~5배 빠름 ──
    BULK_CHUNK = 2000
    for i in range(0, len(bulk_rows), BULK_CHUNK):
        db.bulk_insert_mappings(models.Patient, bulk_rows[i:i + BULK_CHUNK])
        db.commit()

    audit(db, "patient.bulk_import", "",
          f"AI 데이터변환 {len(inserted)}명 추가 (검토필요 {review_inserted}) / {len(skipped)}건 건너뜀")
    db.commit()
    return {
        "inserted": len(inserted),
        "review_inserted": review_inserted,
        "skipped": len(skipped),
        "inserted_patients": inserted,
        "skipped_items": skipped,
    }


# ──────────────── 단계 G #18: 자동 백업 관리 ────────────────

@router.get("/backup/list")
def backup_list(_: bool = Depends(require_admin)):
    """백업 파일 목록 (최신순)."""
    from ..services.backup import list_backups
    return list_backups()


@router.post("/backup/now")
def backup_now_v2(db: Session = Depends(get_db), _: bool = Depends(require_admin)):
    """수동 백업 (지금 백업)."""
    from ..services.backup import make_backup
    result = make_backup()
    if not result.get("ok"):
        raise HTTPException(500, result.get("error", "백업 실패"))
    audit(db, "backup.manual", "", result.get("name", "")); db.commit()
    return result

@router.get("/backup/dir")
def backup_dir_path(_: bool = Depends(require_admin)):
    """백업 폴더 경로 반환."""
    return {"path": str(get_backup_dir())}

@router.post("/backup/restore-latest")
def backup_restore_latest(db: Session = Depends(get_db),
                          _: bool = Depends(require_admin)):
    """가장 최근 백업으로 복원 (안전망 백업 1회 자동 생성)."""
    from ..services.backup import restore_latest
    result = restore_latest()
    if not result.get("ok"):
        raise HTTPException(500, result.get("error", "복원 실패"))
    audit(db, "backup.restore_latest", "", result.get("restored_from", ""))
    db.commit()
    return result

@router.post("/backup/restore-by-name")
def backup_restore_by_name(payload: dict, db: Session = Depends(get_db),
                           _: bool = Depends(require_admin)):
    """지정한 파일명의 백업으로 복원 (안전망 백업 자동 생성)."""
    from ..services.backup import restore_by_name
    filename = (payload or {}).get("filename", "").strip()
    if not filename:
        raise HTTPException(400, "filename 이 필요합니다.")
    result = restore_by_name(filename)
    if not result.get("ok"):
        raise HTTPException(500, result.get("error", "복원 실패"))
    audit(db, "backup.restore_by_name", "", result.get("restored_from", ""))
    db.commit()
    return result

# ──────────────── 감사 로그 ────────────────

@router.get("/audit-logs")
def list_audit_logs(limit: int = 200, db: Session = Depends(get_db),
                    _: bool = Depends(require_admin)):
    rows = (db.query(models.AuditLog)
            .order_by(models.AuditLog.ts.desc()).limit(limit).all())
    return [{"id": r.id, "ts": r.ts.isoformat(), "node_id": r.node_id,
             "actor": r.actor, "action": r.action, "entity_id": r.entity_id,
             "detail": r.detail} for r in rows]


# ──────────────── SMS 설정 ────────────────

def _get_sms_setting(db):
    obj = db.query(models.SmsSetting).filter_by(id=1).first()
    if not obj:
        obj = models.SmsSetting(id=1)
        db.add(obj); db.commit(); db.refresh(obj)
    return obj


@router.get("/sms/setting")
def sms_get(db: Session = Depends(get_db)):
    obj = _get_sms_setting(db)
    return {
        "munjanara_id": obj.munjanara_id,
        "munjanara_pw": "****" if obj.munjanara_pw else "",
        "munjanara_key": obj.munjanara_key[:4] + "****" if obj.munjanara_key else "",
        "sender_phone": obj.sender_phone,
        "clinic_phone": obj.clinic_phone,
        "clinic_name": obj.clinic_name,
        # 문자나라 발송 URL — 관리자가 업체에서 받은 값을 입력
        "api_url": getattr(obj, "api_url", "") or "",
    }

@router.post("/sms/setting")
def sms_set(payload: dict, db: Session = Depends(get_db),
            _: bool = Depends(require_admin)):
    """문자나라 설정 저장.

    v1.2.16 (2026-04-25): 비밀번호 류 빈 값 가드 추가.
    - 페이지 새로고침/업데이트 후 password input 은 보안상 항상 빈 값으로
      그려짐. 그 상태에서 사용자가 다른 필드만 수정해 저장 누르면 빈
      munjanara_pw / munjanara_key 가 같이 PUT 되어 기존 비번이 빈 값으로
      덮어씌워짐 → "업데이트하면 로그인 정보가 초기화" 처럼 보이는 버그.
    - 수정: 비밀번호 류는 빈 값이 들어오면 "사용자가 안 건드림" 으로 해석해
            기존 DB 값을 보존. (마스킹된 값 ****와 동일 정책)
    """
    obj = _get_sms_setting(db)
    PASSWORD_KEYS = {"munjanara_pw", "munjanara_key"}  # 빈 값으로 덮어쓰지 않을 필드
    for k in ("munjanara_id", "munjanara_pw", "munjanara_key", "sender_phone",
              "clinic_phone", "clinic_name", "api_url"):
        if k not in payload or payload[k] is None:
            continue
        val = str(payload[k])
        # 마스킹된 값 (****로 시작) 은 항상 "수정 안 함"
        if val.startswith("****"):
            continue
        # 비밀번호 류: 빈 값도 "수정 안 함" 으로 해석 (기존 DB 값 유지)
        if k in PASSWORD_KEYS and val == "":
            continue
        setattr(obj, k, val)
    audit(db, "sms.setting_update", "", "문자나라 설정 변경"); db.commit()
    return {"ok": True}


# ─────── 단계 F #13: {다음예약항목} 정규화 (도수치료30/60 → 도수치료) ───────
def _normalize_tx_name_for_sms(name: str) -> str:
    """문자 본문 표시용으로 도수치료 시간 수치 제거."""
    if not name:
        return name
    # "도수치료30분" / "도수치료60분" / "도수치료 30분" 등 → "도수치료"
    import re
    cleaned = re.sub(r'(도수치료)\s*\d+\s*분', r'\1', name)
    return cleaned


def _format_tx_summary_for_sms(codes, db) -> str:
    """치료항목 코드 리스트 → 문자용 짧은 표시 ("도수치료, 체외충격파")."""
    names = []
    seen = set()
    by_code = {t.code: t for t in db.query(models.Treatment).all()}
    for c in codes:
        t = by_code.get(c)
        if not t: continue
        normalized = _normalize_tx_name_for_sms(t.name)
        if normalized in seen: continue
        seen.add(normalized)
        names.append(normalized)
    return ", ".join(names) if names else ""


@router.get("/sms/tomorrow-targets")
def sms_tomorrow(db: Session = Depends(get_db)):
    t_local = (datetime.now() + timedelta(days=1)).date()
    ts = datetime.combine(t_local, datetime.min.time())
    te = ts + timedelta(days=1)
    rows = (db.query(models.Appointment)
            .filter(models.Appointment.start_at >= ts,
                    models.Appointment.start_at < te,
                    models.Appointment.status != "canceled").all())
    setting = _get_sms_setting(db)
    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
    out = []
    for a in rows:
        if not a.patient or not a.patient.phone:
            continue
        tt = a.start_at
        wd = weekdays[tt.weekday()]
        # 단계 F #12: 치료항목 추가
        codes = _parse_codes(a.treatment_codes or "[]")
        tx_summary = _format_tx_summary_for_sms(codes, db)
        tx_part = f" {tx_summary}" if tx_summary else ""
        body = (f"[{setting.clinic_name}] {a.patient.name} 님, "
                f"내일({tt.month}/{tt.day} {wd}) {tt.hour:02d}:{tt.minute:02d}{tx_part} "
                f"예약이 있습니다. 변경/취소는 {setting.clinic_phone}")
        out.append({
            "appointment_id": a.id,
            "patient_id": a.patient_id,
            "chart_no": a.patient.chart_no or "-",
            "name": a.patient.name, "phone": a.patient.phone,
            "reserved_at": a.start_at.isoformat(), "body": body,
            "treatment_summary": tx_summary,
        })
    out.sort(key=lambda x: x["reserved_at"])
    return out


# ──────────────── 문자 템플릿 CRUD (단계 F #15) ────────────────

def _serialize_sms_template(t: models.SmsTemplate) -> dict:
    return {
        "id": t.id,
        "name": t.name,
        "body": t.body or "",
        "sort_order": t.sort_order or 0,
        "active": bool(t.active),
        "updated_at": t.updated_at.isoformat() if t.updated_at else None,
    }


@router.get("/sms/templates")
def list_sms_templates(db: Session = Depends(get_db)):
    items = (db.query(models.SmsTemplate)
             .order_by(models.SmsTemplate.sort_order, models.SmsTemplate.name).all())
    return [_serialize_sms_template(t) for t in items]


@router.post("/sms/templates")
def create_sms_template(payload: dict, db: Session = Depends(get_db),
                        _: bool = Depends(require_admin)):
    base_name = (payload.get("name") or "").strip() or "새 템플릿"
    body = payload.get("body") or ""
    # 본문 비어있으면 1번 템플릿 본문 자동 채움 (#15)
    if not body.strip():
        first = (db.query(models.SmsTemplate)
                 .order_by(models.SmsTemplate.sort_order, models.SmsTemplate.name).first())
        if first:
            body = first.body or ""
    # 중복 이름 자동 넘버링: "새 템플릿", "새 템플릿1", "새 템플릿2" ...
    existing_names = {n for (n,) in db.query(models.SmsTemplate.name).all()}
    name = base_name
    if name in existing_names:
        i = 1
        while f"{base_name}{i}" in existing_names:
            i += 1
        name = f"{base_name}{i}"
    sort_order = db.query(models.SmsTemplate).count() + 1
    t = models.SmsTemplate(
        name=name, body=body,
        sort_order=sort_order, active=True,
    )
    db.add(t); db.flush()
    audit(db, "sms_template.create", t.id, name)
    db.commit()
    return _serialize_sms_template(t)


@router.put("/sms/templates/{tid}")
def update_sms_template(tid: str, payload: dict, db: Session = Depends(get_db),
                        _: bool = Depends(require_admin)):
    t = db.get(models.SmsTemplate, tid)
    if not t:
        raise HTTPException(404, "템플릿 없음")
    if "name" in payload: t.name = (payload["name"] or "").strip() or t.name
    if "body" in payload: t.body = payload["body"] or ""
    if "active" in payload: t.active = bool(payload["active"])
    if "sort_order" in payload: t.sort_order = int(payload["sort_order"]) or t.sort_order
    db.flush()
    audit(db, "sms_template.update", tid, t.name)
    db.commit()
    return _serialize_sms_template(t)


@router.delete("/sms/templates/{tid}")
def delete_sms_template(tid: str, db: Session = Depends(get_db),
                        _: bool = Depends(require_admin)):
    t = db.get(models.SmsTemplate, tid)
    if not t:
        raise HTTPException(404, "템플릿 없음")
    # 마지막 1개는 삭제 거부 (안전망)
    if db.query(models.SmsTemplate).count() <= 1:
        raise HTTPException(400, "마지막 템플릿은 삭제할 수 없습니다.")
    audit(db, "sms_template.delete", tid, t.name)
    db.delete(t)
    db.commit()
    return {"ok": True}


def _normalize_phone_for_sms(raw: str) -> str:
    """수신/발신 번호 → 숫자만. 하이픈/공백/점/괄호/+ 모두 제거."""
    if not raw:
        return ""
    import re
    return re.sub(r"[^0-9]", "", str(raw))


def _is_valid_kr_mobile(digits: str) -> bool:
    """한국 휴대폰/전화 번호 형식 체크 (정규화 후 숫자만 문자열)."""
    if not digits:
        return False
    # 휴대폰 010/011/016~019 (10~11자리), 02 (9~10자리), 지역번호 03x~06x/07x/08x (10~11자리)
    if digits.startswith("010") and len(digits) == 11:
        return True
    if digits.startswith(("011", "016", "017", "018", "019")) and len(digits) in (10, 11):
        return True
    if digits.startswith("02") and len(digits) in (9, 10):
        return True
    if len(digits) in (10, 11) and digits[0] == "0":
        return True
    return False


def _mask_phone_for_log(p: str) -> str:
    """로그에 찍을 전화번호 일부 마스킹 (끝 4자리만 노출)."""
    if not p:
        return "(없음)"
    import re as _re
    d = _re.sub(r"[^0-9]", "", str(p))
    if len(d) >= 4:
        return f"***-****-{d[-4:]}"
    return "***"


def _sms_log(line: str):
    """콘솔에 SMS 발송 추적 로그 (stderr, 시각 포함)."""
    import sys, datetime as _dt
    msg = f"[SMS {_dt.datetime.now().strftime('%H:%M:%S')}] {line}"
    try:
        print(msg, file=sys.stderr, flush=True)
    except Exception:
        pass


def _sms_sanitize(text: str, secrets: list) -> str:
    """SMS 관련 텍스트(예외 메시지/서버 응답)에 평문 비밀이 섞여 있어도 마스킹.

    적용 대상:
      - urllib / socket 예외 메시지에 요청 body 일부가 끼어 들어오는 경우
      - SMS 서버가 요청 echo 형태로 응답해 passwd/key 값을 그대로 돌려보내는 경우

    secrets 인자는 마스킹할 비밀 문자열들의 리스트. 빈 값/None 은 스킵.
    """
    if not text:
        return text
    out = str(text)
    for s in secrets:
        if not s:
            continue
        s = str(s)
        # 너무 짧은 비밀은 마스킹 시 의미없는 치환 폭증 → 4자 미만은 스킵
        if len(s) < 4:
            continue
        out = out.replace(s, "***")
    return out


def _smart_decode_response(resp_or_headers, raw_bytes: bytes) -> str:
    """HTTP 응답 본문을 한글 안전하게 디코딩.

    순서:
      1) Content-Type 의 charset 지시를 우선 (예: "text/plain; charset=euc-kr")
      2) UTF-8 → EUC-KR → CP949 순으로 엄격 시도
      3) 다 실패하면 UTF-8 replace (깨진 글자만 치환, 나머지는 살림)

    ⚠ 이전 구현의 `decode("utf-8", errors="ignore")` 는 서버가 EUC-KR 로
       응답할 경우 한글 바이트가 전부 무시돼 응답 메시지가 비거나 깨져
       들어오는 문제 발생 → 이 함수로 통일.
    """
    try:
        ct = ""
        if hasattr(resp_or_headers, "headers"):
            ct = resp_or_headers.headers.get("Content-Type") or ""
        elif hasattr(resp_or_headers, "get"):
            ct = resp_or_headers.get("Content-Type") or ""
        ct = ct.lower()
        declared = None
        if "charset=" in ct:
            try:
                declared = ct.split("charset=")[1].split(";")[0].strip()
            except Exception:
                declared = None
        candidates = []
        if declared:
            candidates.append(declared)
        for enc in ("utf-8", "euc-kr", "cp949"):
            if enc not in candidates:
                candidates.append(enc)
        for enc in candidates:
            try:
                return raw_bytes.decode(enc, errors="strict")
            except Exception:
                continue
        # 어느 것도 엄격 디코딩 안 되면 utf-8 replace 로 최소한 정보 보존
        return raw_bytes.decode("utf-8", errors="replace")
    except Exception:
        return raw_bytes.decode("utf-8", errors="replace")


@router.post("/sms/send")
def sms_send(payload: dict, db: Session = Depends(get_db)):
    """문자나라 발송.

    개선점:
      - API URL 을 SmsSetting.api_url 에서 읽음 (관리자가 설정한 실제 URL 사용)
      - HTTPError 404/5xx 를 상태코드와 함께 detail 에 기록
      - 콘솔 로그에 시각·마스킹 번호·URL·메소드·상태코드·판정 남김
      - 프론트가 식별할 수 있도록 result_code (kind) 도 반환
    """
    import urllib.request, urllib.parse, urllib.error
    setting = _get_sms_setting(db)
    # 문자나라 웹 API 연동 은 "2차 비밀번호" 를 passwd 로 요구함 (일반 로그인 비번 아님).
    # 2026-04-18 확인: 웹사이트 연동 설정 안내 "2차비번 입력" 기준.
    missing = []
    if not setting.munjanara_id: missing.append("아이디")
    if not setting.munjanara_key: missing.append("2차 비밀번호 (API 인증용)")
    if not setting.sender_phone: missing.append("발신번호")
    if not getattr(setting, "api_url", None):
        missing.append("API URL")
    if missing:
        raise HTTPException(400,
            "문자나라 설정을 먼저 완료하세요 (관리자 → 문자나라)\n"
            "누락 항목: " + ", ".join(missing))

    api_url = (setting.api_url or "").strip()
    items = payload.get("items", [])
    sender_digits = _normalize_phone_for_sms(setting.sender_phone)
    results = []

    # 로깅/DB 저장 시 평문 비밀이 섞여 있으면 마스킹할 대상.
    # urllib 예외 메시지 / 서버 echo 응답에 password/key 가 끼어들어가는 사고를 차단.
    _secrets = [getattr(setting, "munjanara_key", None), getattr(setting, "munjanara_pw", None)]

    _sms_log(f"── 발송 요청 시작 · 대상 {len(items)}건 · URL={api_url}")

    for it in items:
        raw_phone = (it.get("phone") or "").strip()
        rphone_digits = _normalize_phone_for_sms(raw_phone)
        masked = _mask_phone_for_log(raw_phone)

        # 사전 유효성 검증
        if not _is_valid_kr_mobile(rphone_digits):
            detail = f"연락처 형식 오류: '{raw_phone}' (숫자 {len(rphone_digits)}자리)"
            _sms_log(f"  [PRECHECK-FAIL] 번호={masked} · {detail}")
            db.add(models.SmsLog(patient_id=it.get("patient_id"),
                phone=raw_phone, body=it.get("body", ""),
                result="fail", detail=detail))
            results.append({"phone": raw_phone, "result": "fail",
                            "kind": "precheck", "detail": detail})
            continue

        # 2026-04-18 실측으로 확정한 문자나라 웹 API 파라미터 이름:
        #   userid / passwd / sender / receiver / message
        #   ※ 문자나라 "웹 연동" 은 passwd 자리에 **2차 비밀번호** 를 요구함
        #     (일반 로그인 비밀번호 munjanara_pw 는 API 호출에 사용되지 않음)
        params = {
            "userid":   setting.munjanara_id,
            "passwd":   setting.munjanara_key,    # 2차 비번 = API 인증
            "sender":   setting.sender_phone,     # 하이픈 유지/제거 둘 다 인식됨
            "receiver": rphone_digits,
            "message":  it.get("body", ""),
        }
        # v1.2.15 (2026-04-25): 한글 메시지 인코딩 UTF-8 → CP949 로 전환.
        # 문자나라 서버가 요청 헤더의 charset=utf-8 을 무시하고 자체적으로
        # EUC-KR/CP949 로 해석하는 것으로 확인됨 (수신 단말에서 한글이
        # `{? 샟역紐` 같이 한글+한자 섞여 깨지는 전형적 패턴 관측).
        # CP949 = EUC-KR 의 superset (Windows 확장) — 표현 가능 범위가 더 넓음.
        # errors='replace' 로 CP949 로 표현 불가한 문자(이모지 등)는 '?' 로
        # 치환해 발송 자체가 막히지 않도록 안전망 추가.
        qs = urllib.parse.urlencode(params, encoding='cp949', errors='replace')
        # 로그에는 passwd/secure 뒷자리만 노출 (변수명 masked 는 전화 마스킹 변수와 충돌하니 별도 이름)
        masked_params = dict(params)
        if masked_params.get("passwd"):
            masked_params["passwd"] = (masked_params["passwd"][:2] + "***")
        if masked_params.get("secure"):
            masked_params["secure"] = (masked_params["secure"][:2] + "***")
        safe_qs = urllib.parse.urlencode(masked_params, encoding='cp949', errors='replace')
        log_url  = f"{api_url}?{safe_qs}"

        _sms_log(f"  [SEND] 번호={masked} · POST {api_url} · body={safe_qs[:160]}")

        # 실제 호출 — POST 방식 (2026-04-21 변경) + CP949 인코딩 (2026-04-25 변경)
        # 이전 GET 방식은 한글 'message' 가 URL 쿼리스트링에 들어가 일부 단말/서버에서
        # 깨짐(특수문자) 현상 발생 → POST 로 우회.
        # 이전 UTF-8 인코딩은 문자나라 서버가 EUC-KR 로 오해석해 한글이 깨져 도착
        #   → CP949 로 percent-encode + charset=euc-kr 헤더로 변경하여 해결.
        # qs 자체는 이미 percent-encoded 라 .encode("ascii") 로 안전하게 바이트화 가능.
        result = "fail"; detail = ""; kind = "unknown"; status_code = None; resp_text = ""
        try:
            req = urllib.request.Request(
                api_url,
                data=qs.encode("ascii"),
                method="POST",
                headers={
                    "User-Agent": "dosu-clinic-sms/1.0",
                    "Content-Type": "application/x-www-form-urlencoded; charset=euc-kr",
                })
            try:
                with urllib.request.urlopen(req, timeout=10) as r:
                    status_code = r.status
                    _raw = r.read()
                    resp_text = _smart_decode_response(r, _raw)
                # 서버 echo 에 passwd/key 평문이 섞여있을 가능성 차단 — 디코딩 직후 마스킹.
                resp_text = _sms_sanitize(resp_text, _secrets)
            except urllib.error.HTTPError as he:
                # 404 / 500 등 HTTP 에러도 응답 본문은 읽을 수 있음
                status_code = he.code
                try:
                    _raw = he.read()
                    resp_text = _smart_decode_response(he, _raw)
                except Exception:
                    resp_text = ""
                resp_text = _sms_sanitize(resp_text, _secrets)
                if he.code == 404:
                    kind = "url_not_found"
                    detail = (f"HTTP 404 — 문자나라 연동 주소 확인 필요. "
                              f"현재 URL: {api_url}")
                elif 500 <= he.code < 600:
                    kind = "server_error"
                    detail = f"문자나라 서버 오류 {he.code}: {resp_text[:200]}"
                else:
                    kind = "http_error"
                    detail = f"HTTP {he.code}: {resp_text[:200]}"
                _sms_log(f"  [HTTP-ERR] 번호={masked} · status={he.code} · body[:80]={resp_text[:80]!r}")
                db.add(models.SmsLog(patient_id=it.get("patient_id"),
                    phone=raw_phone, body=it.get("body", ""),
                    result="fail", detail=f"[{kind}] {detail}"))
                results.append({"phone": raw_phone, "result": "fail",
                    "kind": kind, "status_code": he.code, "detail": detail})
                continue
            except urllib.error.URLError as ue:
                kind = "network_error"
                # ue.reason 이 socket/SSL 객체일 수도, 문자열일 수도 — 일단 str 변환 후 마스킹.
                reason_safe = _sms_sanitize(str(ue.reason), _secrets)
                detail = f"네트워크 오류: {reason_safe}"
                _sms_log(f"  [NET-ERR] 번호={masked} · {reason_safe}")
                db.add(models.SmsLog(patient_id=it.get("patient_id"),
                    phone=raw_phone, body=it.get("body", ""),
                    result="fail", detail=f"[{kind}] {detail}"))
                results.append({"phone": raw_phone, "result": "fail",
                    "kind": kind, "detail": detail})
                continue

            # 문자나라 응답 형식: "코드|필드1|필드2|...|메시지"
            #   성공 (구식)  : 1|잔여SMS|... 또는 성공 텍스트 포함
            #   성공 (신식)  : 9|<메시지ID>|<발송건수>|<발신자명> ← 실제 운영 환경에서 관측됨
            #                  예) "9|103269|1|이중성"  → 메시지 ID 발급 + 1건 발송 성공
            #                  (첫 필드 9 의 의미는 vendor 측에서 명확히 안내 안 함 —
            #                   계정 등급 / 응답 포맷 버전 / 채널 코드 추정. 어쨌든
            #                   2번째 필드에 5자리 이상의 message id 가 발급되면 성공)
            #   userid 오류 : 2|NOT EXISTS USERID !!
            #   파라미터 오류: 1|MISSING PARAMETER ... (성공처럼 보이지만 MISSING 문구로 구분)
            #   진짜 발송 거부: code=9 + 메시지ID 없음 + 안내문 (발신번호 미등록 / 포인트 부족 등)
            stripped = resp_text.strip()
            fields = stripped.split("|") if stripped else []
            code = fields[0].strip() if fields else ""
            vendor_msg = "|".join(fields[1:]).strip() if len(fields) > 1 else stripped

            # 성공 판정:
            #   1) MISSING/ERROR/FAIL 문구가 없어야 함 (대전제)
            #   2) 그리고 둘 중 하나:
            #      - 첫 필드 code=="1" (구식 성공)
            #      - 또는 두 번째 필드가 메시지ID 형식 (3자리 이상 정수) — 신식 성공
            #      - 또는 응답 본문에 "success" 포함
            upper = resp_text.upper()
            looks_error = any(t in upper for t in ("MISSING", "INVALID", "NOT EXISTS", "ERROR", "FAIL", "EMPTY"))
            second = fields[1].strip() if len(fields) > 1 else ""
            has_msg_id = second.isdigit() and len(second) >= 3
            ok = (not looks_error) and (
                code == "1"
                or has_msg_id
                or "success" in resp_text.lower()
            )

            # 알려진 에러코드 힌트 (확인되지 않은 것은 일반 안내)
            hint = ""
            if not ok:
                if code == "2":
                    hint = " · [힌트] 아이디 또는 2차 비밀번호 확인"
                elif code == "9":
                    # 메시지ID 가 없는 진짜 거부일 때만 힌트 표시
                    hint = (" · [힌트] 계정 인증은 통과. 다음 중 하나 확인 필요: "
                            "①발신번호가 문자나라에 사전등록됐는지 "
                            "②잔여 포인트/건수 부족 "
                            "③일일 발송 한도 초과")
                elif "MISSING" in upper or "EMPTY" in upper:
                    hint = " · [힌트] 파라미터 누락 — 아이디/발신번호/수신번호/메시지 재확인"

            result = "success" if ok else "fail"
            kind = "ok" if ok else "rejected_by_vendor"
            # 필드 분해도 detail 에 포함 (디버그 편의)
            fields_str = " / ".join(f"f{i}={v!r}" for i, v in enumerate(fields[:6]))
            detail = (f"status={status_code} · code={code or '(없음)'} · "
                      f"msg={vendor_msg[:200] or '(없음)'}{hint} "
                      f"· raw[{fields_str}]")
            _sms_log(f"  [{'SENT' if ok else 'REJECT'}] 번호={masked} · status={status_code} · resp[:80]={resp_text[:80]!r} → {result.upper()}")
            db.add(models.SmsLog(patient_id=it.get("patient_id"),
                phone=raw_phone, body=it.get("body", ""),
                result=result, detail=f"[{kind}] {detail}"))
            results.append({"phone": raw_phone, "result": result,
                "kind": kind, "status_code": status_code, "detail": detail})
        except Exception as e:
            kind = "exception"
            # 예외 메시지(특히 UnicodeError / 일부 socket error 류) 에 요청 body 일부가
            # 그대로 끼어 들어올 수 있어 마스킹 — passwd/key 평문이 stderr 와 DB 양쪽에 남는 사고 방지.
            detail = _sms_sanitize(f"예외: {e}", _secrets)
            _sms_log(f"  [EX] 번호={masked} · {detail}")
            db.add(models.SmsLog(patient_id=it.get("patient_id"),
                phone=raw_phone, body=it.get("body", ""),
                result="fail", detail=f"[{kind}] {detail}"))
            results.append({"phone": raw_phone, "result": "fail",
                "kind": kind, "detail": detail})
    db.commit()
    sent = sum(1 for r in results if r["result"] == "success")
    failed = len(items) - sent
    _sms_log(f"── 종료 · 성공 {sent} · 실패 {failed}")
    return {
        "sent": sent, "failed": failed, "total": len(items),
        "results": results,
    }


# ──────────────── 통계 (직원별 월별) ────────────────

@router.get("/stats/by-therapist")
def stats_by_therapist(year: int, month: int, mode: str = "reserved", treatment_code: str = "", db: Session = Depends(get_db)):
    """mode=reserved: 예약 기준 / mode=approved: 완료 기준"""
    from calendar import monthrange
    from collections import defaultdict
    days = monthrange(year, month)[1]
    ts = datetime(year, month, 1)
    te = datetime(year + (1 if month == 12 else 0),
                  (1 if month == 12 else month + 1), 1)
    rows = (db.query(models.Appointment)
            .filter(models.Appointment.start_at >= ts,
                    models.Appointment.start_at < te).all())
    _manual_codes_set = set(_get_manual_therapy_codes(db))
    doctor_codes = _doctor_codes_set(db)
    is_doctor_filter = bool(treatment_code and treatment_code not in ('', 'all', 'manual_all') and treatment_code in doctor_codes)

    def _matches(codes: list) -> bool:
        if not treatment_code or treatment_code == "all":
            return True
        if treatment_code == "manual_all":
            return any(c in _manual_codes_set for c in codes)
        return treatment_code in codes

    manual_rows = _get_manual_treatment_rows(db)
    eswt_row = db.query(models.Treatment).filter(
        models.Treatment.code == C.ESWT_CODE,
        models.Treatment.active == True,
    ).first()
    breakdown_codes = [t.code for t in manual_rows]
    breakdown_names = {t.code: t.name for t in manual_rows}
    if eswt_row:
        breakdown_codes.append(C.ESWT_CODE)
        breakdown_names[C.ESWT_CODE] = eswt_row.name

    daily = defaultdict(lambda: defaultdict(int))
    code_counts = defaultdict(lambda: defaultdict(int))
    canceled = defaultdict(int)
    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        if not _matches(codes):
            continue
        if is_doctor_filter:
            handler_id = next((x.handler_id for x in a.assignments if x.treatment_code == treatment_code), None)
            tid = handler_id or "__none__"
        else:
            tid = a.therapist_id or "__none__"
        if a.status == "canceled":
            canceled[tid] += 1
        elif mode == "approved" and a.status == "approved":
            d = a.start_at.day
            daily[tid][d] += 1
            for c in codes:
                if c in breakdown_codes:
                    if not treatment_code or treatment_code == "all":
                        code_counts[tid][c] += 1
                    elif treatment_code == "manual_all":
                        if c in _manual_codes_set:
                            code_counts[tid][c] += 1
                    else:
                        if c == treatment_code:
                            code_counts[tid][c] += 1
        elif mode == "reserved" and a.status != "canceled":
            d = a.start_at.day
            daily[tid][d] += 1
            for c in codes:
                if c in breakdown_codes:
                    if not treatment_code or treatment_code == "all":
                        code_counts[tid][c] += 1
                    elif treatment_code == "manual_all":
                        if c in _manual_codes_set:
                            code_counts[tid][c] += 1
                    else:
                        if c == treatment_code:
                            code_counts[tid][c] += 1
    if is_doctor_filter:
        therapists = {t.id: t.name for t in db.query(models.Employee).filter(models.Employee.role == "doctor").all()}
    else:
        therapists = {t.id: t.name for t in db.query(models.Employee).all()}
    therapists["__none__"] = "미배정"
    result = []
    for tid, day_counts in daily.items():
        counts = [day_counts.get(d, 0) for d in range(1, days + 1)]
        non_zero = [c for c in counts if c > 0]
        manual_breakdown = {code: code_counts[tid].get(code, 0) for code in breakdown_codes}
        result.append({
            "therapist_id": tid,
            "therapist_name": therapists.get(tid, "?"),
            "total": sum(counts),
            "avg_per_day": round(sum(counts) / len(non_zero), 1) if non_zero else 0,
            "max_per_day": max(counts) if counts else 0,
            "min_per_day": min(non_zero) if non_zero else 0,
            "canceled": canceled.get(tid, 0),
            "daily": counts,
            "manual_breakdown": manual_breakdown,
        })
    for tid, cnt in canceled.items():
        if not any(r["therapist_id"] == tid for r in result):
            result.append({
                "therapist_id": tid, "therapist_name": therapists.get(tid, "?"),
                "total": 0, "avg_per_day": 0, "max_per_day": 0, "min_per_day": 0,
                "canceled": cnt, "daily": [0] * days,
                "manual_breakdown": {code: 0 for code in breakdown_codes},
            })
    result.sort(key=lambda x: -x["total"])
    result = [r for r in result if r["total"] > 0]
    return {"year": year, "month": month, "days": days, "items": result,
            "manual_codes": breakdown_codes, "manual_names": breakdown_names}

# ──────────────── 통계 > 도수치료 예약/완료 (치료사 × 시간항목) ────────────────

@router.get("/stats/manual-by-therapist")
def stats_manual_by_therapist(year: int, month: int, db: Session = Depends(get_db)):
    """도수치료 분석 카드 1 전용:
    - 치료사별 × 도수치료 시간항목별 → {reserved, approved}
      * reserved = 발생한 전체 예약 건수 (취소/노쇼 포함 — 운영량 기준)
      * approved = 실제 완료 건수
    - 치료사별 재진률 (완료 예약 중 is_new_patient=False 비율, %)
    치료항목이 추가되어도 DB의 manual% 항목이 자동 반영됨.
    """
    from calendar import monthrange
    from collections import defaultdict
    ts = datetime(year, month, 1)
    te = datetime(year + (1 if month == 12 else 0),
                  (1 if month == 12 else month + 1), 1)

    # 활성 도수치료 시간항목 자동 조회 (role 기반, v1.2.3+)
    manual_rows = _get_manual_treatment_rows(db)
    manual_codes = [t.code for t in manual_rows]
    manual_names = {t.code: t.name for t in manual_rows}
    manual_set = set(manual_codes)

    rows = (db.query(models.Appointment)
            .filter(models.Appointment.start_at >= ts,
                    models.Appointment.start_at < te).all())

    # tid → code → {reserved, approved}
    breakdown = defaultdict(lambda: defaultdict(lambda: {"reserved": 0, "approved": 0}))
    # tid → 재진률 계산용 (완료된 도수 예약 기준)
    revisit = defaultdict(lambda: {"approved_total": 0, "approved_revisit": 0, "approved_new": 0})

    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        manual_in_this = [c for c in codes if c in manual_set]
        if not manual_in_this:
            continue
        tid = a.therapist_id or "__none__"
        # 예약 = 전체 건수 (취소/노쇼 모두 포함), 완료 = approved 만
        for c in manual_in_this:
            breakdown[tid][c]["reserved"] += 1
            if a.status == "approved":
                breakdown[tid][c]["approved"] += 1
        # 재진률: 완료(approved) + 도수치료 포함 예약 단위 카운트
        if a.status == "approved":
            revisit[tid]["approved_total"] += 1
            if getattr(a, "is_new_patient", False):
                revisit[tid]["approved_new"] += 1
            else:
                revisit[tid]["approved_revisit"] += 1

    therapists = {t.id: t.name for t in db.query(models.Employee).all()}
    therapists["__none__"] = "미배정"

    items = []
    for tid, code_map in breakdown.items():
        rv = revisit[tid]
        total_app = rv["approved_total"]
        rate = round((rv["approved_revisit"] / total_app * 100), 1) if total_app > 0 else None
        reserved_total = sum(v["reserved"] for v in code_map.values())
        approved_total_all = sum(v["approved"] for v in code_map.values())
        items.append({
            "therapist_id": tid,
            "therapist_name": therapists.get(tid, "?"),
            "breakdown": {code: {
                "reserved": code_map[code]["reserved"],
                "approved": code_map[code]["approved"],
            } for code in manual_codes},
            "reserved_total": reserved_total,
            "approved_total": approved_total_all,
            "revisit_approved_total": total_app,
            "revisit_new": rv["approved_new"],
            "revisit_revisit": rv["approved_revisit"],
            "revisit_rate": rate,
        })
    # 예약 많은 순
    items.sort(key=lambda x: -x["reserved_total"])
    # 데이터 없는 치료사는 숨김
    items = [x for x in items if x["reserved_total"] > 0]
    return {
        "year": year, "month": month,
        "manual_codes": manual_codes,
        "manual_names": manual_names,
        "items": items,
    }


# ──────────────── 집계 탭 전용 API ────────────────

@router.get("/stats/aggregate")
def stats_aggregate(year: int, month: int, db: Session = Depends(get_db)):
    """집계 탭 전용: 치료사별 도수시간항목별 완료수 + 신환수 + 체외충격파 완료수"""
    from calendar import monthrange
    from collections import defaultdict

    ts = datetime(year, month, 1)
    te = datetime(year + (1 if month == 12 else 0),
                  (1 if month == 12 else month + 1), 1)

    # 도수치료 항목 자동 조회 (role 기반, v1.2.3+ — 새 항목 자동 반영)
    manual_rows = _get_manual_treatment_rows(db)

    eswt_row = db.query(models.Treatment).filter(
        models.Treatment.code == C.ESWT_CODE,
        models.Treatment.active == True,
    ).first()

    manual_codes = [t.code for t in manual_rows]
    manual_names = {t.code: t.name for t in manual_rows}
    eswt_code = C.ESWT_CODE

    # 완료(approved) 예약만 집계
    rows = (db.query(models.Appointment)
            .filter(models.Appointment.start_at >= ts,
                    models.Appointment.start_at < te,
                    models.Appointment.status == "approved")
            .all())

    # 치료사별 집계
    code_counts = defaultdict(lambda: defaultdict(int))  # tid → code → 건수
    new_patient_counts = defaultdict(int)                # tid → 신환수

    for a in rows:
        tid = a.therapist_id or "__none__"
        codes = _parse_codes(a.treatment_codes)
        # 체외충격파는 수동 입력(ManualCount)만 사용 — 예약 기반 카운트에서 제외 (v1.2.8+)
        for c in codes:
            if c in manual_codes:
                code_counts[tid][c] += 1
        if getattr(a, 'is_new_patient', False):
            new_patient_counts[tid] += 1

    # 체외충격파 수동 카운트 합산 (치료사별 월 합계)
    mc_start = ts.strftime("%Y-%m-%d")
    mc_end = te.strftime("%Y-%m-%d")
    mc_rows = (db.query(models.ManualCount)
               .filter(models.ManualCount.count_date >= mc_start,
                       models.ManualCount.count_date < mc_end,
                       models.ManualCount.treatment_code == eswt_code)
               .all())
    for mc in mc_rows:
        tid = mc.therapist_id or "__none__"
        code_counts[tid][eswt_code] += max(0, int(mc.count or 0))

    therapists = {t.id: t.name for t in db.query(models.Employee).filter(
        models.Employee.role == "therapist").all()}
    therapists["__none__"] = "미배정"

    # 집계에 등장한 치료사만 결과에 포함
    all_tids = set(code_counts.keys()) | set(new_patient_counts.keys())

    result = []
    for tid in all_tids:
        bd = {code: code_counts[tid].get(code, 0) for code in manual_codes}
        eswt_count = code_counts[tid].get(eswt_code, 0)
        result.append({
            "therapist_id": tid,
            "therapist_name": therapists.get(tid, "?"),
            "manual_breakdown": bd,
            "new_patient_count": new_patient_counts[tid],
            "eswt_count": eswt_count,
        })

    result.sort(key=lambda x: x["therapist_name"])

    return {
        "year": year, "month": month,
        "manual_codes": manual_codes,
        "manual_names": manual_names,
        "eswt_name": eswt_row.name if eswt_row else "체외충격파",
        "items": result,
    }

# ──────────────── 통계 > 도수치료 코드 자동 조회 ────────────────
def _get_manual_treatment_rows(db):
    """'도수치료' 의 공식 정의 (v1.2.3+):
       role='therapist' AND code != ESWT_CODE AND active=True

    ⚠ 왜 'code LIKE manual%' 가 아닌가?
       관리자 UI 에서 새 치료항목을 추가할 때, 이름이 한글이면
       _slug_code() 가 영문자 시작이 아니라 판단해 'tx_<random>' 으로
       랜덤 코드를 부여함. 그 결과 LIKE 'manual%' 패턴으로는 새 항목이
       집계에 영영 잡히지 않는 버그가 있었음.
       → role 기반 판정으로 통일: 치료사 역할이면서 체외충격파 코드가
         아닌 모든 항목을 '도수치료' 로 간주. code 가 tx_... 든 manual90 이든
         이름이 뭐든 관계없이 자동 반영.
    sort_order 순으로 정렬해서 Treatment 모델 리스트 반환."""
    return (db.query(models.Treatment)
            .filter(models.Treatment.role == "therapist",
                    models.Treatment.code != C.ESWT_CODE,
                    models.Treatment.active == True)
            .order_by(models.Treatment.sort_order).all())


def _get_manual_therapy_codes(db) -> list:
    """code 리스트만 필요한 경우용 얇은 래퍼."""
    return [t.code for t in _get_manual_treatment_rows(db)]

# ──────────────── 집계 탭 전용 API (일별×치료사별) ────────────────

@router.get("/stats/daily-by-therapist")
def stats_daily_by_therapist(year: int = None, month: int = None,
                             date_from: str = "", date_to: str = "",
                             db: Session = Depends(get_db)):
    """집계 탭: 일별 × 치료사별 도수시간항목별 완료수 + 신환수 + 체외충격파 완료수.
    v1.2.9+: date_from/date_to 지정 시 해당 기간 단위로 일별 집계."""
    from collections import defaultdict

    ts, te, range_label = _resolve_stats_range(year, month, date_from, date_to)
    date_keys = _date_list(ts, te)
    days = len(date_keys)

    # 도수치료 항목 자동 조회 (role 기반, v1.2.3+ — 새 항목 자동 반영)
    manual_rows = _get_manual_treatment_rows(db)

    eswt_row = db.query(models.Treatment).filter(
        models.Treatment.code == C.ESWT_CODE,
        models.Treatment.active == True,
    ).first()

    manual_codes = [t.code for t in manual_rows]
    manual_names = {t.code: t.name for t in manual_rows}
    eswt_code = C.ESWT_CODE

    # 활성 치료사 목록
    therapist_list = db.query(models.Employee).filter(
        models.Employee.role == "therapist",
        models.Employee.active == True,
    ).order_by(models.Employee.name).all()
    therapist_ids = [t.id for t in therapist_list]
    therapist_names = {t.id: t.name for t in therapist_list}
    therapist_colors = {t.id: (t.color or "#9CA3AF") for t in therapist_list}

    # 완료(approved) 예약만 집계
    rows = (db.query(models.Appointment)
            .filter(models.Appointment.start_at >= ts,
                    models.Appointment.start_at < te,
                    models.Appointment.status == "approved")
            .all())

    # dateKey(YYYY-MM-DD) → tid → code → count
    daily = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    daily_new = defaultdict(lambda: defaultdict(int))

    for a in rows:
        dk = a.start_at.strftime("%Y-%m-%d")
        tid = a.therapist_id or "__none__"
        codes = _parse_codes(a.treatment_codes)
        # 체외충격파는 수동 입력(ManualCount)만 사용 — 예약 기반 카운트에서 제외 (v1.2.8+)
        for c in codes:
            if c in manual_codes:
                daily[dk][tid][c] += 1
        if getattr(a, 'is_new_patient', False):
            daily_new[dk][tid] += 1

    # ─── 수동 카운트(manual_counts) 합산 (v1.2.7+) ───
    mc_start_str = ts.strftime("%Y-%m-%d")
    mc_end_str = te.strftime("%Y-%m-%d")
    mc_rows = (db.query(models.ManualCount)
               .filter(models.ManualCount.count_date >= mc_start_str,
                       models.ManualCount.count_date < mc_end_str)
               .all())
    daily_manual = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for mc in mc_rows:
        dk = mc.count_date
        tid = mc.therapist_id or "__none__"
        daily[dk][tid][mc.treatment_code] += max(0, int(mc.count or 0))
        daily_manual[dk][tid][mc.treatment_code] = max(0, int(mc.count or 0))

    # 집계에 등장했지만 therapist_list에 없는 치료사(미배정 등) 추가
    all_tids_in_data = set()
    for d_data in daily.values():
        all_tids_in_data.update(d_data.keys())
    for d_data in daily_new.values():
        all_tids_in_data.update(d_data.keys())
    for d_data in daily_manual.values():
        all_tids_in_data.update(d_data.keys())
    for tid in all_tids_in_data:
        if tid not in therapist_ids:
            therapist_ids.append(tid)
            therapist_names[tid] = "미배정"
            therapist_colors[tid] = "#9CA3AF"

    # 결과 조립 — 기간 내 모든 날짜에 대해 한 행씩
    items = []
    for dk in date_keys:
        therapist_data = {}
        for tid in therapist_ids:
            bd = {code: daily[dk][tid].get(code, 0) for code in manual_codes}
            therapist_data[tid] = {
                "manual_breakdown": bd,
                "eswt_count": daily[dk][tid].get(eswt_code, 0),
                "new_patient_count": daily_new[dk].get(tid, 0),
                "eswt_manual_input": daily_manual[dk][tid].get(eswt_code, 0),
            }
        items.append({
            "date": dk,
            "day": int(dk[8:10]),  # 구 필드 호환
            "therapist_data": therapist_data,
        })

    return {
        # 구 필드 (호환용)
        "year": ts.year, "month": ts.month,
        # 신 필드 (v1.2.9+)
        "date_from": ts.strftime("%Y-%m-%d"),
        "date_to": (te - timedelta(days=1)).strftime("%Y-%m-%d"),
        "range_label": range_label,
        "days": days,
        "therapists": [
            {"id": tid, "name": therapist_names[tid], "color": therapist_colors[tid]}
            for tid in therapist_ids
        ],
        "manual_codes": manual_codes,
        "manual_names": manual_names,
        "eswt_name": eswt_row.name if eswt_row else "체외충격파",
        "eswt_code": eswt_code,
        "items": items,
    }

# ──────────────── 수동 카운트 (v1.2.7+) ────────────────
# 체외충격파 등 당일 내방 환자 · 예약 등록 없이 바로 진행한 케이스를
# 집계/통계에 반영하기 위한 "수동 입력 카운트".
# (count_date, therapist_id, treatment_code) 당 1행 · count 덮어쓰기.

@router.post("/manual-counts")
def upsert_manual_count(payload: dict = Body(...),
                        db: Session = Depends(get_db)):
    """수동 카운트 저장 (UPSERT).
    payload: {
      "date": "YYYY-MM-DD",           # 필수
      "therapist_id": "...",          # null 또는 빈값 허용 (미배정)
      "treatment_code": "eswt",       # 필수
      "count": 3                      # 필수 · 0 이면 레코드 삭제
    }
    """
    date = (payload.get("date") or "").strip()
    tcode = (payload.get("treatment_code") or "").strip()
    tid = payload.get("therapist_id") or None
    if isinstance(tid, str) and tid.strip() in ("", "__none__"):
        tid = None
    try:
        count = int(payload.get("count") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "count 는 정수여야 합니다.")

    if not date or len(date) != 10 or date[4] != "-" or date[7] != "-":
        raise HTTPException(400, "date 는 'YYYY-MM-DD' 형식이어야 합니다.")
    if not tcode:
        raise HTTPException(400, "treatment_code 가 필요합니다.")
    if count < 0:
        raise HTTPException(400, "count 는 0 이상이어야 합니다.")

    # 기존 레코드 찾기
    row = (db.query(models.ManualCount)
           .filter(models.ManualCount.count_date == date,
                   models.ManualCount.therapist_id == tid,
                   models.ManualCount.treatment_code == tcode)
           .first())

    if count == 0:
        # 0 이면 삭제 (깨끗하게 유지)
        if row:
            db.delete(row)
            db.commit()
        return {"ok": True, "deleted": True, "count": 0}

    if row:
        row.count = count
    else:
        row = models.ManualCount(
            count_date=date,
            therapist_id=tid,
            treatment_code=tcode,
            count=count,
        )
        db.add(row)

    audit(db, "manual_count.upsert", "",
          f"date={date} tid={tid} code={tcode} count={count}")
    db.commit()
    return {"ok": True, "count": count}


# ──────────────── 통계/집계 공용 기간 해석기 (v1.2.9+) ────────────────

def _resolve_stats_range(year, month, date_from, date_to):
    """통계·집계 엔드포인트 공용 — 입력을 (ts, te, label) 로 변환.
    우선순위: date_from/date_to > year/month > 현재 월.
    ts = 기간 시작 datetime (00:00), te = 기간 종료 다음날 datetime (00:00).
    label = 'YYYY-MM-DD~YYYY-MM-DD' 또는 'YYYY-MM'.
    """
    if date_from and date_to:
        try:
            ts = datetime.strptime(date_from, "%Y-%m-%d")
            te_inc = datetime.strptime(date_to, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "date_from / date_to 는 'YYYY-MM-DD' 형식이어야 합니다.")
        if te_inc < ts:
            raise HTTPException(400, "date_to 는 date_from 이후여야 합니다.")
        return ts, te_inc + timedelta(days=1), f"{date_from}~{date_to}"
    if year and month:
        ts = datetime(year, month, 1)
        te = datetime(year + (1 if month == 12 else 0),
                      (1 if month == 12 else month + 1), 1)
        return ts, te, f"{year:04d}-{month:02d}"
    now = datetime.now()
    ts = datetime(now.year, now.month, 1)
    te = datetime(now.year + (1 if now.month == 12 else 0),
                  (1 if now.month == 12 else now.month + 1), 1)
    return ts, te, f"{now.year:04d}-{now.month:02d}"


def _date_list(ts, te):
    """[ts, te) 기간의 'YYYY-MM-DD' 문자열 리스트."""
    out = []
    cur = ts
    while cur < te:
        out.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return out


# ──────────────── 통계 > 상단 카드용 요약 ────────────────

@router.get("/stats/summary")
def stats_summary(year: int = None, month: int = None,
                  date_from: str = "", date_to: str = "",
                  treatment_code: str = "", db: Session = Depends(get_db)):
    """상단 카드용: 총예약 / 총완료 / 도수예약 / 도수완료 / 취소
    v1.2.9+: date_from/date_to 지정 시 해당 기간 (없으면 year/month 월 범위)."""
    from calendar import monthrange

    ts, te, range_label = _resolve_stats_range(year, month, date_from, date_to)
    days = (te - ts).days

    rows = (
        db.query(models.Appointment)
        .filter(
            models.Appointment.start_at >= ts,
            models.Appointment.start_at < te,
        )
        .all()
    )

    _manual_codes = _get_manual_therapy_codes(db)
    _manual_codes_set = set(_manual_codes)

    def _matches(codes: list) -> bool:
        if not treatment_code or treatment_code == "all":
            return True
        if treatment_code == "manual_all":
            return any(c in _manual_codes_set for c in codes)
        return treatment_code in codes

    total = 0
    manual = 0
    approved = 0
    manual_approved = 0
    canceled = 0
    # 20-3-1 (post-19-P / F-10): 노쇼 별도 카운트 (cancel 의 부분집합)
    no_show_count = 0

    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        if not _matches(codes):
            continue
        total += 1
        if a.status == "canceled":
            canceled += 1
            if getattr(a, "no_show", False):
                no_show_count += 1
        else:
            is_manual = any(c in _manual_codes_set for c in codes)
            if is_manual:
                manual += 1
            if a.status == "approved":
                approved += 1
                if is_manual:
                    manual_approved += 1

    # ⚠ v1.2.8+ 변경: 체외충격파 수동 카운트는 '집계' 탭에만 반영.
    # 통계 탭(이달 요약 / 날짜별 표 / 매출 차트)은 순수 예약 기반으로 복원.

    return {
        # 구 필드 (호환용)
        "year": ts.year,
        "month": ts.month,
        # 신 필드 (v1.2.9+)
        "date_from": ts.strftime("%Y-%m-%d"),
        "date_to": (te - timedelta(days=1)).strftime("%Y-%m-%d"),
        "range_label": range_label,
        "days": days,
        "total": total,
        "manual": manual,
        "approved": approved,
        "manual_approved": manual_approved,
        "canceled": canceled,
        # 20-3-1 (post-19-P / F-10): 노쇼 별도 카운트
        "no_show_count": no_show_count,
        "treatment_code": treatment_code,
    }

# ──────────────── 통계 > 시간대별 예약 수 ────────────────

@router.get("/stats/by-hour")
def stats_by_hour(year: int, month: int, mode: str = "reserved", treatment_code: str = "", db: Session = Depends(get_db)):
    """중간 분석용: 시간대(0~23시)별 건수 / mode=reserved: 예약기준 / mode=approved: 완료기준"""
    from collections import defaultdict

    ts = datetime(year, month, 1)
    te = datetime(
        year + (1 if month == 12 else 0),
        (1 if month == 12 else month + 1),
        1,
    )

    rows = (
        db.query(models.Appointment)
        .filter(
            models.Appointment.start_at >= ts,
            models.Appointment.start_at < te,
        )
        .all()
    )

    _manual_codes_set = set(_get_manual_therapy_codes(db))

    def _matches(codes: list) -> bool:
        if not treatment_code or treatment_code == "all":
            return True
        if treatment_code == "manual_all":
            return any(c in _manual_codes_set for c in codes)
        return treatment_code in codes

    counts = defaultdict(int)
    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        if not _matches(codes):
            continue
        # mode=all: 전체(취소/노쇼 포함) — 운영량 기준
        if mode == "all":
            counts[a.start_at.hour] += 1
        elif mode == "approved" and a.status == "approved":
            counts[a.start_at.hour] += 1
        elif mode == "reserved" and a.status != "canceled":
            counts[a.start_at.hour] += 1

    items = [{"hour": h, "label": f"{h:02d}시", "count": counts[h]} for h in range(24)]

    return {"year": year, "month": month, "items": items}

# ──────────────── 통계 > 요일별 예약 수 ────────────────

@router.get("/stats/by-weekday")
def stats_by_weekday(year: int, month: int, mode: str = "reserved", treatment_code: str = "", db: Session = Depends(get_db)):
    """중간 분석용: 요일별 건수 / mode=reserved: 예약기준 / mode=approved: 완료기준"""
    from collections import defaultdict

    ts = datetime(year, month, 1)
    te = datetime(
        year + (1 if month == 12 else 0),
        (1 if month == 12 else month + 1),
        1,
    )

    rows = (
        db.query(models.Appointment)
        .filter(
            models.Appointment.start_at >= ts,
            models.Appointment.start_at < te,
        )
        .all()
    )

    _manual_codes_set = set(_get_manual_therapy_codes(db))

    def _matches(codes: list) -> bool:
        if not treatment_code or treatment_code == "all":
            return True
        if treatment_code == "manual_all":
            return any(c in _manual_codes_set for c in codes)
        return treatment_code in codes

    counts = defaultdict(int)
    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        if not _matches(codes):
            continue
        # mode=all: 전체(취소/노쇼 포함)
        if mode == "all":
            counts[a.start_at.weekday()] += 1
        elif mode == "approved" and a.status == "approved":
            counts[a.start_at.weekday()] += 1
        elif mode == "reserved" and a.status != "canceled":
            counts[a.start_at.weekday()] += 1

    weekday_names = ["월", "화", "수", "목", "금", "토", "일"]
    items = [
        {"weekday": i, "label": weekday_names[i], "count": counts[i]}
        for i in range(7)
    ]

    return {"year": year, "month": month, "items": items}

# ──────────────── 통계 > 치료종류별 예약 수 ────────────────

@router.get("/stats/by-treatment")
def stats_by_treatment(year: int, month: int, mode: str = "reserved", treatment_code: str = "", db: Session = Depends(get_db)):
    """중간 분석용: 치료항목 코드별 건수 / mode=reserved: 예약기준 / mode=approved: 완료기준"""
    from collections import defaultdict

    ts = datetime(year, month, 1)
    te = datetime(
        year + (1 if month == 12 else 0),
        (1 if month == 12 else month + 1),
        1,
    )

    rows = (
        db.query(models.Appointment)
        .filter(
            models.Appointment.start_at >= ts,
            models.Appointment.start_at < te,
        )
        .all()
    )

    _manual_codes_set = set(_get_manual_therapy_codes(db))
    tx_map = {t.code: t.name for t in db.query(models.Treatment).all()}

    def _matches(codes: list) -> bool:
        if not treatment_code or treatment_code == "all":
            return True
        if treatment_code == "manual_all":
            return any(c in _manual_codes_set for c in codes)
        return treatment_code in codes

    counts = defaultdict(int)
    for a in rows:
        if mode == "approved" and a.status != "approved":
            continue
        if mode == "reserved" and a.status == "canceled":
            continue
        codes = _parse_codes(a.treatment_codes)
        if not _matches(codes):
            continue
        for c in codes:
            counts[c] += 1

    items = sorted(
        [
            {"code": code, "label": tx_map.get(code, code), "count": cnt}
            for code, cnt in counts.items()
        ],
        key=lambda x: -x["count"],
    )

    return {"year": year, "month": month, "items": items}

# ──────────────── 통계 > 날짜별 요약 (하단 표용) ────────────────

@router.get("/stats/daily")
def stats_daily(year: int = None, month: int = None,
                date_from: str = "", date_to: str = "",
                treatment_code: str = "", db: Session = Depends(get_db)):
    """하단 표용: 날짜별 총예약 / 총완료 / 도수항목별 예약 / 취소
    v1.2.9+: date_from/date_to 지정 시 해당 기간 단위로 일별 집계."""
    ts, te, range_label = _resolve_stats_range(year, month, date_from, date_to)
    date_keys = _date_list(ts, te)  # ['YYYY-MM-DD', ...]
    days = len(date_keys)

    rows = (
        db.query(models.Appointment)
        .filter(
            models.Appointment.start_at >= ts,
            models.Appointment.start_at < te,
        )
        .all()
    )

    _manual_codes = _get_manual_therapy_codes(db)
    _manual_codes_set = set(_manual_codes)

    def _matches(codes: list) -> bool:
        if not treatment_code or treatment_code == "all":
            return True
        if treatment_code == "manual_all":
            return any(c in _manual_codes_set for c in codes)
        return treatment_code in codes

    manual_rows = _get_manual_treatment_rows(db)
    manual_codes = [t.code for t in manual_rows]
    manual_names = {t.code: t.name for t in manual_rows}

    daily = {
        dk: {
            "total": 0, "approved": 0, "manual": 0, "manual_approved": 0, "eswt": 0, "canceled": 0,
            "manual_by_code": {code: 0 for code in manual_codes},
            "manual_approved_by_code": {code: 0 for code in manual_codes},
        }
        for dk in date_keys
    }

    for a in rows:
        codes = _parse_codes(a.treatment_codes)
        if not _matches(codes):
            continue
        dk = a.start_at.strftime("%Y-%m-%d")
        if dk not in daily:
            continue  # 안전 가드 (이론상 필터 이미 걸림)
        daily[dk]["total"] += 1
        if a.status == "canceled":
            daily[dk]["canceled"] += 1
        else:
            is_manual = any(c in _manual_codes_set for c in codes)
            is_eswt = "eswt" in codes
            if is_manual:
                daily[dk]["manual"] += 1
                for c in codes:
                    if c in manual_codes:
                        daily[dk]["manual_by_code"][c] += 1
            if is_eswt:
                daily[dk]["eswt"] += 1
            if a.status == "approved":
                daily[dk]["approved"] += 1
                if is_manual:
                    daily[dk]["manual_approved"] += 1
                    for c in codes:
                        if c in manual_codes:
                            daily[dk]["manual_approved_by_code"][c] += 1

    # ⚠ v1.2.8+ 변경: 체외충격파 수동 카운트는 '집계' 탭에만 반영.

    items = [
        {
            "date": dk,
            "day": int(dk[8:10]),   # 구 필드 호환 (월 내에서만 의미 있음)
            "total": daily[dk]["total"],
            "approved": daily[dk]["approved"],
            "manual": daily[dk]["manual"],
            "manual_approved": daily[dk]["manual_approved"],
            "eswt": daily[dk]["eswt"],
            "canceled": daily[dk]["canceled"],
            "manual_by_code": daily[dk]["manual_by_code"],
            "manual_approved_by_code": daily[dk]["manual_approved_by_code"],
        }
        for dk in date_keys
    ]

    return {
        "year": ts.year, "month": ts.month,
        "date_from": ts.strftime("%Y-%m-%d"),
        "date_to": (te - timedelta(days=1)).strftime("%Y-%m-%d"),
        "range_label": range_label,
        "days": days, "items": items,
        "manual_codes": manual_codes, "manual_names": manual_names,
        "treatment_code": treatment_code,
    }


# ──────────────── 도수치료 예약현황 엑셀 다운로드 ────────────────
# 통계가 아닌 "현장에서 보고/적기 위한 인쇄용" 단순 출력. A4 가로 1페이지.

def _lighten_hex(hex_color: str, factor: float) -> str:
    """#RRGGBB -> 흰색 쪽으로 factor (0~1) 만큼 블렌드한 RRGGBB (# 없이).
    factor=0 → 원색, factor=1 → 흰색."""
    h = (hex_color or "").lstrip("#")
    if len(h) != 6:
        return "FFFFFF"
    try:
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except Exception:
        return "FFFFFF"
    f = max(0.0, min(1.0, float(factor)))
    r = int(r + (255 - r) * f)
    g = int(g + (255 - g) * f)
    b = int(b + (255 - b) * f)
    return f"{r:02X}{g:02X}{b:02X}"


@router.get("/export/manual-schedule.xlsx")
def export_manual_schedule(date: str, db: Session = Depends(get_db)):
    """현장 인쇄용 도수치료 예약현황 엑셀 (A4 가로, 가로 1페이지 맞춤).
    - 도수치료 예약만 (role=therapist & code != ESWT_CODE)
    - 컬럼: 활성 치료사(can_manual=True) + 미배정
    - 30분 단위 그리드, 60분 예약은 세로 2칸 병합 (default_minutes / 30)
    - 같은 예약에 ESWT='충', doctor 코드='주' suffix
    - canceled 예약 제외
    - 운영시간 기본, 도수 예약이 운영 종료보다 늦게 끝나면 자동 확장
    """
    try:
        target = datetime.strptime(date, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(400, "date 형식: YYYY-MM-DD")

    cfg = load_config()

    def _hm_to_min(s: str) -> int:
        h, m = (s or "00:00").split(":")
        return int(h) * 60 + int(m)
    def _min_to_hm(m: int) -> str:
        return f"{m // 60:02d}:{m % 60:02d}"

    open_min = _hm_to_min(cfg.get("open_time", "08:30"))
    close_min = _hm_to_min(cfg.get("close_time", "18:30"))

    manual_rows = _get_manual_treatment_rows(db)
    manual_meta = {t.code: t for t in manual_rows}
    eswt_code = C.ESWT_CODE
    doctor_codes = _doctor_codes_set(db)

    therapists = (
        db.query(models.Employee)
        .filter(models.Employee.role == "therapist",
                models.Employee.active == True,
                models.Employee.can_manual == True)
        .order_by(models.Employee.sort_order, models.Employee.name)
        .all()
    )

    start_dt = datetime.combine(target, datetime.min.time())
    end_dt = datetime.combine(target, datetime.max.time())
    appts = (
        db.query(models.Appointment)
        .filter(models.Appointment.start_at >= start_dt,
                models.Appointment.start_at <= end_dt,
                models.Appointment.status != "canceled")
        .all()
    )

    # 그리드 끝 시각: 도수예약이 운영시간보다 늦게 끝나면 자동 확장.
    # ⚠ ap.duration_min 은 충격파/주사가 같이 묶이면 그만큼 늘어나 있을 수 있어
    #    그대로 쓰면 충/주 때문에 엑셀 행이 불필요하게 늘어남. 셀 병합 로직과
    #    동일하게 도수치료 항목의 default_minutes 만 사용.
    for ap in appts:
        try:
            codes = json.loads(ap.treatment_codes or "[]")
        except Exception:
            codes = []
        manual_in = [c for c in codes if c in manual_meta]
        if not manual_in:
            continue
        m_code = sorted(manual_in, key=lambda c: manual_meta[c].sort_order or 0)[0]
        m_dur = manual_meta[m_code].default_minutes or 30
        end_m = ap.start_at.hour * 60 + ap.start_at.minute + m_dur
        if end_m > close_min:
            close_min = end_m

    # 시간 행: 운영 시작(또는 그 이전 30분 정렬값)부터 운영 종료 직전 슬롯까지.
    # 운영 종료 시각 자체는 "예약 시작 슬롯" 이 아니므로 행에 포함하지 않는다.
    #   예) 09:00~18:00 → 09:00, 09:30, ..., 17:30 (18:00 행 없음)
    # 단, 도수예약이 운영 종료 이후까지 이어지면 위 close_min 확장으로 자동 연장됨.
    SLOT = 30
    grid_start = (open_min // SLOT) * SLOT
    # range stop 값으로 close_min 을 그대로 쓰면 자동으로 close_min 직전 슬롯이 마지막.
    # close_min 이 SLOT 배수가 아니면 ceil 해서 그 직전 슬롯까지 표시.
    grid_end = ((close_min + SLOT - 1) // SLOT) * SLOT
    times = list(range(grid_start, grid_end, SLOT))  # 슬롯 시작분 배열
    if not times:
        times = [grid_start]

    cols = list(therapists) + [None]  # 마지막은 미배정

    # ─ 셀 배치
    cell_map = {}  # (row_idx, col_idx) -> (text, span)
    UNASSIGNED_HEX = "#8B5CF6"  # 보라 (UI 미배정 컬럼과 동일 톤)
    pid_cache = {}

    def _patient_name(pid: str) -> str:
        if pid in pid_cache:
            return pid_cache[pid]
        pt = db.get(models.Patient, pid)
        nm = (pt.name if pt else "?") or "?"
        pid_cache[pid] = nm
        return nm

    for ap in appts:
        try:
            codes = json.loads(ap.treatment_codes or "[]")
        except Exception:
            codes = []
        manual_in = [c for c in codes if c in manual_meta]
        if not manual_in:
            continue
        # 같은 예약에 도수 1개 전제. 여러 개면 sort_order 우선 항목 사용.
        m_code = sorted(manual_in, key=lambda c: manual_meta[c].sort_order or 0)[0]
        m_t = manual_meta[m_code]
        m_dur = m_t.default_minutes or SLOT

        s_min = ap.start_at.hour * 60 + ap.start_at.minute
        snapped = (s_min // SLOT) * SLOT
        if snapped not in times:
            continue  # 그리드 밖
        row_idx = times.index(snapped)
        span = max(1, (m_dur + SLOT - 1) // SLOT)
        if row_idx + span > len(times):
            span = len(times) - row_idx

        # 컬럼 매칭 (담당 치료사 / 미배정)
        col_idx = len(cols) - 1  # 기본: 미배정
        if ap.therapist_id:
            for i, t in enumerate(cols):
                if t is not None and t.id == ap.therapist_id:
                    col_idx = i
                    break

        # 텍스트
        m_short = m_t.short or m_t.name or m_code
        text = f"{_patient_name(ap.patient_id)} {m_short}"
        suffix = []
        if eswt_code in codes:
            suffix.append("충")
        if any(c in doctor_codes for c in codes):
            suffix.append("주")
        if suffix:
            text += " " + ",".join(suffix)

        cell_map[(row_idx, col_idx)] = (text, span)

    # ─ openpyxl 워크북 생성
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(target)

    thin = Side(style="thin", color="888888")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    font = Font(name="맑은 고딕", size=11)
    align = Alignment(horizontal="center", vertical="center",
                      shrink_to_fit=True, wrap_text=False)

    # 헤더 행
    head_time = ws.cell(row=1, column=1, value="시간")
    head_time.font = font; head_time.alignment = align; head_time.border = border
    head_time.fill = PatternFill("solid", fgColor="EEEEEE")

    for ci, t in enumerate(cols):
        cell = ws.cell(row=1, column=2 + ci)
        if t is None:
            cell.value = "미배정"
            color = _lighten_hex(UNASSIGNED_HEX, 0.70)
        else:
            cell.value = t.name
            color = _lighten_hex(t.color or "#9CA3AF", 0.70)
        cell.font = font; cell.alignment = align; cell.border = border
        cell.fill = PatternFill("solid", fgColor=color)

    # 데이터 행 + 병합
    skip = set()
    for ri, m in enumerate(times):
        excel_row = ri + 2
        tc = ws.cell(row=excel_row, column=1, value=_min_to_hm(m))
        tc.font = font; tc.alignment = align; tc.border = border
        tc.fill = PatternFill("solid", fgColor="F8F9FA")

        for ci, t in enumerate(cols):
            if (ri, ci) in skip:
                continue
            cell = ws.cell(row=excel_row, column=2 + ci)
            cell.font = font; cell.alignment = align; cell.border = border

            entry = cell_map.get((ri, ci))
            if not entry:
                continue
            text, span = entry
            cell.value = text
            base = UNASSIGNED_HEX if t is None else (t.color or "#9CA3AF")
            light = _lighten_hex(base, 0.85)
            cell.fill = PatternFill("solid", fgColor=light)

            if span > 1:
                end_row = excel_row + span - 1
                ws.merge_cells(start_row=excel_row, start_column=2 + ci,
                               end_row=end_row, end_column=2 + ci)
                for dr in range(1, span):
                    skip.add((ri + dr, ci))
                    sub = ws.cell(row=excel_row + dr, column=2 + ci)
                    sub.font = font; sub.alignment = align; sub.border = border
                    sub.fill = PatternFill("solid", fgColor=light)

    # 행 높이 0.6cm ≈ 17pt 고정
    ws.row_dimensions[1].height = 22  # 헤더만 약간 크게
    for r in range(2, 2 + len(times)):
        ws.row_dimensions[r].height = 17

    # 열 너비
    ws.column_dimensions["A"].width = 8
    for i in range(len(cols)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16

    # 페이지 설정 — A4 가로, 가로 1페이지 맞춤
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 세로는 자동 (여러 페이지 허용)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    fname = f"도수치료_예약현황_{target}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}",
        },
    )


# ──────────────── 통계 탭 엑셀 다운로드 ────────────────
# 보고서용. 3개 시트: 요약 / 일별통계 / 치료사별집계.
# 통계 탭(예약 기반) + 집계 탭(완료 기반 + ManualCount 합산) 패턴을 시트별로 사용.

@router.get("/export/stats.xlsx")
def export_stats_xlsx(date_from: str, date_to: str, db: Session = Depends(get_db)):
    """통계 탭 보고서 엑셀.
    - 요약 시트: 기간 KPI + 날짜별 도수 + 치료사별 도수 + 비교표 2개
    - 일별통계 시트: 기간 전체 일별 표
    - 치료사별집계 시트: 치료사별 완료/신환/재진률/예상매출 (미배정 제외, 합계 행)
    """
    from collections import defaultdict
    from datetime import datetime as _dt
    import io as _io
    from urllib.parse import quote
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    # ─── 날짜 검증
    try:
        d_from = _dt.strptime(date_from, "%Y-%m-%d").date()
        d_to = _dt.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(400, "date_from / date_to 형식: YYYY-MM-DD")
    if d_to < d_from:
        raise HTTPException(400, "date_to 가 date_from 보다 이전입니다.")

    ts = _dt.combine(d_from, _dt.min.time())
    te = _dt.combine(d_to, _dt.min.time()) + timedelta(days=1)
    date_keys = _date_list(ts, te)

    # ─── 치료항목 메타
    manual_rows = _get_manual_treatment_rows(db)
    manual_codes = [t.code for t in manual_rows]
    manual_names = {t.code: t.name for t in manual_rows}
    manual_prices = {t.code: int(t.price or 0) for t in manual_rows}
    eswt_code = C.ESWT_CODE
    eswt_row = (db.query(models.Treatment)
                .filter(models.Treatment.code == eswt_code,
                        models.Treatment.active == True).first())
    eswt_name = eswt_row.name if eswt_row else "체외충격파"
    eswt_price = int(eswt_row.price or 0) if eswt_row else 0
    manual_codes_set = set(manual_codes)

    # ─── 활성 치료사
    therapists = (db.query(models.Employee)
                  .filter(models.Employee.role == "therapist")
                  .order_by(models.Employee.sort_order, models.Employee.name).all())
    ther_by_id = {t.id: t for t in therapists}

    # ─── 예약 로드 (기간 내 전체)
    appts = (db.query(models.Appointment)
             .filter(models.Appointment.start_at >= ts,
                     models.Appointment.start_at < te).all())

    # ─── KPI (통계탭 stats_summary 와 동일 — 예약 기반)
    kpi_total = 0
    kpi_approved = 0
    kpi_manual = 0
    kpi_manual_approved = 0
    kpi_canceled = 0

    # ─── 일별 (통계탭 stats_daily 와 동일 — 예약 기반)
    daily = {dk: {
        "manual": 0, "manual_approved": 0,
        "by_code": {c: 0 for c in manual_codes},
        "eswt": 0, "canceled": 0, "total": 0, "approved": 0,
    } for dk in date_keys}

    # ─── 치료사별 (집계탭 stats_aggregate 와 동일 — 완료 기반)
    th_code_counts = defaultdict(lambda: defaultdict(int))   # tid → code → 완료건수
    th_new_patient = defaultdict(int)                         # tid → 신환수
    th_returning = defaultdict(int)                           # tid → 재진수 (도수 완료 중 is_new_patient=False)
    th_manual_done = defaultdict(int)                         # tid → 도수 완료 총수 (재진률 분모)

    for a in appts:
        codes = _parse_codes(a.treatment_codes)
        kpi_total += 1
        dk = a.start_at.strftime("%Y-%m-%d")
        in_range = dk in daily
        if in_range:
            daily[dk]["total"] += 1

        if a.status == "canceled":
            kpi_canceled += 1
            if in_range:
                daily[dk]["canceled"] += 1
            continue

        is_manual = any(c in manual_codes_set for c in codes)
        is_eswt = eswt_code in codes
        if is_manual:
            kpi_manual += 1
            if in_range:
                daily[dk]["manual"] += 1
                for c in codes:
                    if c in manual_codes_set:
                        daily[dk]["by_code"][c] += 1
        if in_range and is_eswt:
            daily[dk]["eswt"] += 1

        if a.status == "approved":
            kpi_approved += 1
            if in_range:
                daily[dk]["approved"] += 1
            if is_manual:
                kpi_manual_approved += 1
                if in_range:
                    daily[dk]["manual_approved"] += 1

                # 치료사별 집계 (완료 기반, 도수 코드별)
                tid = a.therapist_id  # 미배정 제외 위해 None 그대로
                if tid:
                    for c in codes:
                        if c in manual_codes_set:
                            th_code_counts[tid][c] += 1
                    th_manual_done[tid] += 1
                    if a.is_new_patient:
                        th_new_patient[tid] += 1
                    else:
                        th_returning[tid] += 1

    # 체외충격파 ManualCount 합산 (집계탭 정책 — 치료사별 시트에 반영)
    mc_rows = (db.query(models.ManualCount)
               .filter(models.ManualCount.count_date >= d_from.strftime("%Y-%m-%d"),
                       models.ManualCount.count_date <= d_to.strftime("%Y-%m-%d"),
                       models.ManualCount.treatment_code == eswt_code).all())
    th_eswt = defaultdict(int)
    for mc in mc_rows:
        if mc.therapist_id:  # 미배정은 표에서 제외
            th_eswt[mc.therapist_id] += max(0, int(mc.count or 0))

    # 치료사별 결과 정리 (실제 치료사 + 활동 있는 치료사만, 미배정 제외)
    active_tids = (set(th_code_counts.keys()) | set(th_new_patient.keys())
                   | set(th_eswt.keys()))
    th_rows_data = []
    for t in therapists:
        if t.id not in active_tids:
            continue
        bd = {c: th_code_counts[t.id].get(c, 0) for c in manual_codes}
        total_manual = sum(bd.values())
        new_p = th_new_patient[t.id]
        ret_p = th_returning[t.id]
        denom = th_manual_done[t.id]
        rate = (ret_p / denom * 100.0) if denom > 0 else 0.0
        # 매출 = 도수항목 합계 + 체외충격파 수동 입력분 (통계탭 화면과 동일 정책 — 수가 미입력 시 0 가산)
        revenue = (sum(bd[c] * manual_prices.get(c, 0) for c in manual_codes)
                   + th_eswt[t.id] * eswt_price)
        th_rows_data.append({
            "id": t.id, "name": t.name, "color": t.color or "#9CA3AF",
            "by_code": bd, "total_manual": total_manual,
            "new_patient": new_p, "rate": rate, "revenue": revenue,
            "eswt": th_eswt[t.id],
        })
    # 기본 정렬: 매출 내림차순 (보고서답게)
    th_rows_data.sort(key=lambda x: (-x["revenue"], x["name"]))

    # ─── 스타일 헬퍼 ─────────────────────────────────────
    THIN = Side(style="thin", color="CBD5E1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NAVY_FILL = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
    TOTAL_FILL = PatternFill("solid", fgColor="E2E8F0")
    KPI_FILL = PatternFill("solid", fgColor="F1F5F9")

    F_TITLE = Font(name="맑은 고딕", size=16, bold=True, color="1E293B")
    F_SUBTITLE = Font(name="맑은 고딕", size=10, color="64748B")
    F_SECTION = Font(name="맑은 고딕", size=12, bold=True, color="1E293B")
    F_DESC = Font(name="맑은 고딕", size=9, italic=True, color="64748B")
    F_HEADER = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    F_BODY = Font(name="맑은 고딕", size=10, color="1E293B")
    F_TOTAL = Font(name="맑은 고딕", size=10, bold=True, color="0F172A")
    F_KPI_LABEL = Font(name="맑은 고딕", size=9, color="64748B")
    F_KPI_VALUE = Font(name="맑은 고딕", size=14, bold=True, color="1E293B")

    A_CENTER = Alignment(horizontal="center", vertical="center", shrink_to_fit=False)
    A_RIGHT = Alignment(horizontal="right", vertical="center")
    A_LEFT = Alignment(horizontal="left", vertical="center")

    def _setup_landscape(ws):
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
        ws.print_options.horizontalCentered = True

    def _draw_header_row(ws, row, headers, col_start=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.font = F_HEADER; c.alignment = A_CENTER; c.border = BORDER
            c.fill = NAVY_FILL

    def _style_body_cell(cell, alt=False, bold=False, total=False):
        cell.font = F_TOTAL if (bold or total) else F_BODY
        cell.border = BORDER
        if total:
            cell.fill = TOTAL_FILL
        elif alt:
            cell.fill = ALT_FILL

    def _lighten_hex(hex_color: str, factor: float) -> str:
        return _lighten_hex_inner(hex_color, factor)

    def _money(cell, val):
        cell.value = int(val or 0)
        cell.number_format = '"₩"#,##0'
        cell.alignment = A_RIGHT

    # ─── 워크북 ─────────────────────────────────────
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "요약"
    ws_daily = wb.create_sheet("일별통계")
    ws_th = wb.create_sheet("치료사별집계")

    now_str = _dt.now().strftime("%Y-%m-%d %H:%M")
    period_str = f"{d_from.isoformat()} ~ {d_to.isoformat()}"

    # ────────────────── 요약 시트 ──────────────────
    ws = ws_sum
    _setup_landscape(ws)

    # 1행: 제목
    ws.cell(row=1, column=1, value="도수치료 통계").font = F_TITLE
    ws.cell(row=1, column=1).alignment = A_LEFT
    # 2행: 부제 + 출력일시
    ws.cell(row=2, column=1, value=period_str).font = F_SUBTITLE
    ws.cell(row=2, column=1).alignment = A_LEFT
    # 출력일시는 오른쪽 (G2 정도)
    out_cell = ws.cell(row=2, column=8, value=f"출력일시 {now_str}")
    out_cell.font = F_SUBTITLE; out_cell.alignment = A_RIGHT
    # 3행: 얇은 남색 구분선
    for col in range(1, 13):
        b = ws.cell(row=3, column=col)
        b.border = Border(bottom=Side(style="thin", color="1E3A5F"))

    # ─── 기간 요약 (KPI)
    row = 5
    ws.cell(row=row, column=1, value="기간 요약").font = F_SECTION
    row += 1
    ws.cell(row=row, column=1,
            value="조회 기간 동안의 도수치료 예약/완료 현황입니다.").font = F_DESC
    row += 2

    kpis = [
        ("총 예약", kpi_total),
        ("총 완료", kpi_approved),
        ("도수 예약", kpi_manual),
        ("도수 완료", kpi_manual_approved),
        ("취소", kpi_canceled),
    ]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 2
        # label
        cl = ws.cell(row=row, column=col, value=label)
        cl.font = F_KPI_LABEL; cl.alignment = A_CENTER; cl.fill = KPI_FILL; cl.border = BORDER
        # value
        cv = ws.cell(row=row + 1, column=col, value=value)
        cv.font = F_KPI_VALUE; cv.alignment = A_CENTER; cv.fill = KPI_FILL; cv.border = BORDER
        # 값 셀 옆 빈 칸도 KPI 카드 배경 (가독성)
        cl2 = ws.cell(row=row, column=col + 1)
        cl2.fill = KPI_FILL; cl2.border = BORDER
        cv2 = ws.cell(row=row + 1, column=col + 1)
        cv2.fill = KPI_FILL; cv2.border = BORDER
    row += 3

    # ─── 날짜별 도수치료 현황
    row += 1
    ws.cell(row=row, column=1, value="날짜별 도수치료 현황").font = F_SECTION
    row += 1
    ws.cell(row=row, column=1,
            value="일자별 도수치료 예약, 완료, 치료항목별 건수를 확인할 수 있습니다.").font = F_DESC
    row += 2

    daily_headers = ["날짜", "요일", "도수 예약", "도수 완료"] + \
                    [manual_names[c] for c in manual_codes] + [eswt_name]
    _draw_header_row(ws, row, daily_headers)
    row += 1

    weekday_kr = "월화수목금토일"
    daily_table_start = row
    sum_manual = sum_manual_approved = sum_eswt = 0
    sum_by_code = {c: 0 for c in manual_codes}
    for idx, dk in enumerate(date_keys):
        d = daily[dk]
        # 요일
        dt_obj = _dt.strptime(dk, "%Y-%m-%d")
        wd = weekday_kr[dt_obj.weekday()]
        cells = [dk, wd, d["manual"], d["manual_approved"]] + \
                [d["by_code"][c] for c in manual_codes] + [d["eswt"]]
        alt = (idx % 2 == 1)
        for ci, val in enumerate(cells):
            c = ws.cell(row=row, column=1 + ci, value=val)
            _style_body_cell(c, alt=alt)
            c.alignment = A_CENTER if ci < 2 else A_RIGHT
        sum_manual += d["manual"]
        sum_manual_approved += d["manual_approved"]
        sum_eswt += d["eswt"]
        for c in manual_codes:
            sum_by_code[c] += d["by_code"][c]
        row += 1
    # 합계 행
    total_cells = ["합계", "", sum_manual, sum_manual_approved] + \
                  [sum_by_code[c] for c in manual_codes] + [sum_eswt]
    for ci, val in enumerate(total_cells):
        c = ws.cell(row=row, column=1 + ci, value=val)
        _style_body_cell(c, total=True)
        c.alignment = A_CENTER if ci < 2 else A_RIGHT
    row += 2

    # ─── 치료사별 도수치료 현황
    row += 1
    ws.cell(row=row, column=1, value="치료사별 도수치료 현황").font = F_SECTION
    row += 1
    ws.cell(row=row, column=1,
            value="실제 치료사 기준으로 완료 건수, 신환 수, 재진률, 예상매출을 비교합니다.").font = F_DESC
    row += 2

    th_headers = ["치료사"] + [manual_names[c] for c in manual_codes] + \
                 ["총 도수", "신환", "재진률", "예상매출"]
    _draw_header_row(ws, row, th_headers)
    row += 1

    if not th_rows_data:
        c = ws.cell(row=row, column=1, value="조회 기간 내 치료사별 완료 도수치료 데이터가 없습니다.")
        c.font = F_DESC; c.alignment = A_LEFT
        row += 1
    else:
        for idx, t in enumerate(th_rows_data):
            alt = (idx % 2 == 1)
            therapist_light = _lighten_hex_inner(t["color"], 0.88)
            therapist_fill = PatternFill("solid", fgColor=therapist_light)
            # 치료사
            c = ws.cell(row=row, column=1, value=t["name"])
            _style_body_cell(c, alt=alt); c.alignment = A_LEFT
            c.fill = therapist_fill  # 치료사 색상 아주 연하게 이름 칸에만
            # 항목별
            col = 2
            for code in manual_codes:
                cc = ws.cell(row=row, column=col, value=t["by_code"][code])
                _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT
                col += 1
            # 총 도수
            cc = ws.cell(row=row, column=col, value=t["total_manual"])
            _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT; col += 1
            # 신환
            cc = ws.cell(row=row, column=col, value=t["new_patient"])
            _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT; col += 1
            # 재진률
            cc = ws.cell(row=row, column=col, value=t["rate"] / 100.0)
            cc.number_format = "0.0%"
            _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT; col += 1
            # 예상매출
            cc = ws.cell(row=row, column=col)
            _style_body_cell(cc, alt=alt); _money(cc, t["revenue"])
            row += 1
    row += 2

    # ─── 비교표: 재진률이 높은 치료사
    row += 1
    ws.cell(row=row, column=1, value="재진률이 높은 치료사").font = F_SECTION
    row += 1
    ws.cell(row=row, column=1,
            value="완료된 도수치료 예약 중 기존 환자 비율이 높은 치료사입니다.").font = F_DESC
    row += 2
    _draw_header_row(ws, row, ["치료사", "재진률", "완료"])
    row += 1
    rate_sorted = sorted(th_rows_data, key=lambda x: (-x["rate"], -x["total_manual"]))
    if not rate_sorted:
        c = ws.cell(row=row, column=1, value="해당 데이터가 없습니다.")
        c.font = F_DESC; row += 1
    else:
        for idx, t in enumerate(rate_sorted):
            alt = (idx % 2 == 1)
            c1 = ws.cell(row=row, column=1, value=t["name"])
            _style_body_cell(c1, alt=alt); c1.alignment = A_LEFT
            c2 = ws.cell(row=row, column=2, value=t["rate"] / 100.0)
            c2.number_format = "0.0%"
            _style_body_cell(c2, alt=alt); c2.alignment = A_RIGHT
            c3 = ws.cell(row=row, column=3, value=t["total_manual"])
            _style_body_cell(c3, alt=alt); c3.alignment = A_RIGHT
            row += 1
    row += 2

    # ─── 비교표: 예상매출이 높은 치료사
    row += 1
    ws.cell(row=row, column=1, value="예상매출이 높은 치료사").font = F_SECTION
    row += 1
    ws.cell(row=row, column=1,
            value="치료항목별 수가를 기준으로 계산한 치료사별 예상매출입니다.").font = F_DESC
    row += 2
    _draw_header_row(ws, row, ["치료사", "예상매출", "총 도수"])
    row += 1
    rev_sorted = sorted(th_rows_data, key=lambda x: (-x["revenue"], -x["total_manual"]))
    if not rev_sorted:
        c = ws.cell(row=row, column=1, value="해당 데이터가 없습니다.")
        c.font = F_DESC; row += 1
    else:
        for idx, t in enumerate(rev_sorted):
            alt = (idx % 2 == 1)
            c1 = ws.cell(row=row, column=1, value=t["name"])
            _style_body_cell(c1, alt=alt); c1.alignment = A_LEFT
            c2 = ws.cell(row=row, column=2)
            _style_body_cell(c2, alt=alt); _money(c2, t["revenue"])
            c3 = ws.cell(row=row, column=3, value=t["total_manual"])
            _style_body_cell(c3, alt=alt); c3.alignment = A_RIGHT
            row += 1

    # 요약 시트 열 너비
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 8
    for i in range(3, 3 + len(manual_codes) + 5):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # ────────────────── 일별통계 시트 ──────────────────
    ws = ws_daily
    _setup_landscape(ws)
    ws.cell(row=1, column=1, value="일별 도수치료 통계").font = F_TITLE
    ws.cell(row=2, column=1, value=period_str).font = F_SUBTITLE
    out_cell = ws.cell(row=2, column=8, value=f"출력일시 {now_str}")
    out_cell.font = F_SUBTITLE; out_cell.alignment = A_RIGHT
    for col in range(1, 13):
        ws.cell(row=3, column=col).border = Border(bottom=Side(style="thin", color="1E3A5F"))

    row = 5
    daily2_headers = ["날짜", "요일", "도수 예약", "도수 완료"] + \
                     [manual_names[c] for c in manual_codes] + [eswt_name]
    _draw_header_row(ws, row, daily2_headers)
    row += 1
    for idx, dk in enumerate(date_keys):
        d = daily[dk]
        dt_obj = _dt.strptime(dk, "%Y-%m-%d")
        wd = weekday_kr[dt_obj.weekday()]
        cells = [dk, wd, d["manual"], d["manual_approved"]] + \
                [d["by_code"][c] for c in manual_codes] + [d["eswt"]]
        alt = (idx % 2 == 1)
        for ci, val in enumerate(cells):
            c = ws.cell(row=row, column=1 + ci, value=val)
            _style_body_cell(c, alt=alt)
            c.alignment = A_CENTER if ci < 2 else A_RIGHT
        row += 1
    # 합계
    total_cells = ["합계", "", sum_manual, sum_manual_approved] + \
                  [sum_by_code[c] for c in manual_codes] + [sum_eswt]
    for ci, val in enumerate(total_cells):
        c = ws.cell(row=row, column=1 + ci, value=val)
        _style_body_cell(c, total=True)
        c.alignment = A_CENTER if ci < 2 else A_RIGHT

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    for i in range(3, 5 + len(manual_codes) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # ────────────────── 치료사별집계 시트 ──────────────────
    ws = ws_th
    _setup_landscape(ws)
    ws.cell(row=1, column=1, value="치료사별 도수치료 집계").font = F_TITLE
    ws.cell(row=2, column=1, value=period_str).font = F_SUBTITLE
    out_cell = ws.cell(row=2, column=8, value=f"출력일시 {now_str}")
    out_cell.font = F_SUBTITLE; out_cell.alignment = A_RIGHT
    for col in range(1, 13):
        ws.cell(row=3, column=col).border = Border(bottom=Side(style="thin", color="1E3A5F"))

    row = 5
    th2_headers = ["치료사"] + [manual_names[c] for c in manual_codes] + \
                  ["총 도수", "신환", "재진률", "예상매출"]
    _draw_header_row(ws, row, th2_headers)
    row += 1
    if not th_rows_data:
        c = ws.cell(row=row, column=1, value="조회 기간 내 치료사별 완료 도수치료 데이터가 없습니다.")
        c.font = F_DESC; row += 1
    else:
        sum_by_code_th = {c: 0 for c in manual_codes}
        sum_total = sum_new = 0
        sum_revenue = 0
        sum_returning = 0
        sum_manual_done = 0
        for idx, t in enumerate(th_rows_data):
            alt = (idx % 2 == 1)
            therapist_light = _lighten_hex_inner(t["color"], 0.88)
            cell = ws.cell(row=row, column=1, value=t["name"])
            _style_body_cell(cell, alt=alt); cell.alignment = A_LEFT
            cell.fill = PatternFill("solid", fgColor=therapist_light)
            col = 2
            for code in manual_codes:
                v = t["by_code"][code]
                cc = ws.cell(row=row, column=col, value=v)
                _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT
                sum_by_code_th[code] += v
                col += 1
            cc = ws.cell(row=row, column=col, value=t["total_manual"])
            _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT; sum_total += t["total_manual"]; col += 1
            cc = ws.cell(row=row, column=col, value=t["new_patient"])
            _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT; sum_new += t["new_patient"]; col += 1
            cc = ws.cell(row=row, column=col, value=t["rate"] / 100.0)
            cc.number_format = "0.0%"
            _style_body_cell(cc, alt=alt); cc.alignment = A_RIGHT
            sum_returning += int(round(t["rate"] / 100.0 * t["total_manual"]))
            sum_manual_done += t["total_manual"]
            col += 1
            cc = ws.cell(row=row, column=col)
            _style_body_cell(cc, alt=alt); _money(cc, t["revenue"]); sum_revenue += t["revenue"]
            row += 1
        # 합계 행
        total_cells = ["합계"] + [sum_by_code_th[c] for c in manual_codes] + \
                      [sum_total, sum_new]
        for ci, val in enumerate(total_cells):
            c = ws.cell(row=row, column=1 + ci, value=val)
            _style_body_cell(c, total=True)
            c.alignment = A_CENTER if ci == 0 else A_RIGHT
        # 합계 행 재진률 = 전체 재진수 / 전체 완료 도수 (가중평균)
        avg_rate = (sum_returning / sum_manual_done) if sum_manual_done else 0
        col_rate = 1 + len(manual_codes) + 3
        cc = ws.cell(row=row, column=col_rate, value=avg_rate)
        cc.number_format = "0.0%"
        _style_body_cell(cc, total=True); cc.alignment = A_RIGHT
        # 합계 매출
        cc = ws.cell(row=row, column=col_rate + 1)
        _style_body_cell(cc, total=True); _money(cc, sum_revenue)

    ws.column_dimensions["A"].width = 14
    for i in range(2, 2 + len(manual_codes) + 4):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # ─── 응답
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"도수치료_통계_{d_from.isoformat()}_{d_to.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}",
        },
    )


def _lighten_hex_inner(hex_color: str, factor: float) -> str:
    """#RRGGBB -> 흰색쪽으로 factor (0~1) 만큼 블렌드. 0=원색 / 1=흰색."""
    h = (hex_color or "").lstrip("#")
    if len(h) != 6:
        return "FFFFFF"
    try:
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except Exception:
        return "FFFFFF"
    f = max(0.0, min(1.0, float(factor)))
    r = int(r + (255 - r) * f)
    g = int(g + (255 - g) * f)
    b = int(b + (255 - b) * f)
    return f"{r:02X}{g:02X}{b:02X}"