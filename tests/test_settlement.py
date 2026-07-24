from __future__ import annotations

import io
import sqlite3
import uuid
from datetime import date, timedelta

import openpyxl
import pytest

from app.migrations import m026_backfill_settlement_zero_snapshots
from app.modules.settlement.rules import calculate_incentive_amount, settlement_lock_before


def _unique(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


def _admin_headers(client) -> dict:
    resp = client.post("/api/admin/login", json={"password": "admin1234"})
    assert resp.status_code == 200, resp.text
    return {"X-Admin-Token": resp.json()["token"]}


def _make_category(client, name_prefix: str = "settle-cat") -> dict:
    payload = {
        "name": _unique(name_prefix),
        "color": "#2563EB",
        "active": True,
        "sort_order": 90,
        "default_can_doctor_treatment": False,
        "default_can_manual": True,
        "default_can_eswt": False,
    }
    resp = client.post("/api/employee-categories", json=payload)
    assert resp.status_code == 200, resp.text
    return resp.json()


def _make_treatment(client, headers: dict, category_id: str, *, price=80000,
                    incentive_pct=12.5, incentive_amount=None) -> dict:
    suffix = uuid.uuid4().hex[:8]
    payload = {
        "name": f"settlement-treatment-{suffix}",
        "short": f"S{suffix[:7]}",
        "category_id": category_id,
        "default_minutes": 30,
        "role": "therapist",
        "count_increment": 1,
        "show_in_patient": False,
        "active": True,
        "sort_order": 10,
        "code": f"settle_{suffix}",
        "price": price,
        "incentive_pct": incentive_pct,
        "incentive_amount": incentive_amount,
    }
    resp = client.post("/api/treatments", json=payload, headers=headers)
    assert resp.status_code == 200, resp.text
    return resp.json()


def _make_employee(client, category_id: str, treatment_ids: list[str]) -> dict:
    resp = client.post("/api/employees", json={
        "name": _unique("settlement-employee"),
        "category_id": category_id,
        "color": "#10B981",
        "active": True,
        "treatment_override_enabled": True,
        "treatment_ids": treatment_ids,
    })
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_calculate_incentive_amount():
    assert calculate_incentive_amount(50000, "fixed", 3000, 2) == 6000
    assert calculate_incentive_amount(80000, "percent", 12.5, 2) == 20000
    assert calculate_incentive_amount(33333, "percent", 10, 1) == 3333
    assert calculate_incentive_amount(80000, "none", 0, 4) == 0
    with pytest.raises(ValueError):
        calculate_incentive_amount(80000, "fixed", 1000, 0)


def test_settlement_grid_snapshot_report_update_and_delete(client):
    headers = _admin_headers(client)
    category = _make_category(client)
    treatment = _make_treatment(client, headers, category["id"], price=80000, incentive_pct=12.5)
    employee = _make_employee(client, category["id"], [treatment["id"]])

    payload = {
        "date_from": "2099-08-01",
        "date_to": "2099-08-01",
        "category_id": category["id"],
        "entries": [{
            "performed_on": "2099-08-01",
            "employee_id": employee["id"],
            "treatment_id": treatment["id"],
            "quantity": 2,
        }],
    }
    saved = client.post("/api/settlement/records/grid", json=payload, headers=headers)
    assert saved.status_code == 200, saved.text
    assert saved.json()["summary"]["incentive_total"] == 20000

    report = client.get(
        "/api/settlement/reports/incentives"
        f"?date_from=2099-08-01&date_to=2099-08-01&category_id={category['id']}",
        headers=headers,
    )
    assert report.status_code == 200, report.text
    body = report.json()
    assert body["summary"]["quantity_total"] == 2
    assert body["summary"]["price_total"] == 160000
    assert body["summary"]["incentive_total"] == 20000
    assert body["summary"]["payment_total"] == 20000
    rec = body["records"][0]
    assert rec["employee_name_snapshot"] == employee["name"]
    assert rec["employee_category_name_snapshot"] == category["name"]
    assert rec["treatment_name_snapshot"] == treatment["name"]
    assert rec["price_snapshot"] == 80000
    assert rec["incentive_type_snapshot"] == "percent"
    assert rec["incentive_value_snapshot"] == 12.5

    updated_treatment = dict(treatment)
    updated_treatment.update({
        "price": 100000,
        "incentive_pct": None,
        "incentive_amount": 30000,
    })
    changed = client.put(
        f"/api/treatments/{treatment['id']}",
        json=updated_treatment,
        headers=headers,
    )
    assert changed.status_code == 200, changed.text

    unchanged_report = client.get(
        "/api/settlement/reports/incentives"
        f"?date_from=2099-08-01&date_to=2099-08-01&category_id={category['id']}",
        headers=headers,
    )
    assert unchanged_report.status_code == 200, unchanged_report.text
    assert unchanged_report.json()["summary"]["incentive_total"] == 20000

    payload["entries"][0]["quantity"] = 3
    resaved = client.post("/api/settlement/records/grid", json=payload, headers=headers)
    assert resaved.status_code == 200, resaved.text
    assert resaved.json()["summary"]["quantity_total"] == 3
    assert resaved.json()["summary"]["price_total"] == 300000
    assert resaved.json()["summary"]["incentive_total"] == 90000

    payload["entries"][0]["quantity"] = 0
    deleted = client.post("/api/settlement/records/grid", json=payload, headers=headers)
    assert deleted.status_code == 200, deleted.text
    assert deleted.json()["summary"]["record_count"] == 0
    assert deleted.json()["summary"]["incentive_total"] == 0


def test_settlement_reflection_does_not_mirror_or_feed_aggregate(client):
    """v1.3.37+: 집계는 정산의 업스트림(치료완료/기록/수동)이고, 정산은 단방향
    다운스트림 스냅샷이다.
    - 정산 반영이 ManualCount 를 만들지 않는다 (미러 제거).
    - 집계는 정산 수량을 소스로 읽지 않는다 (예약/기록/수동 없으면 0).
    - manual_counts(수동 보정 델타)는 집계에 그대로 더해진다.
    """
    headers = _admin_headers(client)
    category = _make_category(client, "settle-agg-cat")
    treatment = _make_treatment(client, headers, category["id"], price=50000, incentive_amount=2500, incentive_pct=None)
    employee = _make_employee(client, category["id"], [treatment["id"]])

    payload = {
        "date_from": "2099-08-06",
        "date_to": "2099-08-06",
        "category_id": category["id"],
        "entries": [{
            "performed_on": "2099-08-06",
            "employee_id": employee["id"],
            "treatment_id": treatment["id"],
            "quantity": 5,
        }],
    }
    saved = client.post("/api/settlement/records/grid", json=payload, headers=headers)
    assert saved.status_code == 200, saved.text

    from app.database import SessionLocal
    from app.models import models

    db = SessionLocal()
    try:
        mirrored = db.query(models.ManualCount).filter(
            models.ManualCount.count_date == "2099-08-06",
            models.ManualCount.therapist_id == employee["id"],
            models.ManualCount.treatment_code == treatment["code"],
        ).count()
        assert mirrored == 0  # 미러 제거 — 정산 반영이 ManualCount 를 만들지 않는다.
    finally:
        db.close()

    aggregate = client.get(
        "/api/stats/direct-aggregate"
        f"?date_from=2099-08-06&date_to=2099-08-06&category_id={category['id']}"
    )
    assert aggregate.status_code == 200, aggregate.text
    cell = aggregate.json()["items"][0]["employee_data"][employee["id"]]
    assert cell["counts"][treatment["code"]] == 0   # 정산은 집계 소스가 아니다.
    assert cell["auto"][treatment["code"]] == 0

    # 수동 보정(manual_counts)은 집계에 그대로 더해진다.
    mc = client.post("/api/manual-counts", json={
        "date": "2099-08-06",
        "therapist_id": employee["id"],
        "treatment_code": treatment["code"],
        "count": 3,
    })
    assert mc.status_code == 200, mc.text
    cell2 = client.get(
        "/api/stats/direct-aggregate"
        f"?date_from=2099-08-06&date_to=2099-08-06&category_id={category['id']}"
    ).json()["items"][0]["employee_data"][employee["id"]]
    assert cell2["manual"][treatment["code"]] == 3
    assert cell2["counts"][treatment["code"]] == 3


def test_migration_backfills_zero_settlement_price_snapshots():
    conn = sqlite3.connect(":memory:")
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE treatments (
            id TEXT PRIMARY KEY,
            code TEXT,
            name TEXT,
            short TEXT,
            price INTEGER,
            incentive_pct REAL,
            incentive_amount INTEGER
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE settlement_records (
            id TEXT PRIMARY KEY,
            treatment_id TEXT,
            treatment_code TEXT,
            quantity INTEGER,
            treatment_name_snapshot TEXT,
            treatment_short_snapshot TEXT,
            treatment_code_snapshot TEXT,
            price_snapshot INTEGER,
            incentive_type_snapshot TEXT,
            incentive_value_snapshot REAL,
            incentive_amount INTEGER,
            updated_at TIMESTAMP
        )
        """
    )
    cur.execute(
        "INSERT INTO treatments VALUES (?, ?, ?, ?, ?, ?, ?)",
        ("tx1", "manual30", "도수치료30분", "도수6", 50000, 15.0, None),
    )
    cur.execute(
        "INSERT INTO settlement_records VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)",
        ("sr1", "tx1", "", 2, "", "", "", 0, "none", 0, 0),
    )
    conn.commit()

    m026_backfill_settlement_zero_snapshots.up(conn)

    row = cur.execute(
        """
        SELECT treatment_code, treatment_name_snapshot, treatment_short_snapshot,
               price_snapshot, incentive_type_snapshot, incentive_value_snapshot,
               incentive_amount
          FROM settlement_records
         WHERE id = 'sr1'
        """
    ).fetchone()
    assert row == ("manual30", "도수치료30분", "도수6", 50000, "percent", 15.0, 15000)
    conn.close()


def test_settlement_export_xlsx_period_category_and_no_code_column(client):
    headers = _admin_headers(client)
    category = _make_category(client, "settle-xlsx-cat")
    other_category = _make_category(client, "settle-xlsx-other")
    treatment = _make_treatment(client, headers, category["id"], price=70000, incentive_amount=7000, incentive_pct=None)
    other_treatment = _make_treatment(client, headers, other_category["id"], price=90000, incentive_amount=9000, incentive_pct=None)
    employee = _make_employee(client, category["id"], [treatment["id"]])
    other_employee = _make_employee(client, other_category["id"], [other_treatment["id"]])

    payload = {
        "date_from": "2099-08-04",
        "date_to": "2099-08-05",
        "category_id": category["id"],
        "entries": [{
            "performed_on": "2099-08-04",
            "employee_id": employee["id"],
            "treatment_id": treatment["id"],
            "quantity": 2,
        }],
    }
    saved = client.post("/api/settlement/records/grid", json=payload, headers=headers)
    assert saved.status_code == 200, saved.text

    other_payload = {
        "date_from": "2099-08-04",
        "date_to": "2099-08-05",
        "category_id": other_category["id"],
        "entries": [{
            "performed_on": "2099-08-04",
            "employee_id": other_employee["id"],
            "treatment_id": other_treatment["id"],
            "quantity": 5,
        }],
    }
    other_saved = client.post("/api/settlement/records/grid", json=other_payload, headers=headers)
    assert other_saved.status_code == 200, other_saved.text

    exported = client.get(
        "/api/settlement/reports/incentives.xlsx"
        f"?date_from=2099-08-04&date_to=2099-08-05&category_id={category['id']}",
        headers=headers,
    )
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)
    assert wb.sheetnames == ["직원별 정산 요약", "상세 내역", "치료항목별 합계"]

    summary_ws = wb["직원별 정산 요약"]
    assert [summary_ws.cell(row=4, column=i).value for i in range(1, 8)] == [
        "직원", "과", "총 건수", "총 수가", "세전 인센티브", "조정금액", "최종 지급액"
    ]
    assert summary_ws.cell(row=5, column=1).value == employee["name"]
    assert summary_ws.cell(row=5, column=2).value == category["name"]
    assert summary_ws.cell(row=5, column=3).value == 2
    assert summary_ws.cell(row=5, column=4).value == 140000
    assert summary_ws.cell(row=5, column=5).value == 14000
    assert summary_ws.cell(row=5, column=7).value == 14000

    detail_ws = wb["상세 내역"]
    detail_headers = [detail_ws.cell(row=4, column=i).value for i in range(1, 11)]
    assert "코드" not in detail_headers
    assert detail_headers == [
        "날짜", "과", "직원", "치료항목", "수량", "수가", "금액", "인센티브 규칙", "세전 인센티브", "메모"
    ]
    assert detail_ws.cell(row=5, column=2).value == category["name"]
    assert detail_ws.cell(row=5, column=3).value == employee["name"]
    assert detail_ws.cell(row=5, column=5).value == 2
    assert detail_ws.cell(row=5, column=9).value == 14000


def test_settlement_requires_admin(client):
    resp = client.get("/api/settlement/records?date_from=2099-08-01&date_to=2099-08-01")
    assert resp.status_code == 401


def test_settlement_rejects_unassigned_treatment(client):
    headers = _admin_headers(client)
    category = _make_category(client, "settle-deny-cat")
    allowed = _make_treatment(client, headers, category["id"], incentive_amount=1000, incentive_pct=None)
    denied = _make_treatment(client, headers, category["id"], incentive_amount=1000, incentive_pct=None)
    employee = _make_employee(client, category["id"], [allowed["id"]])

    resp = client.post("/api/settlement/records/grid", json={
        "date_from": "2099-08-02",
        "date_to": "2099-08-02",
        "category_id": category["id"],
        "entries": [{
            "performed_on": "2099-08-02",
            "employee_id": employee["id"],
            "treatment_id": denied["id"],
            "quantity": 1,
        }],
    }, headers=headers)
    assert resp.status_code == 400
    assert "assigned" in resp.text


def test_settlement_rejects_cross_category_employee_treatment_link(client):
    headers = _admin_headers(client)
    category = _make_category(client, "settle-employee-cat")
    other_category = _make_category(client, "settle-treatment-cat")
    allowed = _make_treatment(client, headers, category["id"], incentive_amount=1000, incentive_pct=None)
    other_treatment = _make_treatment(client, headers, other_category["id"], incentive_amount=1000, incentive_pct=None)
    employee = _make_employee(client, category["id"], [allowed["id"]])

    from app.database import SessionLocal
    from app.models import models

    db = SessionLocal()
    try:
        db.add(models.EmployeeTreatment(
            employee_id=employee["id"],
            treatment_id=other_treatment["id"],
        ))
        db.commit()
    finally:
        db.close()

    resp = client.post("/api/settlement/records/grid", json={
        "date_from": "2099-08-06",
        "date_to": "2099-08-06",
        "category_id": other_category["id"],
        "entries": [{
            "performed_on": "2099-08-06",
            "employee_id": employee["id"],
            "treatment_id": other_treatment["id"],
            "quantity": 1,
        }],
    }, headers=headers)
    assert resp.status_code == 400
    assert "assigned" in resp.text


def test_settlement_selected_category_without_treatments_is_empty(client):
    headers = _admin_headers(client)
    empty_category = _make_category(client, "settle-empty-cat")
    other_category = _make_category(client, "settle-other-cat")
    other_treatment = _make_treatment(client, headers, other_category["id"], incentive_amount=1000, incentive_pct=None)

    empty_employee = client.post("/api/employees", json={
        "name": _unique("settle-empty-category-employee"),
        "category_id": empty_category["id"],
        "color": "#22C55E",
        "active": True,
        "treatment_override_enabled": False,
    })
    assert empty_employee.status_code == 200, empty_employee.text
    _make_employee(client, other_category["id"], [other_treatment["id"]])

    resp = client.get(
        "/api/settlement/records"
        f"?date_from=2099-08-03&date_to=2099-08-03&category_id={empty_category['id']}",
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["category_id"] == empty_category["id"]
    assert data["treatments"] == []
    assert data["employees"] == []
    assert data["items"] == [{"date": "2099-08-03", "employee_data": {}}]
    assert data["summary"]["record_count"] == 0


# ──────────── 정산 확정(잠금) — 매월 1일 기준 2달 전 자동 확정 ────────────


def test_settlement_lock_before_boundary():
    """확정 경계 = (이번 달 1일) - 1개월. 배치 없이 날짜만으로 자동 확정된다."""
    # 7/16 → 경계 6/1 : 5월 이하 확정, 6월(1달 전)·7월은 수정 가능
    assert settlement_lock_before(date(2026, 7, 16)) == date(2026, 6, 1)
    # 달이 바뀌는 순간(8/1) 6월이 자동 확정 → 경계 7/1
    assert settlement_lock_before(date(2026, 7, 31)) == date(2026, 6, 1)
    assert settlement_lock_before(date(2026, 8, 1)) == date(2026, 7, 1)
    # 연초 넘김
    assert settlement_lock_before(date(2026, 1, 5)) == date(2025, 12, 1)
    assert settlement_lock_before(date(2026, 2, 28)) == date(2026, 1, 1)


def test_settlement_locked_period_is_not_recalculated(client):
    """확정 기간의 기존 스냅샷은 수량·금액이 다시 계산되지 않는다(급여 근거 보존).

    단, 스냅샷이 아예 없으면 최초 1회는 생성한다 — 한 번도 조회하지 않은 과거 달이
    확정됐다는 이유로 통째로 비어 보이는 것을 막기 위함.
    """
    headers = _admin_headers(client)
    category = _make_category(client)
    treatment = _make_treatment(client, headers, category["id"], price=80000, incentive_pct=10)
    employee = _make_employee(client, category["id"], [treatment["id"]])

    lock = settlement_lock_before()
    locked_day = (lock - timedelta(days=1)).isoformat()  # 확정 기간 (2달 전 이하)
    open_day = lock.isoformat()                          # 경계 당일 = 1달 전 → 수정 가능

    def post(day: str, qty: int):
        resp = client.post("/api/settlement/records/grid", json={
            "date_from": day,
            "date_to": day,
            "category_id": category["id"],
            "entries": [{
                "performed_on": day,
                "employee_id": employee["id"],
                "treatment_id": treatment["id"],
                "quantity": qty,
            }],
        }, headers=headers)
        assert resp.status_code == 200, resp.text
        return resp.json()

    # 1) 확정 기간이라도 스냅샷이 없으면 최초 1회 생성
    d = post(locked_day, 2)
    assert d["changed"]["upserted"] == 1
    assert d["summary"]["quantity_total"] == 2

    # 2) 확정 기간의 기존 스냅샷은 수량이 바뀌지 않음
    d = post(locked_day, 99)
    assert d["changed"]["locked_skipped"] == 1
    assert d["changed"]["upserted"] == 0
    assert d["summary"]["quantity_total"] == 2

    # 3) 확정 기간은 수량 0(삭제)도 무시
    d = post(locked_day, 0)
    assert d["changed"]["deleted"] == 0
    assert d["changed"]["locked_skipped"] == 1
    assert d["summary"]["quantity_total"] == 2

    # 4) 경계 당일(1달 전)은 계속 갱신 가능
    d = post(open_day, 3)
    assert d["changed"]["upserted"] == 1
    assert d["summary"]["quantity_total"] == 3
    d = post(open_day, 5)
    assert d["changed"]["upserted"] == 1
    assert d["summary"]["quantity_total"] == 5


def test_settlement_report_exposes_lock_info(client):
    """정산 리포트가 확정 경계/상태를 내려준다 (화면 안내용, 판정은 백엔드가 단일 원천)."""
    headers = _admin_headers(client)
    lock = settlement_lock_before()
    locked_day = (lock - timedelta(days=1)).isoformat()
    open_day = lock.isoformat()

    def report(date_from: str, date_to: str):
        resp = client.get(
            "/api/settlement/reports/incentives"
            f"?date_from={date_from}&date_to={date_to}",
            headers=headers,
        )
        assert resp.status_code == 200, resp.text
        return resp.json()

    d = report(locked_day, locked_day)
    assert d["lock_before"] == lock.isoformat()
    assert d["locked"] is True
    assert d["partially_locked"] is False

    d = report(open_day, open_day)
    assert d["locked"] is False
    assert d["partially_locked"] is False

    # 확정 기간 ~ 미확정 기간에 걸친 조회
    d = report(locked_day, open_day)
    assert d["locked"] is False
    assert d["partially_locked"] is True
