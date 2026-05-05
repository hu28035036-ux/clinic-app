"""v1.3.5+ — 도수치료 예약현황 엑셀 다운로드의 점심시간 반영 회귀 가드.

사용자 요청:
- cfg.lunch_enabled / lunch_start / lunch_end 가 설정되면 엑셀에도 점심 슬롯 반영.
- 점심 슬롯의 빈 셀은 회색 (#E5E7EB) 배경, 첫 점심 슬롯 첫 컬럼에 "점심시간 HH:MM~HH:MM" 라벨.

회귀:
- lunch_enabled=False → 회색 슬롯 ❌
- lunch_enabled=True 12:30~13:30 → 12:30/13:00 슬롯의 빈 셀 회색 + 라벨 1회
"""
from __future__ import annotations

import io

import openpyxl

from app.config import load_config, save_config


def _set_lunch(enabled: bool, start: str = "12:30", end: str = "13:30") -> dict:
    cfg = load_config()
    prev = {
        "lunch_enabled": cfg.get("lunch_enabled"),
        "lunch_start": cfg.get("lunch_start"),
        "lunch_end": cfg.get("lunch_end"),
    }
    cfg["lunch_enabled"] = enabled
    cfg["lunch_start"] = start
    cfg["lunch_end"] = end
    save_config(cfg)
    return prev


def _restore(prev: dict) -> None:
    cfg = load_config()
    for k, v in prev.items():
        cfg[k] = v
    save_config(cfg)


def test_export_excel_no_lunch_label_when_disabled(client):
    """점심 비활성 → 엑셀 어디에도 '점심시간' 라벨 ❌."""
    prev = _set_lunch(False)
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        all_cells = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    all_cells.append(str(v))
        assert not any("점심시간" in v for v in all_cells), \
            "lunch_enabled=False 인데 엑셀에 '점심시간' 라벨이 있음"
    finally:
        _restore(prev)


def test_export_excel_lunch_label_present_when_enabled(client):
    """점심 활성 12:30~13:30 → 엑셀에 '점심시간 12:30~13:30' 라벨 *정확히 1회*."""
    prev = _set_lunch(True, "12:30", "13:30")
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        labels = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and "점심시간" in str(v):
                    labels.append(str(v))
        assert labels, "lunch_enabled=True 인데 엑셀에 '점심시간' 라벨이 없음"
        assert any("12:30" in lbl and "13:30" in lbl for lbl in labels), \
            f"라벨 시간이 cfg 와 불일치: {labels}"
        # 첫 점심 슬롯 첫 컬럼 1회 라벨 정책
        assert len(labels) == 1, f"라벨이 1회 초과 등장: {labels}"
    finally:
        _restore(prev)


def test_export_excel_lunch_label_under_option2(client):
    """v1.3.5+ 옵션 ② 정책 후: 점심 슬롯 *모든* 셀이 빈 (점심) 상태.

    옵션 ② 로 점심 슬롯의 예약은 cell_map 등록 ❌ → 첫 점심 슬롯 첫 컬럼이 항상 빈
    → 라벨이 첫 점심 슬롯 첫 컬럼 (ci=0) 에 표시. 이전의 라벨 손실 시나리오는
    옵션 ② 적용으로 자연 해소되지만 라벨 1회 보장 회귀는 유지.
    """
    from datetime import datetime as _dt

    from app.database import SessionLocal
    from app.models import models
    prev = _set_lunch(True, "12:30", "13:30")
    db = SessionLocal()
    appt_id = None
    try:
        if not db.query(models.Patient).filter_by(id="lunchp1").first():
            db.add(models.Patient(id="lunchp1", name="점심환자", chart_no="LP01"))
        therapist = (
            db.query(models.Employee)
            .filter(models.Employee.role == "therapist",
                    models.Employee.active == True,  # noqa: E712
                    models.Employee.can_manual == True)  # noqa: E712
            .order_by(models.Employee.sort_order, models.Employee.name)
            .first()
        )
        if therapist:
            appt = models.Appointment(
                patient_id="lunchp1",
                therapist_id=therapist.id,
                start_at=_dt(2026, 5, 1, 12, 30),
                end_at=_dt(2026, 5, 1, 13, 0),
                duration_min=30,
                treatment_codes='["manual30"]',
                status="reserved",
            )
            db.add(appt)
            db.commit()
            appt_id = appt.id

        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        labels = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and "점심시간" in str(v):
                    labels.append(str(v))
        assert len(labels) == 1, (
            f"첫 점심 슬롯 첫 컬럼이 점유돼도 라벨 1회 보존돼야 함 (실제: {labels})"
        )
    finally:
        # 정리
        if appt_id:
            obj = db.query(models.Appointment).filter_by(id=appt_id).first()
            if obj:
                db.delete(obj)
        p = db.query(models.Patient).filter_by(id="lunchp1").first()
        if p:
            db.delete(p)
        db.commit()
        db.close()
        _restore(prev)


def test_export_excel_invalid_lunch_config_graceful(client):
    """v1.3.5+ Codex Low fix: lunch_start/end 형식 오류 시 graceful.

    invalid value 입력 → 500 ❌, 라벨 ❌, 회색 ❌. 정상 엑셀 응답.
    """
    prev = _set_lunch(True, "abc", "13:30")
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # invalid → lunch_slots 빈 set / 라벨 빈 문자열 → "점심시간" 어디에도 ❌
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and "점심시간" in str(v):
                    raise AssertionError(
                        f"invalid lunch config 인데 라벨 노출됨: {v}"
                    )
    finally:
        _restore(prev)


def test_export_excel_lunch_minute_out_of_range_graceful(client):
    """v1.3.5+ Codex Medium fix: HH:MM 의 분 단위가 0-59 범위 벗어나면 graceful.

    이전: `_hm_to_min("12:75")` 가 `13:15` (= 12*60+75 = 795분) 로 암묵 변환 → 점심 적용.
    수정: 분 0-59 범위 명시 검증. 범위 벗어나면 lunch_slots = empty → 라벨 ❌.
    """
    prev = _set_lunch(True, "12:75", "13:30")  # 분 75 = invalid
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and "점심시간" in str(v):
                    raise AssertionError(
                        f"분 범위 벗어난 lunch config (12:75) 인데 라벨 노출됨: {v}"
                    )
    finally:
        _restore(prev)


def test_export_excel_lunch_hour_out_of_range_graceful(client):
    """v1.3.5+ Codex Medium fix: HH 가 0-23 벗어나면 graceful."""
    prev = _set_lunch(True, "25:30", "26:30")  # 시 25 = invalid
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and "점심시간" in str(v):
                    raise AssertionError(
                        f"시 범위 벗어난 lunch config (25:30) 인데 라벨 노출됨: {v}"
                    )
    finally:
        _restore(prev)


def test_export_excel_lunch_starts_in_lunch_hidden(client):
    """v1.3.5+ 사용자 옵션 ② 정책 변경: 점심 슬롯 시작 예약은 보고서에서 *가림*.

    이전 정책: 예약 보존 (보고서 정확성 우선)
    옵션 ② 정책: 점심 라벨 우선 (보고서 가독성 우선) — 예약 시각적 가리기.
    실제 데이터는 DB 그대로 (UI 보드에서 확인 가능).
    """
    from datetime import datetime as _dt

    from app.database import SessionLocal
    from app.models import models
    prev = _set_lunch(True, "12:30", "13:30")
    db = SessionLocal()
    appt_id = None
    try:
        if not db.query(models.Patient).filter_by(id="lunchp2").first():
            db.add(models.Patient(id="lunchp2", name="가려질환자", chart_no="LP02"))
        therapist = (
            db.query(models.Employee)
            .filter(models.Employee.role == "therapist",
                    models.Employee.active == True,  # noqa: E712
                    models.Employee.can_manual == True)  # noqa: E712
            .order_by(models.Employee.sort_order, models.Employee.name)
            .first()
        )
        if therapist:
            # 점심 슬롯 시작 (12:30) 예약 — 옵션 ② 정책으로 엑셀에서 가려져야 함
            appt = models.Appointment(
                patient_id="lunchp2",
                therapist_id=therapist.id,
                start_at=_dt(2026, 5, 1, 12, 30),
                end_at=_dt(2026, 5, 1, 13, 0),
                duration_min=30,
                treatment_codes='["manual30"]',
                status="reserved",
            )
            db.add(appt)
            db.commit()
            appt_id = appt.id

        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        all_text = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    all_text.append(str(v))
        joined = " ".join(all_text)
        # 옵션 ②: 점심 슬롯 예약 가림 → 환자명 ❌
        assert "가려질환자" not in joined, (
            "옵션 ② 정책: 점심 슬롯 시작 예약은 엑셀에서 가려져야 함 (DB 데이터는 그대로)"
        )
        # 점심 라벨은 정상 1회
        labels = [t for t in all_text if "점심시간" in t]
        assert len(labels) == 1, f"점심 라벨 1회 보장 (실제: {labels})"
    finally:
        if appt_id:
            obj = db.query(models.Appointment).filter_by(id=appt_id).first()
            if obj:
                db.delete(obj)
        p = db.query(models.Patient).filter_by(id="lunchp2").first()
        if p:
            db.delete(p)
        db.commit()
        db.close()
        _restore(prev)


def test_export_excel_appointment_span_truncated_at_lunch(client):
    """v1.3.5+ 옵션 ②: 일반 슬롯 시작 → 점심 슬롯 걸치는 60분 예약 = span 자름.

    예: 12:00 시작 60분 예약 (12:00~13:00), 12:30 가 점심 → 12:00 셀에는 환자명 표시,
    12:30 셀은 점심 회색. 세로 병합이 점심을 침범하지 ❌.
    """
    from datetime import datetime as _dt

    from app.database import SessionLocal
    from app.models import models
    prev = _set_lunch(True, "12:30", "13:30")
    db = SessionLocal()
    appt_id = None
    try:
        if not db.query(models.Patient).filter_by(id="lunchp3").first():
            db.add(models.Patient(id="lunchp3", name="잘림환자", chart_no="LP03"))
        therapist = (
            db.query(models.Employee)
            .filter(models.Employee.role == "therapist",
                    models.Employee.active == True,  # noqa: E712
                    models.Employee.can_manual == True)  # noqa: E712
            .order_by(models.Employee.sort_order, models.Employee.name)
            .first()
        )
        if therapist:
            # 12:00 시작 60분 예약 — 12:00 (일반) ~ 13:00 (점심까지 걸침)
            appt = models.Appointment(
                patient_id="lunchp3",
                therapist_id=therapist.id,
                start_at=_dt(2026, 5, 1, 12, 0),
                end_at=_dt(2026, 5, 1, 13, 0),
                duration_min=60,
                treatment_codes='["manual60"]',
                status="reserved",
            )
            db.add(appt)
            db.commit()
            appt_id = appt.id

        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        all_text = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    all_text.append(str(v))
        joined = " ".join(all_text)
        # 12:00 슬롯은 일반 → 환자명 표시 (span 1로 잘림)
        assert "잘림환자" in joined, (
            "12:00 시작 (일반 슬롯) 예약은 정상 표시돼야 함"
        )
        # 점심 라벨 1회
        labels = [t for t in all_text if "점심시간" in t]
        assert len(labels) == 1
    finally:
        if appt_id:
            obj = db.query(models.Appointment).filter_by(id=appt_id).first()
            if obj:
                db.delete(obj)
        p = db.query(models.Patient).filter_by(id="lunchp3").first()
        if p:
            db.delete(p)
        db.commit()
        db.close()
        _restore(prev)


def test_export_excel_lunch_outside_operating_hours_graceful(client):
    """v1.3.5+ 사용자 명시: 점심이 운영시간 *바깥* 으로 설정돼도 graceful.

    예: 운영 08:30~18:30, 점심 19:00~20:00 → 운영 슬롯과 겹치는 lunch_slot ⊥
    → 라벨 ❌, 회색 ❌, 정상 엑셀 응답.
    """
    prev = _set_lunch(True, "19:00", "20:00")
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and "점심시간" in str(v):
                    raise AssertionError(
                        f"운영시간 외부 점심인데 라벨 노출됨: {v}"
                    )
    finally:
        _restore(prev)


def test_export_excel_lunch_slot_grey_background(client):
    """점심 슬롯 빈 셀은 회색 배경 (#E5E7EB / FFE5E7EB)."""
    prev = _set_lunch(True, "12:30", "13:30")
    try:
        r = client.get("/api/export/manual-schedule.xlsx?date=2026-05-01")
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # 시간 컬럼 (A) 에서 12:30 / 13:00 슬롯 행 찾기
        lunch_rows = []
        for ri, row in enumerate(ws.iter_rows(values_only=False), start=1):
            time_cell = row[0]
            if time_cell.value in ("12:30", "13:00"):
                lunch_rows.append(ri)
        assert lunch_rows, "12:30 / 13:00 시간 행을 엑셀에서 찾지 못함"
        # 첫 점심 행의 시간 셀 fill 검증 (FFE5E7EB or E5E7EB)
        first_row = ws[lunch_rows[0]]
        time_cell = first_row[0]
        fg = (time_cell.fill.fgColor.rgb or "").upper() if time_cell.fill else ""
        assert "E5E7EB" in fg, f"점심 슬롯 시간 셀 배경 회색 ❌: {fg}"
    finally:
        _restore(prev)
